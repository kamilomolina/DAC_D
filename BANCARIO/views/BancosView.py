import base64
import csv
import hashlib
import json
import logging
import os
import pandas as pd
import pyodbc
import pymssql
import pymysql
import requests
import subprocess
import tempfile
import time
import xlrd
import jwt
from decimal import Decimal
import gzip
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
import decimal
from django.db.utils import OperationalError
from django.views.decorators.http import require_POST
from calendar import monthrange
from urllib.parse import quote_plus
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.db import connections, transaction
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, Http404
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from sqlalchemy import create_engine, text
from sqlalchemy.dialects.mysql import insert
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import sessionmaker
from CWS.views.LoginView import EXPIRATION_MINUTES, SECRET_KEY

from CONTABLE.views.utilsContable import *
from CONTABLE.views.utilsContable import _callproc_fetchall
from BANCARIO.views.utilsBancario import *


def setEnlaceMovimiento(request):
    id_mov = request.POST.get("id_mov")
    id_enlace = request.POST.get("id_enlace")
    tipo = request.POST.get("tipo")
    userName = request.session.get("userName", "")

    try:
        with connections["bankConn"].cursor() as cursor:
            cursor.callproc(
                "CONTA_SET_ENLACE_MOVIMIENTO", [id_mov, id_enlace, tipo, userName]
            )
        return JsonResponse({"save": 1})
    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)})


def dataDestinosProcedenciasMovimientos(request):
    empresa = request.POST.get("empresa")
    tipo_enlance = request.POST.get("tipo_enlance")
    estado = request.POST.get("estado")

    user_name = request.session.get("userName", "CONTABLE")

    data = []

    try:
        udcConn = connections["bankConn"]
        with udcConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_DESTINOS__PROCEDENCIAS_SIN_ENLAZAR",
                [tipo_enlance, estado, empresa],
            )
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        udcConn.close()

        return JsonResponse({"data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataTiposEnlancesDestinosProcedencias(request):
    tipo_enlance = request.POST.get("tipo_enlance")
    estado = request.POST.get("estado")

    user_name = request.session.get("userName", "CONTABLE")

    data = []

    try:
        udcConn = connections["bankConn"]
        with udcConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_ENLANCES_DESTINOS_PROCEDENCIAS", [tipo_enlance, estado]
            )
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        udcConn.close()

        return JsonResponse({"data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def api_get_caja_salidas_pendientes(request):
    sucursal = request.GET.get("sucursal", 0)
    try:
        with connections["super"].cursor() as cursor:
            cursor.callproc("CONTA_GET_CAJA_SALIDAS_PENDIENTES", [sucursal])
            rows = dictfetchall(cursor)

            # Formatear la salida serializable
            for row in rows:
                if "Salidas" in row and row["Salidas"]:
                    row["Salidas"] = float(row["Salidas"])
                if "SaldoDisponible" in row and row["SaldoDisponible"]:
                    row["SaldoDisponible"] = float(row["SaldoDisponible"])
                if "Creado" in row and row["Creado"]:
                    row["Creado"] = row["Creado"].strftime("%Y-%m-%d %H:%M:%S")

            return JsonResponse({"data": rows})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def bank_enlazar_destinos_procedencias(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("bancarioAdminIT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        context = {
            "empresasData": empresasData,
            "adminIT": adminIT,
        }

        return render(request, "bancario/enlazar_destinos_procedencias.html", context)


def saveEnlaceDestinoProcedencia(request):

    id_enlace = int(request.POST.get("id_enlace", 0))
    tipo = int(request.POST.get("tipo"))
    nombre = request.POST.get("nombre")
    cuenta = request.POST.get("cuenta")

    userName = request.session.get("userName")

    with connections["bankConn"].cursor() as cursor:
        cursor.callproc(
            "CONTA_SAVE_ENLACE_DESTINO_PROCEDENCIA",
            [id_enlace, tipo, nombre, cuenta, userName],
        )
        result = dictfetchall(cursor)

    return JsonResponse({"save": 1})


def toggleEnlaceDestinoProcedencia(request):

    id_enlace = int(request.POST.get("id_enlace"))
    userName = request.session.get("userName")

    with connections["bankConn"].cursor() as cursor:
        cursor.callproc(
            "CONTA_TOGGLE_ENLACE_DESTINO_PROCEDENCIA", [id_enlace, userName]
        )

    return JsonResponse({"save": 1})


def dataEnlacesAdmin(request):

    with connections["bankConn"].cursor() as cursor:
        cursor.callproc("CONTA_GET_ENLACES_ADMIN")
        result = dictfetchall(cursor)

    return JsonResponse({"data": result})


def panel_bancario(request):
    user_id = request.session.get("user_id", "")
    if user_id == "":
        return HttpResponseRedirect(reverse("login"))

    get_accesos_bancario(request)

    error_msg = request.session.get("error_msg", None)
    if error_msg:
        del request.session["error_msg"]

    context = {"error_msg": error_msg}
    return render(request, "panel_bancario.html", context)


def api_get_ingresos_pendientes(request):

    modo = int(request.GET.get("modo", 1))
    proveedor_id = int(request.GET.get("proveedor_id", 0))
    empresa_id = int(request.GET.get("empresa_id", 0))
    valor_maximo = request.GET.get("valor_maximo", 0)

    with connections["bankConn"].cursor() as cursor:
        cursor.callproc(
            "BK_GET_INGRESOS_PENDIENTES", [modo, proveedor_id, empresa_id, valor_maximo]
        )
        columns = [col[0] for col in cursor.description]
        data = [dict(zip(columns, row)) for row in cursor.fetchall()]

    return JsonResponse({"data": data})


def api_get_solicitudes_acreedor(request):

    cod_proveedor = int(request.GET.get("cod_proveedor", 0))
    empresa_id = int(request.GET.get("empresa_id", 0))

    with connections["bankConn"].cursor() as cursor:
        cursor.callproc("BK_GET_SOLICITUD_ACREEDORES", [cod_proveedor, empresa_id])
        columns = [col[0] for col in cursor.description]
        data = [dict(zip(columns, row)) for row in cursor.fetchall()]

    return JsonResponse({"data": data})


def bank_liquidacion_caja_turnos(request):

    user = request.session.get("user_id", "")
    if user == "":
        return redirect("login")

    acceso_denegado = verificar_acceso_bancario(request, "40111")
    if acceso_denegado:
        return acceso_denegado

    empresa = request.GET.get("empresa", "3")
    adminIT = request.session.get("bancarioAdminIT", 0)
    fecha_actual = datetime.today()

    context = {
        "fecha_actual": fecha_actual,
        "id_empresa": empresa,
        "adminIT": adminIT,
        "empresasData": obtener_empresas_usuario(request),
        "sucursalesData": obtener_sucursales_usuario(request, empresa),
        "date1": datetime.now().replace(day=1).strftime("%Y-%m-%d"),
        "date2": getActualDate(),
    }

    return render(request, "caja/reporte_liquidaciones.html", context)


def getDataTurnosLiquidar(request):

    sucursal = request.GET.get("sucursal")
    date1 = request.GET.get("date1")
    date2 = request.GET.get("date2")
    tipo = request.GET.get("tipo", 1)

    data = []

    with connections["super"].cursor() as cursor:
        cursor.callproc("CONTA_LIQ_GET_TURNOS", [sucursal, date1, date2, tipo])
        column_names = [desc[0] for desc in cursor.description]
        raw_rows = cursor.fetchall()
        data = []
        for row in raw_rows:
            row_dict = {}
            for col, val in zip(column_names, row):
                # Convertir campos tipo BIT (bytes) a int (0 o 1)
                if isinstance(val, bytes):
                    row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                else:
                    row_dict[str(col)] = val
            data.append(row_dict)

    return JsonResponse({"data": data})


def getDataLiquidacionMovDiario(request):

    sucursal = request.GET.get("sucursal")
    date1 = request.GET.get("date1")
    date2 = request.GET.get("date2")
    tipo = request.GET.get("tipo", 1)

    data = []

    with connections["bankConn"].cursor() as cursor:
        cursor.callproc("CONTA_GET_LIQ_CAJA_MOV_DIARIO", [date1, date2, sucursal, tipo])
        column_names = [desc[0] for desc in cursor.description]
        raw_rows = cursor.fetchall()
        data = []
        for row in raw_rows:
            row_dict = {}
            for col, val in zip(column_names, row):
                # Convertir campos tipo BIT (bytes) a int (0 o 1)
                if isinstance(val, bytes):
                    row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                else:
                    row_dict[str(col)] = val
            data.append(row_dict)

    return JsonResponse({"data": data})


def actualizar_turno_liquidacion(request):

    pkturno = request.POST.get("pkturno")
    tipo = request.POST.get("tipo")
    userName = request.session.get("userName", "")

    data = {}

    with connections["super"].cursor() as cursor:
        cursor.callproc("CONTA_LIQ_ACTUALIZAR_TURNO", [pkturno, tipo, userName])
        row = cursor.fetchone()
        columns = [col[0] for col in cursor.description]
        data = dict(zip(columns, row))

    return JsonResponse(data)


def listado_anticipo_clientes(request):

    user = request.session.get("user_id", "")
    if user == "":
        return redirect("login")

    acceso_denegado = verificar_acceso_bancario(request, "40113")
    if acceso_denegado:
        return acceso_denegado

    empresa = request.GET.get("empresa", "3")
    adminIT = request.session.get("bancarioAdminIT", 0)
    fecha_actual = datetime.today()

    context = {
        "fecha_actual": fecha_actual,
        "id_empresa": empresa,
        "adminIT": adminIT,
        "empresasData": obtener_empresas_usuario(request),
        "bancosData": get_bancos(),
        "estadosData": get_estados_movimientos(),
        "clientesData": get_clientes(),
        "date1": getDate1(),
        "date2": getActualDate(),
    }

    return render(request, "bancario/anticipos_clientes_listado.html", context)


def bk_get_lote_abonos_listado(request):
    if request.method == "GET":

        idEmpresa = request.GET.get("idEmpresa", "0")
        idBanco = request.GET.get("idBanco", "0")
        idCuenta = request.GET.get("idCuenta", "0")
        desde = request.GET.get("desde")
        hasta = request.GET.get("hasta")
        estado = request.GET.get("estado", "0")  # 0/1/2

        data = []

        try:
            with connections["bankConn"].cursor() as cursor:
                cursor.callproc(
                    "BK_GET_CXC_LOTE_ABONOS_LISTADO",
                    [
                        int(idEmpresa),
                        int(idBanco),
                        int(idCuenta),
                        desde,
                        hasta,
                        int(estado),
                    ],
                )

                rows = cursor.fetchall()
                cols = [c[0] for c in cursor.description]

                for r in rows:
                    item = dict(zip(cols, r))
                    data.append(item)

            return JsonResponse({"success": True, "data": data})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e), "data": []})


def bk_autorizar_abonos_detalle(request):
    if request.method == "POST":

        userName = request.session.get("userName", "")

        ids = request.POST.getlist("ids[]")  # viene como lista
        ids = [str(int(x)) for x in ids if str(x).isdigit()]

        if len(ids) == 0:
            return JsonResponse(
                {"success": False, "error": "No hay registros seleccionados."}
            )

        ids_csv = ",".join(ids)

        try:
            with connections["bankConn"].cursor() as cursor:
                cursor.callproc("BK_AUTORIZAR_ABONOS_DETALLE", [ids_csv, userName])
                row = cursor.fetchone()
                cols = [c[0] for c in cursor.description]
                resp = dict(zip(cols, row))

            return JsonResponse(
                {
                    "success": True,
                    "codigo": resp.get("CodigoAutorizacion"),
                    "rows_updated": resp.get("rows_updated", 0),
                }
            )

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})


def bk_get_abonos_por_autorizacion(request):
    if request.method == "GET":

        codigo = request.GET.get("codigo", "").strip()
        if codigo == "":
            return JsonResponse({"success": False, "error": "Debe indicar codigo."})

        data = []
        try:
            with connections["bankConn"].cursor() as cursor:
                cursor.callproc("BK_GET_ABONOS_POR_AUTORIZACION", [codigo])
                rows = cursor.fetchall()
                cols = [c[0] for c in cursor.description]
                for r in rows:
                    data.append(dict(zip(cols, r)))

            return JsonResponse({"success": True, "data": data})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e), "data": []})


def bk_get_lotes_autorizados(request):

    idEmpresa = request.GET.get("idEmpresa", 0)
    desde = request.GET.get("desde")
    hasta = request.GET.get("hasta")

    try:
        with connections["bankConn"].cursor() as cursor:
            cursor.callproc("BK_GET_LOTES_AUTORIZADOS", [int(idEmpresa), desde, hasta])
            data = dictfetchall(cursor)

        return JsonResponse({"success": True, "data": data})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


def bk_get_detalle_lote_aut(request):

    codigo = request.GET.get("codigo")

    try:
        with connections["bankConn"].cursor() as cursor:
            cursor.callproc("BK_GET_DETALLE_LOTE_AUT", [codigo])
            data = dictfetchall(cursor)

        return JsonResponse({"success": True, "data": data})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


def bk_descargar_anticipos_clientes_autorizados(request):

    lotes = request.GET.get("lotes", "")

    if not lotes:
        return JsonResponse({"success": False, "error": "No se recibieron lotes"})

    try:
        with connections["bankConn"].cursor() as cursor:
            cursor.callproc("BK_GET_ABONOS_MULTIPLE_AUT", [lotes])
            data = dictfetchall(cursor)

        if not data:
            return JsonResponse({"success": False, "error": "No hay datos"})

        wb = Workbook()
        ws = wb.active
        ws.title = "Anticipos"

        headers = [
            "Código AUT",
            "Empresa",
            "Banco",
            "Cuenta",
            "Referencia",
            "Fecha",
            "Cliente",
            "Documento",
            "Fecha Documento",
            "Saldo",
            "Abono",
        ]

        ws.append(headers)

        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)

        for row in data:
            ws.append(
                [
                    row["CodigoAutorizacion"],
                    row["EmpresaNombre"],
                    row["BancoNombre"],
                    row["CuentaBancaria"],
                    row["NumeroReferenciaBanco"],
                    str(row["Fecha"]),
                    row["NombreCliente"],
                    row["NoDocumento"],
                    str(row["FechaDocumento"]),
                    float(row["SaldoDocumento"]),
                    float(row["ValorAbono"]),
                ]
            )

        # Formato moneda columnas 10 y 11
        for row in ws.iter_rows(min_row=2, min_col=10, max_col=11):
            for cell in row:
                cell.number_format = "#,##0.00"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = "attachment; filename=Anticipos_Lotes.xlsx"

        wb.save(response)

        return response

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


def crear_anticipo_clientes(request):

    user = request.session.get("user_id", "")
    if user == "":
        return redirect("login")

    acceso_denegado = verificar_acceso_bancario(request, "40112")
    if acceso_denegado:
        return acceso_denegado

    empresa = request.GET.get("empresa", "3")
    adminIT = request.session.get("bancarioAdminIT", 0)
    fecha_actual = datetime.today()

    context = {
        "fecha_actual": fecha_actual,
        "id_empresa": empresa,
        "adminIT": adminIT,
        "empresasData": obtener_empresas_usuario(request),
        "bancosData": get_bancos(),
        "estadosData": get_estados_movimientos(),
        "clientesData": get_clientes(),
    }

    return render(request, "bancario/crear_anticipo_clientes.html", context)


def get_cxc_pendientes_cliente(request):

    codigo_cliente = request.GET.get("codigo_cliente")

    data = []

    try:
        with connections["super"].cursor() as cursor:
            cursor.callproc("CONTA_GET_CXC_PENDIENTES_CLIENTE", [codigo_cliente])

            for row in cursor.fetchall():
                data.append(
                    {
                        "PKcxc": row[0],
                        "FechaDoc": str(row[1]),
                        "NoDocumento": row[2],
                        "ValorInicial": float(row[3]),
                        "Saldo": float(row[4]),
                    }
                )

        return JsonResponse({"data": data})

    except Exception as e:
        return JsonResponse({"data": [], "error": str(e)})


def bank_guardar_lote_abonos(request):

    if request.method == "POST":

        userName = request.session.get("userName")

        if not request.POST.get("fkBanco") or request.POST.get("fkBanco") == "0":
            return JsonResponse({"success": False, "error": "Banco no válido"})

        if (
            not request.POST.get("fkCuentaBancaria")
            or request.POST.get("fkCuentaBancaria") == "0"
        ):
            return JsonResponse(
                {"success": False, "error": "Cuenta bancaria no válida"}
            )

        if not request.POST.get("numero"):
            return JsonResponse({"success": False, "error": "Número requerido"})

        if not request.POST.get("fecha"):
            return JsonResponse({"success": False, "error": "Fecha requerida"})

        if (
            not request.POST.get("valorAnticipo")
            or float(request.POST.get("valorAnticipo")) <= 0
        ):
            return JsonResponse({"success": False, "error": "Valor inválido"})

        if not request.POST.get("concepto"):
            return JsonResponse({"success": False, "error": "Concepto requerido"})

        try:

            with connections["bankConn"].cursor() as cursor:

                cursor.callproc(
                    "BK_INSERT_CXC_LOTE_COMPLETO",
                    [
                        request.POST.get("fkEmpresa"),
                        request.POST.get("empresaNombre"),
                        request.POST.get("fkBanco"),
                        request.POST.get("bancoNombre"),
                        request.POST.get("fkCuentaBancaria"),
                        request.POST.get("cuentaBancaria"),
                        request.POST.get("moneda"),
                        request.POST.get("numero"),
                        request.POST.get("fecha"),
                        request.POST.get("valorAnticipo"),
                        request.POST.get("concepto"),
                        userName,
                        request.POST.get("detalle"),
                    ],
                )

            return JsonResponse({"success": True})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})


def guardar_movimiento_credito(request):

    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método inválido"})

    try:
        data = json.loads(request.body)

        encabezado = data.get("encabezado", {})
        detalle = data.get("detalle", [])
        liquidaciones = data.get("liquidaciones", [])

        if not encabezado or not detalle:
            return JsonResponse({"success": False, "error": "Datos incompletos"})

        empresa = encabezado.get("empresa")
        sucursal = encabezado.get("sucursal")
        cuenta = encabezado.get("cuenta")
        fecha = encabezado.get("fecha")
        numero = encabezado.get("numero")
        valor_doc = float(encabezado.get("valorDoc", 0))
        concepto_general = encabezado.get("conceptoGeneral")
        tipo_movimiento = encabezado.get("tipoMovimiento")

        fk_usuario = request.session.get("user_id")

        if not fk_usuario:
            return JsonResponse({"success": False, "error": "Sesión inválida"})

        valor_lps = valor_doc
        tasa_cambio = 1

        # =============================
        # TRANSACCIÓN COMPLETA
        # =============================

        with connections["bankConn"].cursor() as cursor:

            try:
                cursor.execute("START TRANSACTION")

                # -------------------------
                # INSERT MOVIMIENTO PRINCIPAL
                # -------------------------

                cursor.callproc(
                    "BK_INSERT_MOV_BANCARIO",
                    [
                        None,
                        tipo_movimiento,
                        cuenta,
                        None,
                        fecha,
                        numero,
                        None,
                        valor_doc,
                        valor_lps,
                        concepto_general,
                        fk_usuario,
                        None,
                        4,
                        fk_usuario,
                        None,
                        fk_usuario,
                        5,
                        0,
                        0,
                        0,
                        0,
                        empresa,
                        0,
                        0,
                        tasa_cambio,
                        fecha,
                        0,
                        sucursal,
                    ],
                )

                result = cursor.fetchone()

                if not result:
                    cursor.execute("ROLLBACK")
                    return JsonResponse(
                        {
                            "success": False,
                            "error": "No se pudo obtener el ID del movimiento",
                        }
                    )

                pk_mov = result[0]

                # Limpiar posibles resultsets pendientes
                while cursor.nextset():
                    pass

                # -------------------------
                # INSERT DETALLES
                # -------------------------

                for d in detalle:

                    cursor.callproc(
                        "SP_InsertDetalleMov",
                        [
                            pk_mov,
                            d.get("fkProcedencia"),
                            d.get("fkTarjeta"),
                            d.get("numero"),
                            d.get("fkTipoDocumento"),
                            d.get("valor"),
                            d.get("valor"),
                            d.get("concepto"),
                            d.get("fkBanco"),
                            d.get("fkCuenta"),
                            d.get("librador"),
                            0,
                            0,
                            d.get("fkMovimientoXAplicar"),
                            d.get("fkCliente"),
                        ],
                    )

                    while cursor.nextset():
                        pass

                # -------------------------
                # INSERT LIQUIDACIONES CAJA
                # -------------------------
                if liquidaciones and len(liquidaciones) > 0:
                    with connections["super"].cursor() as cursor_super:
                        for liq in liquidaciones:
                            pkCajaMov = liq.get("pkCajaMov")
                            valorLiquidado = liq.get("valorAplicado", 0)

                            if pkCajaMov and float(valorLiquidado) > 0:
                                cursor_super.callproc(
                                    "CONTA_SAVE_CAJA_LIQUIDACION",
                                    [pkCajaMov, pk_mov, valorLiquidado],
                                )
                                # Limpiar resultsets en el otro cursor
                                while cursor_super.nextset():
                                    pass
                        cursor_super.execute("COMMIT")

                cursor.execute("COMMIT")

                return JsonResponse({"success": True, "pk": pk_mov})

            except Exception as e:
                cursor.execute("ROLLBACK")
                return JsonResponse({"success": False, "error": str(e)})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


def validar_mes_movimiento(request):
    fecha = request.GET.get("fecha")
    empresa = request.session.get("empresa_id")

    fecha_dt = datetime.strptime(fecha, "%Y-%m-%d")
    anio = fecha_dt.year
    mes = fecha_dt.month

    with connections["universal"].cursor() as cursor:
        cursor.callproc("CONTA_VALIDAR_MES_ABIERTO", [anio, mes, empresa])
        result = cursor.fetchone()

    cerrado = result[0] if result else 0

    return JsonResponse({"cerrado": cerrado})


def crear_movimiento_credito(request):

    user = request.session.get("user_id", "")
    if user == "":
        return redirect("login")

    tipo = request.GET.get("tipo", "1")

    # Validar permisos
    OPCIONES_CREDITO = {
        "5": "90303",  # Credito Electronico
        "6": "90302",  # Nota de Credito
        "7": "90301",  # Deposito
    }
    permiso_requerido = OPCIONES_CREDITO.get(str(tipo), "")
    if permiso_requerido:
        acceso_denegado = verificar_acceso_bancario(request, permiso_requerido)
        if acceso_denegado:
            return acceso_denegado
    empresa = request.GET.get("empresa", "3")
    adminIT = request.session.get("bancarioAdminIT", 0)
    fecha_actual = datetime.today()

    TITULOS = {
        "5": "CREDITO ELECTRONICO",
        "6": "NOTA DE CREDITO",
        "7": "DEPOSITO",
    }

    context = {
        "titulo_tipo": TITULOS.get(tipo, "MOVIMIENTO"),
        "fecha_actual": fecha_actual,
        "fkTipoMovimiento": tipo,
        "id_empresa": empresa,
        "adminIT": adminIT,
        "empresasData": obtener_empresas_usuario(request),
        "bancosData": get_bancos(),
        "estadosData": get_estados_movimientos(),
        "destinosData": get_destinos_usuario(empresa, user, adminIT),
        "procedenciasData": get_procedencias_usuario(empresa, user, adminIT),
        "acreedoresData": get_acreedores(),
        "proveedoresData": get_proveedores(),
        "sucursalesData": obtener_sucursales_usuario(request, empresa),
        "tiposDoctoData": get_tipos_documentos(),
        "clientesData": get_clientes(),
        "tarjetasData": get_tarjetas_credito(empresa),
    }

    return render(request, "bancario/crear_movimiento_credito.html", context)


def crear_movimiento_debito(request):

    user = request.session.get("user_id", "")
    if user == "":
        return redirect("login")

    tipo = request.GET.get("tipo", "1")

    # Validar permisos
    OPCIONES_DEBITO = {
        "1": "90101",  # Cheque
        "2": "90103",  # Debito Electronico
        "3": "90102",  # Nota de Debito
        "4": "90104",  # Retiro en Ventanilla
    }
    permiso_requerido = OPCIONES_DEBITO.get(str(tipo), "")
    if permiso_requerido:
        acceso_denegado = verificar_acceso_bancario(request, permiso_requerido)
        if acceso_denegado:
            return acceso_denegado
    empresa = request.GET.get("empresa", "3")
    adminIT = request.session.get("bancarioAdminIT", 0)
    fecha_actual = datetime.today()

    TITULOS = {
        "1": "CHEQUE",
        "2": "DEBITO ELECTRONICO",
        "3": "NOTA DE DEBITO",
        "4": "RETIRO EN VENTANILLA",
    }

    context = {
        "titulo_tipo": TITULOS.get(tipo, "MOVIMIENTO"),
        "fecha_actual": fecha_actual,
        "fkTipoMovimiento": tipo,
        "id_empresa": empresa,
        "adminIT": adminIT,
        "empresasData": obtener_empresas_usuario(request),
        "bancosData": get_bancos(),
        "estadosData": get_estados_movimientos(),
        "destinosData": get_destinos_usuario(empresa, user, adminIT),
        "acreedoresData": get_acreedores(),
        "proveedoresData": get_proveedores(),
    }

    return render(request, "bancario/crear_movimiento_debito.html", context)


def api_get_saldo_cuenta(request):
    try:

        fk_cuenta = int(request.GET.get("cuenta", 0))

        if fk_cuenta <= 0:
            return JsonResponse({"success": False, "message": "Cuenta inválida"})

        with connections["bankConn"].cursor() as cursor:

            cursor.execute("SELECT fn_GetSaldoCuenta(%s)", [fk_cuenta])
            row = cursor.fetchone()
            saldo = float(row[0] or 0)

        return JsonResponse({"success": True, "saldo": saldo})

    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)})


def saveMovimientoBancarioDebito(request):
    try:
        payload = json.loads(request.POST.get("payload", "{}"))

        fk_usuario = int(request.session.get("user_id", 0))
        if fk_usuario <= 0:
            return JsonResponse({"success": False, "message": "Usuario inválido"})

        # ======================================================
        # DATOS BASE
        # ======================================================
        fk_empresa = int(payload.get("empresa", 1))
        destino = int(payload.get("destino", 0))
        tipo = int(payload.get("tipo", 0))

        cuenta = payload.get("cuenta")
        banco = payload.get("banco")
        estado = payload.get("estado")

        fecha = payload.get("fecha")
        numero = payload.get("numero")

        beneficiario = payload.get("beneficiario")
        concepto = payload.get("concepto")
        entidad_id = payload.get("entidad_id")

        valor = float(payload.get("valor", 0))
        valor_lps = float(payload.get("valorLps", 0))

        ingresos = payload.get("ingresos", [])

        ck_taf = bool(payload.get("ck_taf", False))

        es_cheque = bool(payload.get("es_cheque", False))
        pago_proveedores = bool(payload.get("pago_proveedores", False))

        if not banco:
            return JsonResponse(
                {"success": False, "message": "No ha seleccionado banco"}
            )
        if not cuenta:
            return JsonResponse(
                {"success": False, "message": "No ha seleccionado cuenta"}
            )
        if not fecha:
            return JsonResponse(
                {"success": False, "message": "No ha seleccionado fecha"}
            )
        if not numero:
            return JsonResponse({"success": False, "message": "No ha ingresado número"})
        if valor <= 0 or valor_lps <= 0:
            return JsonResponse({"success": False, "message": "Valor inválido"})
        if destino <= 0:
            return JsonResponse({"success": False, "message": "Destino requerido"})
        if not concepto:
            return JsonResponse({"success": False, "message": "Sinopsis requerida"})

        tasa_cambio_mov = valor_lps / valor

        pk_mov_ts = 0
        pk_mov_debito = 0
        pk_mov_credito = 0

        # ======================================================
        # TRANSACCION
        # ======================================================
        with transaction.atomic(using="bankConn"):

            with connections["bankConn"].cursor() as cursor:

                # ==================================================
                # OBTENER FLAGS TS DESDE BD (NO FRONTEND)
                # ==================================================
                cursor.execute("SELECT fn_GetGeneraTSCuenta(%s)", [cuenta])
                ts_cuenta = cursor.fetchone()[0]

                cursor.execute("SELECT fn_GetGeneraTSDestino(%s)", [destino])
                ts_destino = cursor.fetchone()[0]

                # ==================================================
                # BLOQUE TS / TAF
                # ==================================================
                if ts_cuenta and ts_destino:

                    valor_ts = float(payload.get("valor_ts", 0))

                    concepto_ts = (
                        "Pago Tasa de Seguridad sobre "
                        f"{valor:,.2f} {payload.get('moneda','Lps')} "
                        f"de {tipo} {numero}"
                    )

                    # ==============================
                    # TAF
                    # ==============================
                    if ck_taf:

                        cursor.execute("SELECT fn_GetFkDestinoTAF(%s)", [fk_empresa])
                        destino_taf = cursor.fetchone()[0]

                        concepto_taf = (
                            f"Valor transferido desde la Cuenta Bancaria "
                            f"{payload.get('cuenta_taf_text','')} del Banco "
                            f"{payload.get('banco_text','')} para la Cuenta "
                            f"{payload.get('cuenta_text','')}, Pago de {concepto_ts}"
                        )

                        # ---------- NOTA DEBITO TAF ----------
                        cursor.callproc(
                            "SP_InsertMovimiento",
                            [
                                destino_taf,
                                2,
                                payload.get("cuenta_taf"),
                                None,
                                fecha,
                                numero,
                                None,
                                valor_ts,
                                valor_ts * tasa_cambio_mov,
                                concepto_taf,
                                fk_usuario,
                                None,
                                4,
                                fk_usuario,
                                None,
                                fk_usuario,
                                5,
                                0,
                                0,
                                0,
                                0,
                                fk_empresa,
                                0,
                                0,
                                tasa_cambio_mov,
                                fecha,
                                0,
                            ],
                        )
                        drain_results(cursor)

                        cursor.execute(
                            "SELECT fn_GetUltimoPkMovimiento(%s)", [fk_usuario]
                        )
                        pk_mov_debito = cursor.fetchone()[0]

                        cursor.callproc(
                            "SP_UpdateNumeroMov",
                            [pk_mov_debito, f"{fecha.replace('-','')}{pk_mov_debito}"],
                        )
                        drain_results(cursor)

                        # ---------- NOTA CREDITO TAF ----------
                        cursor.callproc(
                            "SP_InsertMovimiento",
                            [
                                None,
                                6,
                                cuenta,
                                None,
                                fecha,
                                numero,
                                None,
                                valor_ts,
                                valor_ts * tasa_cambio_mov,
                                concepto_taf,
                                fk_usuario,
                                None,
                                4,
                                fk_usuario,
                                None,
                                fk_usuario,
                                5,
                                0,
                                0,
                                0,
                                0,
                                fk_empresa,
                                0,
                                0,
                                tasa_cambio_mov,
                                fecha,
                                0,
                            ],
                        )
                        drain_results(cursor)

                        cursor.execute(
                            "SELECT fn_GetUltimoPkMovimiento(%s)", [fk_usuario]
                        )
                        pk_mov_credito = cursor.fetchone()[0]

                        cursor.callproc(
                            "SP_InsertDetalleMov",
                            [
                                pk_mov_credito,
                                1 if fk_empresa == 1 else 22,
                                None,
                                f"{fecha.replace('-','')}{pk_mov_credito}",
                                "Transferencia",
                                valor_ts,
                                valor_ts,
                                concepto_taf,
                                banco,
                                payload.get("cuenta_taf_text"),
                                payload.get("nombre_empresa"),
                                0,
                                0,
                                None,
                                None,
                            ],
                        )
                        drain_results(cursor)

                    # ==============================
                    # MOVIMIENTO TS FINAL
                    # ==============================
                    cursor.execute("SELECT fn_GetFkDestinoTS(%s)", [fk_empresa])
                    destino_ts = cursor.fetchone()[0]

                    cursor.callproc(
                        "SP_InsertMovimiento",
                        [
                            destino_ts,
                            2,
                            cuenta,
                            None,
                            fecha,
                            numero,
                            None,
                            valor_ts,
                            valor_ts * tasa_cambio_mov,
                            concepto_ts,
                            fk_usuario,
                            None,
                            4,
                            fk_usuario,
                            None,
                            fk_usuario,
                            5,
                            pk_mov_debito,
                            pk_mov_credito,
                            0,
                            0,
                            fk_empresa,
                            0,
                            0,
                            tasa_cambio_mov,
                            fecha,
                            0,
                        ],
                    )
                    drain_results(cursor)

                    cursor.execute("SELECT fn_GetUltimoPkMovimiento(%s)", [fk_usuario])
                    pk_mov_ts = cursor.fetchone()[0]

                # ==================================================
                # MOVIMIENTO PRINCIPAL
                # ==================================================
                cursor.callproc(
                    "SP_InsertMovimiento",
                    [
                        destino,
                        tipo,
                        cuenta,
                        entidad_id,
                        fecha,
                        numero,
                        beneficiario if es_cheque else None,
                        valor,
                        valor_lps,
                        concepto,
                        fk_usuario,
                        None,
                        None if es_cheque else 4,
                        None if es_cheque else fk_usuario,
                        None,
                        None if es_cheque else fk_usuario,
                        estado,
                        pk_mov_debito,
                        pk_mov_credito,
                        pk_mov_ts,
                        0,
                        fk_empresa,
                        valor if pago_proveedores else 0,
                        valor_lps if pago_proveedores else 0,
                        tasa_cambio_mov,
                        None if es_cheque else fecha,
                        0,
                    ],
                )
                drain_results(cursor)

                cursor.execute("SELECT fn_GetUltimoPkMovimiento(%s)", [fk_usuario])
                pk_mov = cursor.fetchone()[0]

                cursor.callproc(
                    "SP_UpdateNumeroMov", [pk_mov, f"{fecha.replace('-','')}{pk_mov}"]
                )
                drain_results(cursor)

                # ==================================================
                # REGISTRO PAGOS
                # ==================================================
                for row in ingresos:

                    valor_pago = float(row.get("valor_pago_lps", 0))

                    cursor.callproc(
                        "SP_InsertRegistroPago",
                        [
                            4,
                            None,
                            None,
                            pk_mov,
                            row.get("doc_id"),
                            None,
                            fecha,
                            valor_pago,
                            valor_pago,
                            0,
                            valor_pago,
                            1,
                            tasa_cambio_mov,
                            fk_usuario,
                            fk_empresa,
                            0,
                        ],
                    )
                    drain_results(cursor)

                    if row.get("doc_id"):
                        cursor.callproc(
                            "SP_SetAbonoIngreso", [row.get("doc_id"), valor_pago]
                        )
                        drain_results(cursor)

        return JsonResponse({"success": True, "pk_movimiento": pk_mov})

    except Exception as e:

        return JsonResponse({"success": False, "message": str(e)})


def api_get_sucursales_usuario(request):
    try:
        empresa = request.GET.get("empresa", 0)
        sucursales = obtener_sucursales_usuario(request, empresa)
        return JsonResponse({"success": True, "data": sucursales})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


def get_cuentas_bancarias(request):

    empresa = request.GET.get("empresa")
    status = request.GET.get("status")
    banco = request.GET.get("banco")

    data = []

    with connections["bankConn"].cursor() as cursor:

        cursor.callproc("BK_GET_CUENTAS_BANCARIAS", [empresa, status, banco])
        column_names = [desc[0] for desc in cursor.description]
        raw_rows = cursor.fetchall()
        data = []
        for row in raw_rows:
            row_dict = {}
            for col, val in zip(column_names, row):
                # Convertir campos tipo BIT (bytes) a int (0 o 1)
                if isinstance(val, bytes):
                    row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                else:
                    row_dict[str(col)] = val
            data.append(row_dict)

    return JsonResponse({"data": data})


def get_movimientos_caja_liquidaciones(request):
    try:

        sucursal = request.GET.get("sucursal")
        date1 = request.GET.get("date1")
        date2 = request.GET.get("date2")
        tipoCuenta = request.GET.get("tipoCuenta", "SALIDA")

        movimientos = []
        ids_banco = set()

        with connections["super"].cursor() as cursor:
            cursor.callproc(
                "BK_CAJA_MOVIMIENTOS_LIQUIDADOS", [sucursal, date1, date2, tipoCuenta]
            )

            columns = [col[0] for col in cursor.description]

            for row in cursor.fetchall():
                data = dict(zip(columns, row))
                movimientos.append(data)

                if data["fkMovimientoBanco"]:
                    ids_banco.add(str(data["fkMovimientoBanco"]))

            while cursor.nextset():
                pass

        referencias = {}

        if ids_banco:

            ids = ",".join(ids_banco)

            with connections["bankConn"].cursor() as cursor:
                cursor.callproc("BK_OBTENER_REFERENCIAS_X_PK_MOV", [ids])

                columns = [col[0] for col in cursor.description]

                for row in cursor.fetchall():
                    data = dict(zip(columns, row))
                    referencias[data["PkMovimientoBancario"]] = data

                while cursor.nextset():
                    pass

        for mov in movimientos:
            ref_data = referencias.get(mov["fkMovimientoBanco"], {})
            mov.update(
                {
                    "Referencia": ref_data.get("Referencia"),
                    "FechaMovimiento": (
                        str(ref_data.get("FechaMovimiento"))
                        if ref_data.get("FechaMovimiento")
                        else None
                    ),
                    "Numero": ref_data.get("Numero"),
                    "Banco": ref_data.get("Banco"),
                    "CuentaBancaria": ref_data.get("CuentaBancaria"),
                    "TipoCuenta": ref_data.get("TipoCuenta"),
                }
            )

        return JsonResponse({"data": movimientos})

    except Exception as e:
        return JsonResponse({"error": str(e)})
