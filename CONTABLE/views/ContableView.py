import base64
import csv
import hashlib
import json
import logging
import os
import pandas as pd
import pyodbc
import pymssql
import pymysql
import requests
import subprocess
import tempfile
import time
import xlrd
import jwt

from decimal import Decimal
import gzip
from io import BytesIO

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter

from datetime import date, datetime, timedelta
import decimal

from django.db.utils import OperationalError
from django.views.decorators.http import require_POST

from calendar import monthrange
from urllib.parse import quote_plus

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.db import connections, transaction
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, Http404
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt

from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken

from sqlalchemy import create_engine, text
from sqlalchemy.dialects.mysql import insert
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import sessionmaker

from CWS.views.LoginView import EXPIRATION_MINUTES, SECRET_KEY

from .utilsContable import *
from .utilsContable import _callproc_fetchall


CHUNK_SIZE = 200
TOKEN = "2e078366ee3366544e4132ebb24eb2948270bbce69aa8ff22a30a2422cc12a7e"


def get_accesos_ctrl_gestion(request):
    user_id = request.session.get("user_id", "")
    request.session["contabilidadAdminIT"] = 0
    request.session["ctrlGestionAdminIT"] = 0

    conn = connections["global_nube"]

    # 1) ¿Es admin IT?
    adminITQuery = _callproc_fetchall(conn, "WEB_GET_ADMIN_IT", [user_id, 12])
    if adminITQuery:
        request.session["contabilidadAdminIT"] = 1

    adminITCrtlGestionQuery = _callproc_fetchall(
        conn, "WEB_GET_ADMIN_IT", [user_id, 13]
    )
    if adminITQuery:
        request.session["ctrlGestionAdminIT"] = 1

    # 2) Menús contables (grupo 12)
    menusContables = _callproc_fetchall(
        conn, "WEB_GET_MENUS_GRUPO_USUARIO", [user_id, 12]
    )
    if menusContables:
        for menu in menusContables:
            posicion_menu = str(menu[2])
            permiso_menu = menu[6]
            request.session[posicion_menu] = 1 if permiso_menu == 1 else 0

    # 3) Menús de control de gestión (grupo 13)
    menuCtrlGestion = _callproc_fetchall(
        conn, "WEB_GET_MENUS_GRUPO_USUARIO", [user_id, 13]
    )
    if menuCtrlGestion:
        for menu in menuCtrlGestion:
            posicion_menu = str(menu[2])
            permiso_menu = menu[6]
            request.session[posicion_menu] = 1 if permiso_menu == 1 else 0

            if permiso_menu == 1:
                print("{} PERMISO {}".format(posicion_menu, permiso_menu))


def get_accesos_contabilidad(request):
    user_id = request.session.get("user_id", "")
    request.session["contabilidadAdminIT"] = 0

    appConn = connections["global_nube"]
    with appConn.cursor() as cursor:
        cursor.callproc("WEB_GET_ADMIN_IT", [user_id, 12])
        adminITQuery = cursor.fetchall()

        # Procesar los menús y establecer valores en la sesión
        if adminITQuery:
            request.session["contabilidadAdminIT"] = 1

        appConn.close()

    appConn = connections["global_nube"]
    with appConn.cursor() as cursor:
        cursor.callproc("WEB_GET_MENUS_GRUPO_USUARIO", [user_id, 12])
        menuQuery = cursor.fetchall()

        # Procesar los menús y establecer valores en la sesión
        if menuQuery:
            for menu in menuQuery:
                posicion_menu = menu[2]
                permiso_menu = menu[6]

                if posicion_menu == "10000" and permiso_menu == 1:
                    request.session["tabContabilidad_CNT"] = 1
                elif posicion_menu == "10000" and permiso_menu == 0:
                    request.session["tabContabilidad_CNT"] = 0

                if posicion_menu == "10100" and permiso_menu == 1:
                    request.session["grupoLibros_CNT"] = 1
                elif posicion_menu == "10100" and permiso_menu == 0:
                    request.session["grupoLibros_CNT"] = 0

                if posicion_menu == "10101" and permiso_menu == 1:
                    request.session["nuevaPartidaManual_CNT"] = 1
                elif posicion_menu == "10101" and permiso_menu == 0:
                    request.session["nuevaPartidaManual_CNT"] = 0

                if posicion_menu == "10102" and permiso_menu == 1:
                    request.session["cargarDatosExcel_CNT"] = 1
                elif posicion_menu == "10102" and permiso_menu == 0:
                    request.session["cargarDatosExcel_CNT"] = 0

                if posicion_menu == "10103" and permiso_menu == 1:
                    request.session["libroDiario_CNT"] = 1
                elif posicion_menu == "10103" and permiso_menu == 0:
                    request.session["libroDiario_CNT"] = 0

                if posicion_menu == "10104" and permiso_menu == 1:
                    request.session["libroMayor_CNT"] = 1
                elif posicion_menu == "10104" and permiso_menu == 0:
                    request.session["libroMayor_CNT"] = 0

                if posicion_menu == "10105" and permiso_menu == 1:
                    request.session["libroMayor_Codigo_CNT"] = 1
                elif posicion_menu == "10105" and permiso_menu == 0:
                    request.session["libroMayor_Codigo_CNT"] = 0

                if posicion_menu == "10200" and permiso_menu == 1:
                    request.session["grupoCierreContables_CNT"] = 1
                elif posicion_menu == "10200" and permiso_menu == 0:
                    request.session["grupoCierreContables_CNT"] = 0

                if posicion_menu == "10201" and permiso_menu == 1:
                    request.session["cierreMes_CNT"] = 1
                elif posicion_menu == "10201" and permiso_menu == 0:
                    request.session["cierreMes_CNT"] = 0

                if posicion_menu == "10202" and permiso_menu == 1:
                    request.session["entregarMes_CNT"] = 1
                elif posicion_menu == "10202" and permiso_menu == 0:
                    request.session["entregarMes_CNT"] = 0

                if posicion_menu == "10203" and permiso_menu == 1:
                    request.session["actualizarDatosMesesCerrados_CNT"] = 1
                elif posicion_menu == "10203" and permiso_menu == 0:
                    request.session["actualizarDatosMesesCerrados_CNT"] = 0

                if posicion_menu == "10204" and permiso_menu == 1:
                    request.session["entregarPeriodo_CNT"] = 1
                elif posicion_menu == "10204" and permiso_menu == 0:
                    request.session["entregarPeriodo_CNT"] = 0

                if posicion_menu == "10205" and permiso_menu == 1:
                    request.session["verMesesCerradosEntregados_CNT"] = 1
                elif posicion_menu == "10205" and permiso_menu == 0:
                    request.session["verMesesCerradosEntregados_CNT"] = 0

                if posicion_menu == "20101" and permiso_menu == 1:
                    request.session["estadoResultadosAnuales_CNT"] = 1
                elif posicion_menu == "20101" and permiso_menu == 0:
                    request.session["estadoResultadosAnuales_CNT"] = 0

                if posicion_menu == "20102" and permiso_menu == 1:
                    request.session["estadoSituacionFinanciera_CNT"] = 1
                elif posicion_menu == "20102" and permiso_menu == 0:
                    request.session["estadoSituacionFinanciera_CNT"] = 0

                if posicion_menu == "20103" and permiso_menu == 1:
                    request.session["gastosPorMes_CNT"] = 1
                elif posicion_menu == "20103" and permiso_menu == 0:
                    request.session["gastosPorMes_CNT"] = 0

                if posicion_menu == "20104" and permiso_menu == 1:
                    request.session["estadoResultadoIntegral_CNT"] = 1
                elif posicion_menu == "20104" and permiso_menu == 0:
                    request.session["estadoResultadoIntegral_CNT"] = 0

                if posicion_menu == "20105" and permiso_menu == 1:
                    request.session["balanzaComprobacion_CNT"] = 1
                elif posicion_menu == "20105" and permiso_menu == 0:
                    request.session["balanzaComprobacion_CNT"] = 0

                if posicion_menu == "20106" and permiso_menu == 1:
                    request.session["estadoFlujoEfectivo_CNT"] = 1
                elif posicion_menu == "20106" and permiso_menu == 0:
                    request.session["estadoFlujoEfectivo_CNT"] = 0

                if posicion_menu == "20107" and permiso_menu == 1:
                    request.session["razonesFinancieras_CNT"] = 1
                elif posicion_menu == "20107" and permiso_menu == 0:
                    request.session["razonesFinancieras_CNT"] = 0

                if posicion_menu == "20108" and permiso_menu == 1:
                    request.session["effYdr_CNT"] = 1
                elif posicion_menu == "20108" and permiso_menu == 0:
                    request.session["effYdr_CNT"] = 0

                if posicion_menu == "20109" and permiso_menu == 1:
                    request.session["nofYcc_CNT"] = 1
                elif posicion_menu == "20109" and permiso_menu == 0:
                    request.session["nofYcc_CNT"] = 0

                if posicion_menu == "20110" and permiso_menu == 1:
                    request.session["nof_CNT"] = 1
                elif posicion_menu == "20110" and permiso_menu == 0:
                    request.session["nof_CNT"] = 0

                if posicion_menu == "20111" and permiso_menu == 1:
                    request.session["roeDupont_CNT"] = 1
                elif posicion_menu == "20111" and permiso_menu == 0:
                    request.session["roeDupont_CNT"] = 0

                if posicion_menu == "20112" and permiso_menu == 1:
                    request.session["freeCashFlow_CNT"] = 1
                elif posicion_menu == "20112" and permiso_menu == 0:
                    request.session["freeCashFlow_CNT"] = 0

                if posicion_menu == "30101" and permiso_menu == 1:
                    request.session["empresas_CNT"] = 1
                elif posicion_menu == "30101" and permiso_menu == 0:
                    request.session["empresas_CNT"] = 0

                if posicion_menu == "30102" and permiso_menu == 1:
                    request.session["cuentasContables_CNT"] = 1
                elif posicion_menu == "30102" and permiso_menu == 0:
                    request.session["cuentasContables_CNT"] = 0

                if posicion_menu == "30103" and permiso_menu == 1:
                    request.session["cuentasBalance_CNT"] = 1
                elif posicion_menu == "30103" and permiso_menu == 0:
                    request.session["cuentasBalance_CNT"] = 0

                if posicion_menu == "30104" and permiso_menu == 1:
                    request.session["sucursales_CNT"] = 1
                elif posicion_menu == "30104" and permiso_menu == 0:
                    request.session["sucursales_CNT"] = 0

                if posicion_menu == "30105" and permiso_menu == 1:
                    request.session["departamentos_CNT"] = 1
                elif posicion_menu == "30105" and permiso_menu == 0:
                    request.session["departamentos_CNT"] = 0

                if posicion_menu == "30106" and permiso_menu == 1:
                    request.session["cuentasGastos_CNT"] = 1
                elif posicion_menu == "30106" and permiso_menu == 0:
                    request.session["cuentasGastos_CNT"] = 0

                if posicion_menu == "30107" and permiso_menu == 1:
                    request.session["encargadosDepartamentos_CNT"] = 1
                elif posicion_menu == "30107" and permiso_menu == 0:
                    request.session["encargadosDepartamentos_CNT"] = 0

                if posicion_menu == "30108" and permiso_menu == 1:
                    request.session["reglasUsoCuentas_CNT"] = 1
                elif posicion_menu == "30108" and permiso_menu == 0:
                    request.session["reglasUsoCuentas_CNT"] = 0

                if posicion_menu == "20000" and permiso_menu == 1:
                    request.session["pestañaEstadosFinancieros_CNT"] = 1
                elif posicion_menu == "20000" and permiso_menu == 0:
                    request.session["pestañaEstadosFinancieros_CNT"] = 0

                if posicion_menu == "20100" and permiso_menu == 1:
                    request.session["grupoEstadosFinancieros_CNT"] = 1
                elif posicion_menu == "20100" and permiso_menu == 0:
                    request.session["grupoEstadosFinancieros_CNT"] = 0

                if posicion_menu == "30000" and permiso_menu == 1:
                    request.session["pestañaCatalogos_CNT"] = 1
                elif posicion_menu == "30000" and permiso_menu == 0:
                    request.session["pestañaCatalogos_CNT"] = 0

                if posicion_menu == "30100" and permiso_menu == 1:
                    request.session["grupoCatalogo_CNT"] = 1
                elif posicion_menu == "30100" and permiso_menu == 0:
                    request.session["grupoCatalogo_CNT"] = 0

                if posicion_menu == "30200" and permiso_menu == 1:
                    request.session["grupoNIC_CNT"] = 1
                elif posicion_menu == "30200" and permiso_menu == 0:
                    request.session["grupoNIC_CNT"] = 0

                if posicion_menu == "30201" and permiso_menu == 1:
                    request.session["perfilesNIC_CNT"] = 1
                elif posicion_menu == "30201" and permiso_menu == 0:
                    request.session["perfilesNIC_CNT"] = 0

                if posicion_menu == "30202" and permiso_menu == 1:
                    request.session["detallesNIC_CNT"] = 1
                elif posicion_menu == "30202" and permiso_menu == 0:
                    request.session["detallesNIC_CNT"] = 0

                if posicion_menu == "30300" and permiso_menu == 1:
                    request.session["grupoConfiguraciones_CNT"] = 1
                elif posicion_menu == "30300" and permiso_menu == 0:
                    request.session["grupoConfiguraciones_CNT"] = 0

                if posicion_menu == "30301" and permiso_menu == 1:
                    request.session["firmasEstadosFinancieros_CNT"] = 1
                elif posicion_menu == "30301" and permiso_menu == 0:
                    request.session["firmasEstadosFinancieros_CNT"] = 0

                if posicion_menu == "30302" and permiso_menu == 1:
                    request.session["definirDecimales_CNT"] = 1
                elif posicion_menu == "30302" and permiso_menu == 0:
                    request.session["definirDecimales_CNT"] = 0

                if posicion_menu == "30109" and permiso_menu == 1:
                    request.session["cuentasContablesPorSD_CNT"] = 1
                elif posicion_menu == "30109" and permiso_menu == 0:
                    request.session["cuentasContablesPorSD_CNT"] = 0

                if posicion_menu == "30500" and permiso_menu == 1:
                    request.session["clasificacionCuentas_CNT"] = 1
                elif posicion_menu == "30500" and permiso_menu == 0:
                    request.session["clasificacionCuentas_CNT"] = 0

                if posicion_menu == "30510" and permiso_menu == 1:
                    request.session["libroDiarioVerTodo_CNT"] = 1
                elif posicion_menu == "30510" and permiso_menu == 0:
                    request.session["libroDiarioVerTodo_CNT"] = 0

                if posicion_menu == "30520" and permiso_menu == 1:
                    request.session["tasaCambio_CNT"] = 1
                elif posicion_menu == "30520" and permiso_menu == 0:
                    request.session["tasaCambio_CNT"] = 0

                if posicion_menu == "30530" and permiso_menu == 1:
                    request.session["catPartidasContables_CNT"] = 1
                elif posicion_menu == "30530" and permiso_menu == 0:
                    request.session["catPartidasContables_CNT"] = 0

        appConn.close()


def conta_reporte_ingregracion_cuentas(request):
    user_id = request.session.get("user_id", "")

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        anio_actual = datetime.now().year

        context = {
            "anio": anio_actual,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
            "empresasData": obtener_empresas(),
            "cuentas_bases": obtener_cuentas_base(),
        }

        return render(request, "contabilidad/reporte_integracion_cuentas.html", context)


def dataReporteIntegracionCuentasContables(request):
    try:
        empresa = request.POST.get("empresa")
        date1 = request.POST.get("date1")
        date2 = request.POST.get("date2")
        filtro = request.POST.get("filtro")  # 1 = solo con saldo, 0 = todo

        with connections["contable"].cursor() as cursor:
            cursor.callproc(
                "CONTA_REPORTE_INTEGRACION_CUENTAS_CONTABLES",
                [empresa, date1, date2, filtro],
            )
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"data": data})

    except Exception as e:
        return JsonResponse({"error": str(e)})


def conta_configuraciones_contables(request):
    user_id = request.session.get("user_id", "")

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        anio_actual = datetime.now().year

        context = {
            "anio": anio_actual,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "contabilidad/configuraciones_contables.html", context)


def dataConfiguraciones(request):
    try:
        with connections["contable"].cursor() as cursor:
            cursor.callproc("CONTA_GET_CONFIGURACIONES")
            column_names = [desc[0] for desc in cursor.description]
            data = [dict(zip(map(str, column_names), row)) for row in cursor.fetchall()]

        return JsonResponse({"data": data})

    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataConfiguracionesHistorico(request):
    try:
        id_configuracion = request.POST.get("id_configuracion")

        with connections["contable"].cursor() as cursor:
            cursor.callproc("CONTA_GET_CONFIGURACIONES_HISTORICO", [id_configuracion])
            column_names = [desc[0] for desc in cursor.description]
            data = [dict(zip(map(str, column_names), row)) for row in cursor.fetchall()]

        return JsonResponse({"data": data})

    except Exception as e:
        return JsonResponse({"error": str(e)})


def conta_insert_update_configuraciones(request):
    try:
        opc = request.POST.get("opcion")
        id_configuracion = request.POST.get("id_configuracion")
        configuracion = request.POST.get("configuracion")
        valor = request.POST.get("valor")

        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["contable"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_INSERT_UPDATE_CONFIGURACIONES",
                [opc, id_configuracion, configuracion, valor, userName],
            )
            results = cursor.fetchall()

        appConn.close()

        return JsonResponse({"save": 1, "result": results})

    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)})


def conta_update_status_configuraciones(request):
    try:
        id_configuracion = request.POST.get("id_configuracion")
        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["contable"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_UPDATE_STATUS_CONFIGURACION", [id_configuracion, userName]
            )
            results = cursor.fetchall()

        appConn.close()

        return JsonResponse({"save": 1, "result": results})

    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)})


def panel_contabilidad(request):
    user_id = request.session.get("user_id", "")

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        get_accesos_contabilidad(request)
        get_accesos_ctrl_gestion(request)

        anio_actual = datetime.now().year

        context = {
            "anio": anio_actual,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "panel_contable.html", context)


def conta_clasificacion_cuentas(request):
    user_id = request.session.get("user_id", "")
    clasificacionCuentas_CNT = request.session.get("clasificacionCuentas_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if clasificacionCuentas_CNT == 1 or adminIT == 1:
            return render(request, "catalogos/clasificacionCuentas.html")
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def conta_clasificacion_cuentas_nic(request):
    user_id = request.session.get("user_id", "")
    clasificacionCuentas_NIC_CNT = request.session.get("detallesNIC_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if clasificacionCuentas_NIC_CNT == 1 or adminIT == 1:

            tiposNICData = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_TIPOS_NIC", [1])
                    column_names = [desc[0] for desc in cursor.description]
                    tiposNICData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                # Cierra la conexión
                udcConn.close()
            except Exception as e:
                # Manejo de excepciones, puedes personalizar esto según tus necesidades
                return JsonResponse({"error": str(e)})

            cuentasBaseData = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_CUENTAS_BASES", [1])
                    column_names = [desc[0] for desc in cursor.description]
                    cuentasBaseData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                # Cierra la conexión
                udcConn.close()
            except Exception as e:
                # Manejo de excepciones, puedes personalizar esto según tus necesidades
                return JsonResponse({"error": str(e)})

            empresasData = obtener_empresas()

            context = {
                "tiposNICData": tiposNICData,
                "cuentasBaseData": cuentasBaseData,
                "empresasData": empresasData,
            }

            return render(request, "catalogos/clasificacionNIC.html", context)
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def cuentas_nic(request):
    user_id = request.session.get("user_id", "")

    cuentas_NIC_CNT = request.session.get("perfilesNIC_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if cuentas_NIC_CNT == 1 or adminIT == 1:

            tiposNICData = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_TIPOS_NIC", [1])
                    column_names = [desc[0] for desc in cursor.description]
                    tiposNICData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                udcConn.close()
            except Exception as e:
                return JsonResponse({"error": str(e)})

            cuentasBaseData = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_CUENTAS_BASES", [1])
                    column_names = [desc[0] for desc in cursor.description]
                    cuentasBaseData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                udcConn.close()
            except Exception as e:
                return JsonResponse({"error": str(e)})

            context = {
                "tiposNICData": tiposNICData,
                "cuentasBaseData": cuentasBaseData,
            }

            return render(request, "catalogos/cuentasNIC.html", context)
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def catalogo_partidas_contables(request):
    user_id = request.session.get("user_id", "")

    catPartidasContables_CNT = request.session.get("catPartidasContables_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if catPartidasContables_CNT == 1 or adminIT == 1:

            modulosData = getModulos()
            hermanasData = conta_get_partidas_hermanas()

            cuentasData = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_CAT_CUENTAS", [3])
                    column_names = [desc[0] for desc in cursor.description]
                    cuentasData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]
                udcConn.close()
            except Exception as e:
                return JsonResponse({"error": str(e)})

            context = {
                "adminIT": adminIT,
                "modulosData": modulosData,
                "cuentasData": cuentasData,
                "hermanasData": hermanasData,
            }

            return render(request, "catalogos/catalogo_partidas.html", context)
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def dataTiposPartidas(request):
    opcion = request.POST.get("opcion")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_TIPOS_PARTIDAS", [opcion])
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            partidasData = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir BIT (bytes) a int
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                partidasData.append(row_dict)

        udcConn.close()
        return JsonResponse({"data": partidasData})

    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataEstructuraTipoPartida(request):
    codigo = request.POST.get("codigo")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_DETALLES_TIPOS_PARTIDAS_X_CODIGO", [codigo])
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            detallesPartidasData = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir BIT (bytes) a int
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                detallesPartidasData.append(row_dict)

        udcConn.close()
        return JsonResponse({"data": detallesPartidasData})

    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataDetallesTiposPartidas(request):
    partida_id = request.POST.get("partida_id")
    opcion = request.POST.get("opcion")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_DETALLES_TIPOS_PARTIDAS", [partida_id, opcion])
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            detallesPartidasData = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir BIT (bytes) a int
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                detallesPartidasData.append(row_dict)

        udcConn.close()
        return JsonResponse({"data": detallesPartidasData})

    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataCuentasDisponiblesNIC(request):
    empresa = request.POST.get("empresa")
    tipoNIC = request.POST.get("tipoNIC")
    padre = request.POST.get("padre")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_CUENTAS_X_CLASIFICAR_NIC", [empresa, tipoNIC, padre])
            column_names = [desc[0] for desc in cursor.description]
            cuentasData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"data": cuentasData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataCuentasAsociadasNIC(request):
    empresa = request.POST.get("empresa")
    tipoNIC = request.POST.get("tipoNIC")
    padre = request.POST.get("padre")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_CUENTAS_X_NIC", [empresa, tipoNIC, padre])
            column_names = [desc[0] for desc in cursor.description]
            cuentasData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"data": cuentasData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def insert_remove_cuenta_x_nic(request):
    try:
        empresa = request.POST.get("empresa")
        cuenta_nic = request.POST.get("cuenta_nic")
        id_cuenta = request.POST.get("id_cuenta")
        codigo = request.POST.get("codigo")
        opcion = request.POST.get("opcion")

        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_ADD_REMOVE_CUENTA_X_NIC",
                [cuenta_nic, id_cuenta, codigo, empresa, opcion, userName],
            )
            results = cursor.fetchall()

        appConn.close()

        datos = {"save": 1}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def dataClasificacionCuentas(request):
    opcion = request.POST.get("opcion")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_CUENTAS_BASES", [opcion])
            column_names = [desc[0] for desc in cursor.description]
            cuentasData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"data": cuentasData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataCuentasNIC(request):
    tipoNIC = request.POST.get("tipoNIC")
    opcion = request.POST.get("opcion")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_CUENTAS_NIC_X_TIPO", [tipoNIC, opcion])
            column_names = [desc[0] for desc in cursor.description]
            cuentasData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"data": cuentasData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def cuentas_contables(request):
    user_id = request.session.get("user_id", "")

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        context = {
            "cuentas_bases": obtener_cuentas_base(),
            "tipo_saldo": obtener_tipo_saldo(),
            "empresasData": obtener_empresas(),
        }

        return render(request, "catalogos/cuentas_contables.html", context)


def cuentas_gastos(request):
    user_id = request.session.get("user_id", "")

    cuentasGastos_CNT = request.session.get("cuentasGastos_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if cuentasGastos_CNT == 1 or adminIT == 1:

            context = {
                "empresasData": obtener_empresas(),
            }

            return render(request, "catalogos/cuentas_gastos.html", context)
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def dataCuentasGastos(request):
    empresa = request.POST.get("empresa")
    opcion = request.POST.get("opcion")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_CUENTAS_GASTOS", [empresa, opcion])
            column_names = [desc[0] for desc in cursor.description]
            cuentasData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"data": cuentasData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataCuentasContables(request):
    empresa = request.POST.get("empresa")
    opcion = request.POST.get("opcion")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_MOSTRAR_CUENTAS_CONTABLES", [opcion, empresa])
            column_names = [desc[0] for desc in cursor.description]
            cuentasData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"data": cuentasData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def update_status_tipos_partidas(request):
    try:
        partida_id = request.POST.get("partida_id")
        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc("CONTA_UPDATE_STATUS_TIPO_PARTIDA", [partida_id, userName])
            results = cursor.fetchall()

        appConn.close()

        datos = {"save": 1}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def update_status_cuentas_gastos(request):
    try:
        cuenta_id = request.POST.get("cuenta_id")
        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc("CONTA_STATUS_UPDATE_CUENTA_GASTO", [cuenta_id, userName])
            results = cursor.fetchall()

        appConn.close()

        datos = {"save": 1}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def get_tasa_cambio_hoy(request):
    try:
        # Fecha de hoy (server). Si quieres la del usuario, pásala desde JS
        with connections["universal"].cursor() as cursor:
            cursor.callproc("CONTA_GET_TASA_CAMBIO_HOY")
            row = cursor.fetchone()

        data = {}
        if row:
            # Orden según el SELECT del SP
            data = {
                "fecha": str(row[0]),
                "tasa_cambio": float(row[1]),
                "tasa_compra": float(row[2]),
                "tasa_banco": float(row[3]),
            }

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)}, status=500)


def obtener_cuentas_base():
    with connections["universal"].cursor() as cursor:
        cursor.callproc("CONTA_CUENTAS_BASES", [1])
        results = cursor.fetchall()

    return results


def obtener_tipo_saldo():
    with connections["universal"].cursor() as cursor:
        cursor.callproc("CONTA_GET_TIPO_SALDO")
        results = cursor.fetchall()

    return results


def obtener_subcuentas(request):
    cuenta_padre = request.GET.get("cuenta_padre")
    if cuenta_padre:
        with connections["universal"].cursor() as cursor:
            cursor.callproc("CONTA_OBTENER_SUBCUENTAS", [cuenta_padre])
            subcuentas = cursor.fetchall()

        data = [{"id": sub[0], "nombre": sub[2]} for sub in subcuentas]
        return JsonResponse({"subcuentas": data}, safe=False)
    return JsonResponse({"error": "No se proporcionó cuenta_padre"}, status=400)


def generar_codigo(request):
    cuenta_padre = request.GET.get("cuenta_padre", None)
    nuevo_codigo = None

    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc("CONTA_GENERAR_CODIGO", [cuenta_padre, "@nuevo_codigo"])

            cursor.execute("SELECT @nuevo_codigo")
            resultado = cursor.fetchone()

            if resultado:
                nuevo_codigo = resultado[0]
            else:
                nuevo_codigo = None

    except Exception as e:
        nuevo_codigo = None

    return JsonResponse({"nuevo_codigo": nuevo_codigo})


def conta_update_es_cuenta_padre(request):
    try:
        cuenta_id = request.POST.get("cuenta_id", 0)
        es_padre = request.POST.get("es_padre", 0)
        userName = request.session.get("userName", "CONTABLE")

        with connections["universal"].cursor() as cursor:
            cursor.callproc(
                "CONTA_UPDATE_ES_CUENTA_PADRE", [cuenta_id, es_padre, userName]
            )

        datos = {"save": 1}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def conta_update_es_padre_cuentas_gastos(request):
    try:
        cuenta_id = request.POST.get("cuenta_id", 0)
        es_padre = request.POST.get("es_padre", 0)

        userName = request.session.get("userName", "CONTABLE")

        with connections["universal"].cursor() as cursor:
            cursor.callproc(
                "CONTA_UPDATE_ES_PADRE_CUENTA_GASTOS", [cuenta_id, es_padre, userName]
            )

        datos = {"save": 1}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def normalize_row(row):
    fixed = []
    for v in row:
        if isinstance(v, (bytes, bytearray)):  # para bit(1)
            fixed.append(int.from_bytes(v, byteorder="little"))
        elif v is None:  # para NoneType
            fixed.append(None)  # o '' si querés string vacío
        else:
            fixed.append(v)
    return tuple(fixed)


def sync_temp_cuentas():
    try:
        print(">>> Iniciando sync_temp_cuentas()")

        # Traer data oficial
        with connections["universal"].cursor() as cursor:
            print(">>> Ejecutando SELECT * FROM conta_cuentas_contables")
            cursor.execute(
                """
                SELECT id_cuenta, codigo, nombre, descripcion, tratamiento,
                    tipo_saldo, tipo_cuenta, id_tipo_movimiento, creado_por, fecha_hora_creado,
                    es_cuenta_padre, padre_id, id_cuenta_base, id_empresa
                FROM conta_cuentas_contables
            """
            )
            rows = cursor.fetchall()
            print(f">>> Registros obtenidos: {len(rows)}")

        # Vaciar temporal
        with connections["contable_zeus"].cursor() as cursor:
            print(">>> Ejecutando TRUNCATE TABLE temp_cuentascontables")
            cursor.execute("TRUNCATE TABLE temp_cuentascontables")
            print(">>> Tabla temporal vaciada correctamente")

        # Insertar en temporal
        insert_sql = """
            INSERT INTO temp_cuentascontables
            (id_cuenta, codigo, nombre, descripcion, tratamiento,
             tipo_saldo, tipo_cuenta, id_tipo_movimiento, creado_por, fecha_hora_creado,
             es_cuenta_padre, padre_id, id_cuenta_base, id_empresa)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """

        # Normalizar todas las filas
        rows_normalized = [normalize_row(r) for r in rows]

        with connections["contable_zeus"].cursor() as cursor:
            print(">>> Ejecutando INSERT batch en temp_cuentascontables")
            cursor.executemany(insert_sql, rows_normalized)
            print(">>> Inserción completada en la tabla temporal")

        print(">>> Llamando a migrar_cuentas_contables()")
        result = migrar_cuentas_contables()
        print(f">>> Resultado migración: {result}")

        print(">>> sync_temp_cuentas() completado")
        return JsonResponse(
            {
                "status": "success",
                "message": f"{len(rows)} registros copiados a temporal",
            }
        )

    except Exception as e:
        print(">>> ERROR en sync_temp_cuentas:", str(e))
        return JsonResponse({"status": "fail", "error": str(e)})


def migrar_cuentas_contables():
    try:
        print(">>> Iniciando migrar_cuentas_contables()")
        with connections["contable_zeus"].cursor() as cursor:
            print(">>> Llamando SP CONTA_MIGRAR_CUENTAS_CONTABLES")
            cursor.callproc("CONTA_MIGRAR_CUENTAS_CONTABLES")
            result = cursor.fetchone()
            print(f">>> SP ejecutado, resultado: {result}")

        return JsonResponse(
            {"status": "success", "message": result[0] if result else "OK"}
        )
    except Exception as e:
        print(">>> ERROR en migrar_cuentas_contables:", str(e))
        return JsonResponse({"status": "fail", "error": str(e)})


def insertar_actualizar_cuenta(request):
    if request.method == "POST":

        id_cuenta = request.POST.get("id_cuenta")
        codigo = request.POST.get("codigo")
        codigo_viejo = request.POST.get("codigo_viejo")
        nombre = request.POST.get("nombre")
        tipo_cuenta = request.POST.get("tipo_cuenta")
        cuenta_padre = request.POST.get("cuenta_padre")
        descripcion = request.POST.get("descripcion")
        tipo_saldo = request.POST.get("tipo_saldo")
        se_acredita_cuando = request.POST.get("se_acredita_cuando")
        tratamiento_cuenta = request.POST.get("tratamiento_cuenta")
        FkCuentaBase = request.POST.get("FkCuentaBase")
        opcion = int(request.POST.get("opcion", 1))
        creado_por = request.session.get("userName", "CONTABLE")
        es_cuenta_padre = int(request.POST.get("es_cuenta_padre", 0))
        empresa = int(request.POST.get("empresa", 0))
        crearTodosCentros = int(request.POST.get("crearTodosCentros", 0))

        # Manejo de valores nulos
        cuenta_padre = None if not cuenta_padre else cuenta_padre
        tipo_saldo = None if not tipo_saldo else tipo_saldo
        se_acredita_cuando = "-"

        try:
            appConn = connections["universal"]
            with appConn.cursor() as cursor:
                cursor.callproc(
                    "CONTA_INSERT_UPDATE_CUENTA_CONTABLE",
                    [
                        id_cuenta,  # ID de la cuenta
                        codigo,  # Código de la cuenta
                        codigo_viejo,  # Código viejo de la cuenta
                        nombre,  # Nombre de la cuenta
                        tipo_cuenta,  # Tipo de cuenta
                        cuenta_padre,  # ID de la cuenta padre
                        creado_por,  # Usuario que realiza la acción
                        descripcion,  # Descripción
                        tipo_saldo,  # Tipo de saldo
                        se_acredita_cuando,  # Se acredita cuando
                        tratamiento_cuenta,
                        FkCuentaBase,
                        es_cuenta_padre,
                        crearTodosCentros,
                        empresa,
                        opcion,  # Opción: Insertar o Actualizar
                    ],
                )

                # Obtener el resultado del procedimiento almacenado
                result = cursor.fetchone()
                existe = result[0] if result else 0

            # Respuesta según el resultado
            if existe == 0:
                return JsonResponse(
                    {
                        "status": "success",
                        "message": "Operación realizada exitosamente.",
                    }
                )
            else:
                return JsonResponse(
                    {
                        "status": "fail",
                        "message": "Registro duplicado. Verifique el código o el nombre.",
                    }
                )

            appConn.close()

            sync_temp_cuentas()

        except Exception as e:
            # Manejo de errores
            return JsonResponse({"status": "fail", "error": str(e)})

    return JsonResponse({"status": "fail", "error": "Método no permitido"}, status=405)


def update_status_cuentas_contables(request):
    try:
        cuenta_id = request.POST.get("cuenta_id")
        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_STATUS_UPDATE_CUENTAS_CONTABLES", [cuenta_id, userName]
            )
            results = cursor.fetchall()

        appConn.close()

        sync_temp_cuentas()

        datos = {"save": 1}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def insert_update_tipos_partidas(request):
    try:
        existe = 0

        partida_id = request.POST.get("partida_id")
        codigo = request.POST.get("codigo")
        nombre = request.POST.get("nombre")
        notas = request.POST.get("notas")
        diaria = request.POST.get("diaria")
        sistema = request.POST.get("sistema")
        manual = request.POST.get("manual")
        hermana = request.POST.get("hermana")
        origen = request.POST.get("origen")
        manual_automatica = request.POST.get("manual_automatica")
        opcion = request.POST.get("opcion")

        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_INSERT_UPDATE_TIPOS_PARTIDAS",
                [
                    partida_id,
                    codigo,
                    nombre,
                    notas,
                    diaria,
                    sistema,
                    manual,
                    hermana,
                    origen,
                    manual_automatica,
                    opcion,
                    userName,
                ],
            )
            results = cursor.fetchall()

            if results:
                for result in results:
                    existe = result[0]
                    lastID = result[1]
            else:
                existe = 0
                lastID = 0

        appConn.close()

        datos = {"save": 1, "existe": existe, "lastID": lastID}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def insert_update_detalles_tipos_partidas(request):
    try:
        existe = 0

        detalle_id = request.POST.get("detalle_id")
        partida_id = request.POST.get("partida_id")
        cuenta = request.POST.get("cuenta")
        tipo_movimiento = request.POST.get("tipo_movimiento")
        formula_monto = request.POST.get("formula_monto")
        valor_prueba = request.POST.get("valor_prueba")
        descripcion = request.POST.get("descripcion")
        opcion = request.POST.get("opcion")

        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_INSERT_UPDATE_DETALLE_CONFIG_PARTIDA",
                [
                    detalle_id,
                    partida_id,
                    cuenta,
                    tipo_movimiento,
                    formula_monto,
                    valor_prueba,
                    descripcion,
                    opcion,
                    userName,
                ],
            )
            results = cursor.fetchall()

            if results:
                for result in results:
                    existe = result[0]
                    lastID = result[1]
            else:
                existe = 0
                lastID = 0

        appConn.close()

        datos = {"save": 1, "existe": existe, "lastID": lastID}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def insert_update_clasificacion_cuentas(request):
    try:
        existe = 0

        clasificacion_id = request.POST.get("clasificacion_id")
        clasificacion = request.POST.get("clasificacion")
        descripcion = request.POST.get("descripcion")
        opcion = request.POST.get("opcion")

        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_INSERT_UPDATE_CLASIFICACION_CUENTA",
                [clasificacion_id, clasificacion, descripcion, opcion, userName],
            )
            results = cursor.fetchall()

            if results:
                for result in results:
                    existe = result[0]
            else:
                existe = 0

        appConn.close()

        datos = {"save": 1, "existe": existe}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def realizar_cierre_mes(request):
    try:
        empresa_id = request.POST.get("empresa_id")
        mes = request.POST.get("mes")
        anio = request.POST.get("anio")
        userName = request.session.get("userName", "CONTABLE")

        save = 0
        existe = 0

        with connections["universal"].cursor() as cursor:
            cursor.callproc(
                "CONTA_REALIZAR_CIERRE_CONTABLE_MES", [empresa_id, mes, anio, userName]
            )
            resultado = cursor.fetchone()
            if resultado:
                existe = resultado[0]  # 0 = creado, 1 = ya existía
                save = 1

        return JsonResponse({"save": save, "existe": existe})
    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)})


def actualizar_estado_cierre_contable(request):
    try:
        cierre_id = request.POST.get("id")
        nuevo_estado = request.POST.get("cerrado")
        userName = request.session.get("userName", "CONTABLE")

        with connections["universal"].cursor() as cursor:
            cursor.callproc(
                "CONTA_ACTUALIZAR_ESTADO_CIERRE_MES",
                [cierre_id, nuevo_estado, userName],
            )

        return JsonResponse({"save": 1})
    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)})


def insert_update_cuenta_nic(request):
    try:
        existe = 0

        nic_id = request.POST.get("nic_id")
        concepto = request.POST.get("concepto")
        cuenta_base = request.POST.get("cuenta_base")
        tipo_nic = request.POST.get("tipo_nic")
        opcion = request.POST.get("opcion")

        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_INSERT_UPDATE_CUENTA_NIC",
                [nic_id, concepto, cuenta_base, tipo_nic, opcion, userName],
            )
            results = cursor.fetchall()

            if results:
                for result in results:
                    existe = result[0]
            else:
                existe = 0

        appConn.close()

        datos = {"save": 1, "existe": existe}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def insert_update_cuentas_gastos(request):
    try:
        existe = 0

        cuenta_id = request.POST.get("cuenta_id")
        nombre = request.POST.get("nombre")
        codigo = request.POST.get("codigo")
        padre = request.POST.get("padre")
        empresa = request.POST.get("empresa")
        opcion = request.POST.get("opcion")

        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_INSERT_UPDATE_CUENTA_GASTO",
                [cuenta_id, nombre, codigo, padre, empresa, opcion, userName],
            )
            results = cursor.fetchall()

            if results:
                for result in results:
                    existe = result[0]
            else:
                existe = 0

        appConn.close()

        datos = {"save": 1, "existe": existe}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def update_status_detalle_tipos_partidas(request):
    try:
        detalle_id = request.POST.get("detalle_id")
        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_UPDATE_STATUS_CONFIG_DETALLE_PARTIDA", [detalle_id, userName]
            )
            results = cursor.fetchall()

        appConn.close()

        datos = {"save": 1}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def update_status_clasificacion_cuentas(request):
    try:
        clasificacion_id = request.POST.get("clasificacion_id")
        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_UPDATE_STATUS_CLASIFICACION_CUENTA", [clasificacion_id, userName]
            )
            results = cursor.fetchall()

        appConn.close()

        datos = {"save": 1}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def update_status_cuentas_nic(request):
    try:
        nic_id = request.POST.get("nic_id")
        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc("CONTA_UPDATE_STATUS_CUENTAS_NIC", [nic_id, userName])
            results = cursor.fetchall()

        appConn.close()

        datos = {"save": 1}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def firmas_view(request):
    user_id = request.session.get("user_id", "")
    firmas_CNT = request.session.get("firmasEstadosFinancieros_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if firmas_CNT == 1 or adminIT == 1:
            empresasData = obtener_empresas()

            empleadosData = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_EMPLEADOS_CARGOS", [0])
                    column_names = [desc[0] for desc in cursor.description]
                    empleadosData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                # Cierra la conexión
                udcConn.close()
            except Exception as e:
                # Manejo de excepciones, puedes personalizar esto según tus necesidades
                return JsonResponse({"error": str(e)})

            secuenciasFirmasData = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_TIPOS_FIRMAS_NUMERACION", [1])
                    column_names = [desc[0] for desc in cursor.description]
                    secuenciasFirmasData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                # Cierra la conexión
                udcConn.close()
            except Exception as e:
                # Manejo de excepciones, puedes personalizar esto según tus necesidades
                return JsonResponse({"error": str(e)})

            context = {
                "empresasData": empresasData,
                "empleadosData": empleadosData,
                "secuenciasFirmasData": secuenciasFirmasData,
            }

            return render(request, "catalogos/firmas_financieras.html", context)
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def dataFirmasFinancieras(request):
    empresa = request.POST.get("empresa")
    opcion = request.POST.get("opcion")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_FIRMAS", [empresa, opcion])
            column_names = [desc[0] for desc in cursor.description]
            firmasData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"data": firmasData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def insert_update_firma_empresa(request):
    existe = 0

    varId = request.POST.get("varId")
    varTipoFirma = request.POST.get("varTipoFirma")
    varFirmaEmpleado = request.POST.get("varFirmaEmpleado")
    varEmpresa = request.POST.get("varEmpresa")
    varOpcion = request.POST.get("opc")

    user_name = request.session.get("userName", "CONTABLE")

    try:
        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_INSERT_UPDATE_FIRMA_EMPRESA",
                [
                    varId,
                    varTipoFirma,
                    varFirmaEmpleado,
                    varEmpresa,
                    varOpcion,
                    user_name,
                ],
            )
            results = cursor.fetchall()

            if results:
                for result in results:
                    existe = result[0]
            else:
                existe = 0

        appConn.close()

        datos = {"save": 1, "existe": existe}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def update_status_firmas(request):
    try:
        firma_id = request.POST.get("firma_id")
        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc("CONTA_UPDATE_STATUS_FIRMAS", [firma_id, userName])
            results = cursor.fetchall()

        appConn.close()

        datos = {"save": 1}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def insert_new_firma(request):
    existe = 0
    ultimoId = 0
    descripcion = request.POST.get("descripcion")
    user_name = request.session.get("userName", "CONTABLE")

    try:
        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc("CONTA_INSERT_NEW_FIRMA", [descripcion, user_name])

            results = cursor.fetchall()

            if results:
                for result in results:
                    existe = result[0]
                    ultimoId = result[1]
            else:
                existe = 0

            appConn.close()

        datos = {"save": 1, "existe": existe, "ultimoId": ultimoId}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def get_tipos_firmas_numeracion(request):
    secuenciasFirmasData = ""
    try:
        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_TIPOS_FIRMAS_NUMERACION", [1])
            column_names = [desc[0] for desc in cursor.description]
            secuenciasFirmasData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        appConn.close()

    except Exception as e:
        return JsonResponse({"error": str(e)})

    return JsonResponse({"data": secuenciasFirmasData})


def conta_new_edit_partida(request, partida, empresa):
    user_id = request.session.get("user_id", "")
    libroDiario_CNT = request.session.get("libroDiario_CNT", 0)
    nuevaPartidaManual_CNT = request.session.get("nuevaPartidaManual_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if nuevaPartidaManual_CNT == 1 or libroDiario_CNT == 1 or adminIT == 1:
            empresasData = obtener_empresas()

            fecha_hora_actual = datetime.now().strftime("%Y-%m-%d")

            secuenciasFirmasData = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_TIPOS_FIRMAS_NUMERACION", [1])
                    column_names = [desc[0] for desc in cursor.description]
                    secuenciasPartidasData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                udcConn.close()
            except Exception as e:
                return JsonResponse({"error": str(e)})

            cuentasData = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, 15])
                    column_names = [desc[0] for desc in cursor.description]
                    cuentasData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                udcConn.close()
            except Exception as e:
                return JsonResponse({"error": str(e)})

            num_partida = ""
            sinopsis = ""
            tasaCambio = 0.00
            balance = 0.00
            FkSucursal = 0
            NumeroDocto = ""
            TipoDocto = ""

            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_DATA_X_PARTIDA", [partida])
                    result = [col[0] for col in cursor.description]
                    partidaQuery = [dict(zip(result, row)) for row in cursor.fetchall()]
                    cursor.nextset()

                    if partidaQuery:
                        for p in partidaQuery:
                            partida_id = p["PkEncPartida"]
                            num_partida = p["Npartida"]
                            sinopsis = p["Sinopsis"]
                            fecha_hora_actual = p["FechaPartida"].strftime("%Y-%m-%d")
                            tasaCambio = p["TasaCambio"]
                            balance = p["Balance"]
                            empresa = p["FkEmpresa"]
                            NumeroDocto = p["NumeroDocto"]
                            TipoDocto = p["TipoDocto"]
                            FkSucursal = p["FkSucursal"]
            except Exception as e:
                return JsonResponse({"error": str(e)})

            detallesPartidaData = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc(
                        "CONTA_GET_PARTIDA_DETALLE_X_FECHA_NPARTIDA", [partida]
                    )
                    column_names = [desc[0] for desc in cursor.description]
                    detallesPartidaData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                udcConn.close()
            except Exception as e:
                return JsonResponse({"error": str(e)})

            tiposPartidasData = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_TIPOS_PARTIDAS", [1])
                    column_names = [desc[0] for desc in cursor.description]
                    tiposPartidasData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                udcConn.close()
            except Exception as e:
                return JsonResponse({"error": str(e)})

            sucursalesData = ""
            try:
                udcConn = connections["global_nube"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("GS_GET_SUCURSALES_X_EMPRESA", [empresa])
                    column_names = [desc[0] for desc in cursor.description]
                    sucursalesData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                udcConn.close()
            except Exception as e:
                return JsonResponse({"error": str(e)})

            centrosData = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("VAIC_GET_CENTROS_X_EMPRESA", [empresa])
                    column_names = [desc[0] for desc in cursor.description]
                    centrosData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]
                udcConn.close()
            except Exception as e:
                return JsonResponse({"error": str(e)})

            context = {
                "partida_id": partida,
                "sinopsis": sinopsis,
                "tasaCambio": tasaCambio,
                "balance": balance,
                "empresa": empresa,
                "NumeroDocto": NumeroDocto,
                "TipoDocto": TipoDocto,
                "FkSucursal": int(FkSucursal),
                "sucursalesData": sucursalesData,
                "empresasData": empresasData,
                "secuenciasFirmasData": secuenciasFirmasData,
                "fecha_hora_actual": fecha_hora_actual,
                "tiposPartidasData": tiposPartidasData,
                "cuentasData": cuentasData,
                "detallesPartidaData": detallesPartidaData,
                "centrosData": centrosData,
            }

            return render(request, "contabilidad/partida_manual_v2.html", context)
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def insert_update_partidas_header_details(request):
    try:
        existe = 0
        varID = 0
        numPartida = ""

        partida_id = request.POST.get("partida_id")
        tipo_partida = request.POST.get("tipo_partida")
        sinopsisPartida = request.POST.get("sinopsisPartida")
        fecha_partida = request.POST.get("fecha_partida")  # 'YYYY-MM-DD'
        debe_total = request.POST.get("debe")
        haber_total = request.POST.get("haber")
        empresa = request.POST.get("empresa")
        tasa = request.POST.get("tasa")
        balance = request.POST.get("balance")
        balanceD = request.POST.get("balanceD")
        sucursal = request.POST.get("sucursal")
        referencia = request.POST.get("referencia")
        opcion = request.POST.get("opcion")

        sistema = 12
        arr_detalles = json.loads(request.POST.get("arrDetalles", "[]"))
        userName = request.session.get("userName", "CONTABLE")

        # Mes / Año desde fecha_partida
        try:
            dt = datetime.strptime(fecha_partida, "%Y-%m-%d")
            varMes = dt.month
            varAnio = dt.year
        except Exception:
            varMes = 0
            varAnio = 0

        # 1) Encabezado
        with connections["contable"].cursor() as cursor:
            cursor.callproc(
                "CONTA_INSERT_UPDATE_ENCABEZADO_PARTIDA",
                [
                    partida_id,
                    tipo_partida,
                    sinopsisPartida,
                    sistema,
                    fecha_partida,
                    tasa,
                    balance,
                    balanceD,
                    empresa,
                    userName,
                    sucursal,
                    referencia,
                    opcion,
                ],
            )
            results = cursor.fetchall()
            if results:
                for result in results:
                    existe = result[0]
                    varID = result[1]
                    numPartida = result[2]

        # 2) Borrar detalles anteriores (estado=3)
        with connections["contable"].cursor() as cursor:
            cursor.callproc("CONTA_DELETE_DETALLES_X_PARTIDA", [partida_id, userName])

        # 3) Obtener id_encabezado de gastos (una sola vez)
        with connections["contable"].cursor() as cursor:
            cursor.callproc("VAIC_GET_MAX_ID_ENCABEZADO_GASTOS", [])
            rs = cursor.fetchall()
        id_encabezado = rs[0][0] if rs else None

        # 4) Insertar cada detalle (y si aplica, su gasto vinculado)
        for detalle in arr_detalles:
            detalle_id = detalle.get("detalle_id")
            codigo = detalle.get("codigo")  # número de cuenta o id_presupuesto
            texto = detalle.get("texto")  # concepto de la cuenta
            sinopsis = detalle.get("sinopsis", "")
            debe = detalle.get("debe", 0)
            haber = detalle.get("haber", 0)
            aplica_gastos = int(detalle.get("aplica_gastos", 0) or 0)
            id_centro = int(detalle.get("id_centro", 0) or 0)

            with connections["contable"].cursor() as cursor:
                cursor.callproc(
                    "CONTA_INSERT_UPDATE_DETALLE_PARTIDA_N_GASTOS",
                    [
                        detalle_id,  # varDetalle (lo dejamos por compatibilidad)
                        varID,  # varPartida (fk encabezado)
                        numPartida,  # varNumPartida
                        fecha_partida,  # varFechaHora (usa misma fecha)
                        codigo,  # varCodigo
                        texto,  # varConcepto (nombre de cuenta)
                        sinopsis,  # varSinopsis
                        debe,  # varDebe
                        haber,  # varHaber
                        aplica_gastos,  # varAplicaGasto
                        id_centro,  # varCentro
                        id_encabezado,  # varIdEncabezado (para enlazar gastos de esta partida)
                        sinopsisPartida,  # varSinopsisPartida (concepto principal)
                        varMes,  # varMes
                        varAnio,  # varAnio
                        userName,  # userName
                        opcion,  # opc
                    ],
                )

        datos = {"save": 1, "existe": existe}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def get_tasa_cambio_x_fecha(request):
    try:
        existe = 0
        varID = 0
        tasa_cambio = 0.00

        fecha_partida = request.POST.get("fecha_partida")

        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_TASA_CAMBIO_X_FECHA", [fecha_partida])
            result = [col[0] for col in cursor.description]
            tasaCambioQuery = [dict(zip(result, row)) for row in cursor.fetchall()]
            cursor.nextset()

            if tasaCambioQuery:
                for tc in tasaCambioQuery:
                    tasa_cambio = tc["TasaCambio"]

        datos = {"save": 1, "tasa_cambio": tasa_cambio}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def conta_tasa_cambio(request):
    user_id = request.session.get("user_id", "")
    tasaCambio_CNT = request.session.get("tasaCambio_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if tasaCambio_CNT == 1 or adminIT == 1:
            empresasData = obtener_empresas()

            context = {
                "empresasData": empresasData,
            }

            return render(request, "catalogos/tasa_cambio.html", context)
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def dataTasaCambio(request):
    opcion = request.POST.get("opcion")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_TASA_CAMBIO", [opcion])
            column_names = [desc[0] for desc in cursor.description]
            tasaCambioData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"data": tasaCambioData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def insert_update_tasa_cambio(request):
    try:
        existe = 0

        tasa_id = request.POST.get("tasa_id")
        fecha = request.POST.get("fecha")
        tasaCambio = request.POST.get("tasaCambio")
        tasaCompra = request.POST.get("tasaCompra")
        tasaBanco = request.POST.get("tasaBanco")

        opcion = request.POST.get("opcion")

        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_INSERT_UPDATE_TASA_CAMBIO",
                [tasa_id, fecha, tasaCambio, tasaCompra, tasaBanco, userName, opcion],
            )
            results = cursor.fetchall()

            if results:
                for result in results:
                    existe = result[0]
            else:
                existe = 0

        appConn.close()

        datos = {"save": 1, "existe": existe}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def update_status_tasa_cambio(request):
    try:
        tasa_id = request.POST.get("tasa_id")
        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc("CONTA_STATUS_UPDATE_TASA_CAMBIO", [tasa_id, userName])
            results = cursor.fetchall()

        appConn.close()

        datos = {"save": 1}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def conta_secuencia_partidas(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    empresa = 3

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        cuentasData = []
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, 0])
                column_names = [desc[0] for desc in cursor.description]
                cuentasData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            cuentasData = "Error {}".format(str(e))

        tiposNICData = []
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_TIPOS_NIC", [1])
                column_names = [desc[0] for desc in cursor.description]
                tiposNICData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            tiposNICData = "Error {}".format(str(e))

        partidasAutomaticas = ""
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_PARTIDAS_AUTOMATICAS", [0])
                column_names = [desc[0] for desc in cursor.description]
                partidasAutomaticas = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            # Cierra la conexión
            udcConn.close()
        except Exception as e:
            # Manejo de excepciones, puedes personalizar esto según tus necesidades
            return JsonResponse({"error": str(e)})

        fecha_actual = datetime.now()
        varMonth = datetime.now().month
        varAnio = datetime.now().year

        context = {
            "empresa": empresa,
            "empresasData": empresasData,
            "partidasAutomaticas": partidasAutomaticas,
            "cuentasData": cuentasData,
            "tiposNICData": tiposNICData,
            "varMonth": varMonth,
            "varAnio": varAnio,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "auditorias/secuencia_partidas.html", context)


def conta_diferencia_cambio(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    empresa = 3

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        cuentasData = []
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, 0])
                column_names = [desc[0] for desc in cursor.description]
                cuentasData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            cuentasData = "Error {}".format(str(e))

        tiposNICData = []
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_TIPOS_NIC", [1])
                column_names = [desc[0] for desc in cursor.description]
                tiposNICData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            tiposNICData = "Error {}".format(str(e))

        fecha_actual = datetime.now()
        varMonth = datetime.now().month
        varAnio = datetime.now().year

        context = {
            "empresa": empresa,
            "empresasData": empresasData,
            "cuentasData": cuentasData,
            "tiposNICData": tiposNICData,
            "varMonth": varMonth,
            "varAnio": varAnio,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "contabilidad/diferencia_cambio.html", context)


def conta_flujo_efectivo(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    empresa = 3

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        cuentasData = []
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, 0])
                column_names = [desc[0] for desc in cursor.description]
                cuentasData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            cuentasData = "Error {}".format(str(e))

        tiposNICData = []
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_TIPOS_NIC", [1])
                column_names = [desc[0] for desc in cursor.description]
                tiposNICData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            tiposNICData = "Error {}".format(str(e))

        fecha_actual = datetime.now()
        varMonth = datetime.now().month
        varAnio = datetime.now().year

        context = {
            "empresa": empresa,
            "empresasData": empresasData,
            "cuentasData": cuentasData,
            "tiposNICData": tiposNICData,
            "varMonth": varMonth,
            "varAnio": varAnio,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "contabilidad/flujo_efectivo.html", context)


def conta_balance_general(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    empresa = 3

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        cuentasData = []
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, 0])
                column_names = [desc[0] for desc in cursor.description]
                cuentasData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            cuentasData = "Error {}".format(str(e))

        tiposNICData = []
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_TIPOS_NIC", [1])
                column_names = [desc[0] for desc in cursor.description]
                tiposNICData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            tiposNICData = "Error {}".format(str(e))

        fecha_actual = datetime.now()
        varMonth = datetime.now().month
        varAnio = datetime.now().year

        context = {
            "empresa": empresa,
            "empresasData": empresasData,
            "cuentasData": cuentasData,
            "tiposNICData": tiposNICData,
            "varMonth": varMonth,
            "varAnio": varAnio,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "contabilidad/balance_general.html", context)


def list_all_cuentas_grado(request):
    grado = (
        int(request.POST.get("grado"))
        if request.POST.get("grado") not in [None, "", "0"]
        else 15
    )
    empresa = (
        int(request.POST.get("empresa"))
        if request.POST.get("empresa") not in [None, "", "0"]
        else 3
    )

    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, grado])
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in cursor.fetchall()]

        return JsonResponse({"success": True, "data": data})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


def conta_antiguedad_saldos_clientes(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    empresa = 3

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        cuentasData = []
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, 0])
                column_names = [desc[0] for desc in cursor.description]
                cuentasData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            cuentasData = "Error {}".format(str(e))

        fecha_actual = datetime.now()
        varMonth = datetime.now().month
        varAnio = datetime.now().year

        context = {
            "empresa": empresa,
            "empresasData": empresasData,
            "cuentasData": cuentasData,
            "varMonth": varMonth,
            "varAnio": varAnio,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "cxc/antiguedad_saldos_clientes.html", context)


def dataAgingClientes(request):
    if request.method == "POST":
        empresa = request.POST.get("empresa")
        fechaCorte = request.POST.get("fechaCorte")

        rows = []
        try:
            with connections["universal"].cursor() as cursor:
                cursor.callproc("CONTA_AGING_CLIENTES", [fechaCorte, empresa])
                cols = [col[0] for col in cursor.description]
                data = cursor.fetchall()
                for r in data:
                    rows.append(dict(zip(cols, r)))
        except Exception as e:
            return JsonResponse({"save": 0, "error": str(e)})

        return JsonResponse({"save": 1, "data": rows})
    return JsonResponse({"save": 0, "error": "Método inválido"})


def conta_antiguedad_saldos_proveedores(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    empresa = 3

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        cuentasData = []
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, 0])
                column_names = [desc[0] for desc in cursor.description]
                cuentasData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            cuentasData = "Error {}".format(str(e))

        fecha_actual = datetime.now()
        varMonth = datetime.now().month
        varAnio = datetime.now().year

        context = {
            "empresa": empresa,
            "empresasData": empresasData,
            "cuentasData": cuentasData,
            "varMonth": varMonth,
            "varAnio": varAnio,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "cxp/antiguedad_saldos_proveedores.html", context)


def dataAgingProveedores(request):
    if request.method == "POST":
        empresa = request.POST.get("empresa")
        fechaCorte = request.POST.get("fechaCorte")

        try:
            with connections["universal"].cursor() as cursor:
                cursor.callproc("CONTA_AGING_PROVEEDORES", [fechaCorte, empresa])
                cols = [c[0] for c in cursor.description]
                rows = [dict(zip(cols, r)) for r in cursor.fetchall()]
            return JsonResponse({"data": rows})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


def conta_libro_compras(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    empresa = 3

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        cuentasData = []
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, 0])
                column_names = [desc[0] for desc in cursor.description]
                cuentasData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            cuentasData = "Error {}".format(str(e))

        fecha_actual = datetime.now()
        varMonth = datetime.now().month
        varAnio = datetime.now().year

        context = {
            "empresa": empresa,
            "empresasData": empresasData,
            "cuentasData": cuentasData,
            "varMonth": varMonth,
            "varAnio": varAnio,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "contabilidad/libro_compras.html", context)


def conta_libro_ventas(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    empresa = 3

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        cuentasData = []
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, 0])
                column_names = [desc[0] for desc in cursor.description]
                cuentasData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            cuentasData = "Error {}".format(str(e))

        fecha_actual = datetime.now()
        varMonth = datetime.now().month
        varAnio = datetime.now().year

        context = {
            "empresa": empresa,
            "empresasData": empresasData,
            "cuentasData": cuentasData,
            "varMonth": varMonth,
            "varAnio": varAnio,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "contabilidad/libro_ventas.html", context)


""" def dataLibroVentas(request):
    if request.method == "POST":
        empresa = request.POST.get("empresa")
        fini = request.POST.get("fechaInicial")
        ffin = request.POST.get("fechaFinal")

        data = []
        try:
            with connections['super'].cursor() as cursor:
                cursor.callproc("CONTA_LIBRO_VENTAS", [fini, ffin, empresa])
                columns = [col[0] for col in cursor.description]
                for row in cursor.fetchall():
                    data.append(dict(zip(columns, row)))
        except Exception as e:
            return JsonResponse({"error": str(e)})

        return JsonResponse({"data": data}) """


""" def dataLibroCompras(request):
    if request.method == "POST":
        empresa = request.POST.get("empresa")
        fini = request.POST.get("fechaInicial")
        ffin = request.POST.get("fechaFinal")

        data = []
        try:
            with connections['super'].cursor() as cursor:
                cursor.callproc("CONTA_LIBRO_COMPRAS", [fini, ffin, empresa])
                columns = [col[0] for col in cursor.description]
                for row in cursor.fetchall():
                    data.append(dict(zip(columns, row)))
        except Exception as e:
            return JsonResponse({"error": str(e)})

        return JsonResponse({"data": data})
 """


def dataLibroVentas(request):
    if request.method == "POST":
        empresa = request.POST.get("empresa")
        fini = request.POST.get("fechaInicial")
        ffin = request.POST.get("fechaFinal")
        user_name = request.session.get("userName")  # o el campo que uses
        npartida = 0
        opcion = 0
        visualizarTodo = 1

        facturas = []
        partidas = []
        result = []

        try:
            # 1. Traer facturas de ventas (super)
            with connections["super"].cursor() as cursor:
                cursor.callproc("CONTA_LIBRO_VENTAS", [fini, ffin, empresa])
                columns = [col[0] for col in cursor.description]
                facturas = [dict(zip(columns, row)) for row in cursor.fetchall()]

            # 2. Traer partidas del diario (contable)
            with connections["universal"].cursor() as cursor:
                cursor.callproc(
                    "CONTA_GET_PARTIDAS_X_FECHA_NPARTIDA",
                    [fini, ffin, empresa, user_name, npartida, opcion, visualizarTodo],
                )
                columns = [col[0] for col in cursor.description]
                partidas = [dict(zip(columns, row)) for row in cursor.fetchall()]

            # 3. Mapear partidas por referencia / documento
            partidas_map = {}
            for p in partidas:
                key = str(p.get("NumeroDocto") or "").strip()
                nf = key.split(" |")[0].strip()
                if nf:
                    partidas_map[nf] = p.get("Npartida")

            # 4. Unir facturas con su partida
            for f in facturas:
                ref = str(f.get("NoFactura") or "").strip()
                f["partida_generada"] = partidas_map.get(ref, None)
                result.append(f)

        except Exception as e:
            return JsonResponse({"error": str(e)})

        return JsonResponse({"data": result})


def dataLibroCompras(request):
    if request.method == "POST":
        empresa = request.POST.get("empresa")
        fini = request.POST.get("fechaInicial")
        ffin = request.POST.get("fechaFinal")
        user_name = request.session.get("userName")  # o el campo que uses
        npartida = 0
        opcion = 0
        visualizarTodo = 1

        compras = []
        partidas = []
        result = []

        try:
            # 1. Traer compras ingresadas (super)
            with connections["super"].cursor() as cursor:
                cursor.callproc("CONTA_LIBRO_COMPRAS", [fini, ffin, empresa])
                columns = [col[0] for col in cursor.description]
                compras = [dict(zip(columns, row)) for row in cursor.fetchall()]

            # 2. Traer partidas del diario (contable)
            with connections["universal"].cursor() as cursor:
                cursor.callproc(
                    "CONTA_GET_PARTIDAS_X_FECHA_NPARTIDA",
                    [fini, ffin, empresa, user_name, npartida, opcion, visualizarTodo],
                )
                columns = [col[0] for col in cursor.description]
                partidas = [dict(zip(columns, row)) for row in cursor.fetchall()]

            # 3. Mapear partidas por referencia / documento
            partidas_map = {}
            for p in partidas:
                key = str(p.get("NumeroDocto") or "").strip()
                nf = key.split(" |")[0].strip()
                if nf:
                    partidas_map[nf] = p.get("Npartida")

            # 4. Unir compras con su partida
            for f in compras:
                ref = str(f.get("NoPedido") or "").strip()
                f["partida_generada"] = partidas_map.get(ref, None)
                result.append(f)

        except Exception as e:
            return JsonResponse({"error": str(e)})

        return JsonResponse({"data": result})


def conta_auxiliar_proveedores(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    empresa = 3

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        cuentasData = []
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, 0])
                column_names = [desc[0] for desc in cursor.description]
                cuentasData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            cuentasData = "Error {}".format(str(e))

        fecha_actual = datetime.now()
        varMonth = datetime.now().month
        varAnio = datetime.now().year

        context = {
            "empresa": empresa,
            "empresasData": empresasData,
            "cuentasData": cuentasData,
            "varMonth": varMonth,
            "varAnio": varAnio,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "cxp/auxiliar_proveedores.html", context)


def dataAuxiliarProveedores(request):
    empresa = request.POST.get("empresa")
    fini = request.POST.get("fechaInicial")
    ffin = request.POST.get("fechaFinal")

    rows = []
    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc("CONTA_AUXILIAR_PROVEEDORES", [fini, ffin, empresa])
            cols = [col[0] for col in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]
    except Exception as e:
        return JsonResponse({"error": str(e), "data": []})
    return JsonResponse({"data": rows})


def conta_meses_abiertos(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    empresa = 3

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        fecha_actual = datetime.now()
        varMonth = datetime.now().month
        varAnio = datetime.now().year

        context = {
            "empresa": empresa,
            "empresasData": empresasData,
            "varMonth": varMonth,
            "varAnio": varAnio,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "auditorias/meses_abiertos.html", context)


def dataMesesAbiertos(request):
    if request.method == "POST":
        empresa = request.POST.get("empresa")
        anio = request.POST.get("anio")
        dias = request.POST.get("dias", 30)

        data = []
        try:
            with connections["contable"].cursor() as cursor:
                cursor.callproc("CONTA_MESES_ABIERTOS_ALERTA", [empresa, anio, dias])
                columns = [col[0] for col in cursor.description]
                for row in cursor.fetchall():
                    data.append(dict(zip(columns, row)))
        except Exception as e:
            return JsonResponse({"error": str(e)})

        return JsonResponse({"data": data})


def conta_asiento_cuentas_inactivas(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    empresa = 3

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        fecha_actual = datetime.now()
        varMonth = datetime.now().month
        varAnio = datetime.now().year

        context = {
            "empresa": empresa,
            "empresasData": empresasData,
            "varMonth": varMonth,
            "varAnio": varAnio,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "auditorias/asiento_cuentas_inactivas.html", context)


def dataCuentasInactivas(request):
    if request.method == "POST":
        empresa = request.POST.get("empresa")
        fini = request.POST.get("fechaInicial")
        ffin = request.POST.get("fechaFinal")

        data = []
        try:
            with connections["universal"].cursor() as cursor:
                cursor.callproc("CONTA_AUDIT_CUENTAS_INACTIVAS", [fini, ffin, empresa])
                columns = [col[0] for col in cursor.description]
                for row in cursor.fetchall():
                    data.append(dict(zip(columns, row)))
        except Exception as e:
            return JsonResponse({"error": str(e)})

        return JsonResponse({"data": data})


def conta_estado_resultado(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    empresa = 3

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        fecha_actual = datetime.now()
        varMonth = datetime.now().month
        varAnio = datetime.now().year

        context = {
            "empresa": empresa,
            "empresasData": empresasData,
            "varMonth": varMonth,
            "varAnio": varAnio,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "contabilidad/estado_resultado.html", context)


def dataEstadoResultado(request):
    empresa = request.GET.get("empresa")
    anio = request.GET.get("anio")
    mes = request.GET.get("mes")
    userName = request.session.get("userName")

    data = []
    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc("CONTA_GET_ESTADO_RESULTADO", [empresa, anio, mes])
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]
            data = rows

        return JsonResponse({"success": True, "data": data})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


def dataEstadoResultadoDetalle(request):
    empresa = request.GET.get("empresa")
    anio = request.GET.get("anio")
    mes = request.GET.get("mes")

    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc("CONTA_GET_ESTADO_RESULTADO_DETALLE", [empresa, anio, mes])
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return JsonResponse({"success": True, "data": rows})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


def conta_auxiliar_clientes(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    empresa = 3

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        cuentasData = []
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, 0])
                column_names = [desc[0] for desc in cursor.description]
                cuentasData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            cuentasData = "Error {}".format(str(e))

        fecha_actual = datetime.now()
        varMonth = datetime.now().month
        varAnio = datetime.now().year

        context = {
            "empresa": empresa,
            "empresasData": empresasData,
            "cuentasData": cuentasData,
            "varMonth": varMonth,
            "varAnio": varAnio,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "cxc/auxiliar_clientes.html", context)


def dataAuxiliarClientes(request):
    empresa_id = request.POST.get("empresa")
    fecha_inicial = request.POST.get("fechaInicial")
    fecha_final = request.POST.get("fechaFinal")

    rows = []
    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc(
                "CONTA_AUXILIAR_CLIENTES", [fecha_inicial, fecha_final, empresa_id]
            )
            columns = [col[0] for col in cursor.description]
            rows = [dict(zip(columns, r)) for r in cursor.fetchall()]

    except Exception as e:
        return JsonResponse({"data": [], "error": str(e)})

    return JsonResponse({"data": rows})


def dataDifCambioResumen(request):
    empresa = int(request.POST.get("empresa", 0))
    fechaInicial = request.POST.get("fechaInicial")
    fechaFinal = request.POST.get("fechaFinal")
    pref_ingreso = request.POST.get("prefijoIngreso", "4-09")
    pref_gasto = request.POST.get("prefijoGasto", "6-09")

    data = {"resumen": [], "detalle": []}
    try:
        with connections["universal"].cursor() as cur:
            cur.callproc(
                "CONTA_RESUMEN_DIF_CAMBIO_OPT",
                [fechaInicial, fechaFinal, empresa, pref_ingreso, pref_gasto],
            )

            cols = [c[0] for c in cur.description]
            data["resumen"] = [dict(zip(cols, r)) for r in cur.fetchall()]

            if cur.nextset() and cur.description:
                cols2 = [c[0] for c in cur.description]
                data["detalle"] = [dict(zip(cols2, r)) for r in cur.fetchall()]

        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def dataSecuenciaPartidas(request):
    empresa = int(request.POST.get("empresa", 0))
    fechaInicial = request.POST.get("fechaInicial")
    fechaFinal = request.POST.get("fechaFinal")
    tipoDocto = request.POST.get("tipoDocto", "").strip()

    out = {"duplicados": [], "huecos": []}
    try:
        with connections["universal"].cursor() as cur:
            cur.callproc(
                "CONTA_AUDIT_SECUENCIA_PARTIDAS",
                [fechaInicial, fechaFinal, empresa, tipoDocto],
            )

            cols = [c[0] for c in cur.description]
            out["duplicados"] = [dict(zip(cols, r)) for r in cur.fetchall()]

            if cur.nextset() and cur.description:
                cols2 = [c[0] for c in cur.description]
                out["huecos"] = [dict(zip(cols2, r)) for r in cur.fetchall()]

        return JsonResponse(out)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def dataBalanceGeneral(request):
    empresa = int(request.GET.get("empresa", 0))
    anio = int(request.GET.get("anio"))
    mes = int(request.GET.get("mes"))
    tipoNIC = int(request.GET.get("tipoNIC", 2))  # 1=NIC, 2=NIC2010
    util = int(request.GET.get("utilidad", 0))  # 0/1

    f_ini = date(anio, mes, 1).strftime("%Y-%m-%d")
    f_fin = date(anio, mes, monthrange(anio, mes)[1]).strftime("%Y-%m-%d")

    out = {"success": False, "resumen": [], "detalle": [], "totales": [], "sin_map": []}

    try:
        with connections["universal"].cursor() as cur:
            cur.callproc(
                "CONTA_EST_BG_NIC_DETALLES_SALDO_MANUAL",
                [f_ini, f_fin, empresa, tipoNIC, util],
            )

            # 1) Resumen (Rubro, Subrubro, Saldo)
            cols = [c[0] for c in cur.description]
            out["resumen"] = [dict(zip(cols, r)) for r in cur.fetchall()]

            # 2) Detalle (Rubro, Subrubro, CuentaMayor, NombreMayor, Saldo)
            if cur.nextset() and cur.description:
                cols2 = [c[0] for c in cur.description]
                out["detalle"] = [dict(zip(cols2, r)) for r in cur.fetchall()]

            # (Opcional) 3) Totales por Rubro (si agregas 2º result set de totales en tu SP)
            if cur.nextset() and cur.description:
                cols3 = [c[0] for c in cur.description]
                out["totales"] = [dict(zip(cols3, r)) for r in cur.fetchall()]

            # (Opcional) 4) Sin NIC mapeo (si lo agregas)
            if cur.nextset() and cur.description:
                cols4 = [c[0] for c in cur.description]
                out["sin_map"] = [dict(zip(cols4, r)) for r in cur.fetchall()]

        out["success"] = True
    except Exception as e:
        out["error"] = str(e)

    return JsonResponse(out)


def guardarSaldoManualBG(request):
    out = {"success": False}

    try:
        empresa = int(request.GET.get("empresa", 0))
        anio = int(request.GET.get("anio", 0))
        mes = int(request.GET.get("mes", 0))
        cuenta = (request.GET.get("cuenta", "") or "").strip()
        saldo = request.GET.get("saldo", "0")
        obs = (request.GET.get("obs", "") or "").strip()
        userName = request.session.get("userName", "")

        # Validaciones mínimas
        if empresa <= 0 or anio <= 0 or mes <= 0 or mes > 12:
            out["error"] = "Parámetros inválidos."
            return JsonResponse(out)

        if not cuenta:
            out["error"] = "Cuenta requerida."
            return JsonResponse(out)

        # 🔒 SOLO grado 5 (4 guiones)
        if cuenta.count("-") != 4:
            out["error"] = (
                "Solo se permite registrar saldo manual en cuentas de grado 5."
            )
            return JsonResponse(out)

        try:
            saldo = float(str(saldo).replace(",", ""))
        except:
            saldo = 0.0

        with connections["contable"].cursor() as cur:
            cur.callproc(
                "CONTA_BG_GUARDAR_SALDO_MANUAL",
                [empresa, anio, mes, cuenta, saldo, obs, userName],
            )
            cols = [c[0] for c in cur.description]
            row = cur.fetchone()
            if row:
                out.update(dict(zip(cols, row)))

        out["success"] = True
        return JsonResponse(out)

    except Exception as e:
        out["error"] = str(e)
        return JsonResponse(out)


def conta_balanza_de_comprobacion(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    empresa = 3

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        cuentasData = ""
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, 0])
                column_names = [desc[0] for desc in cursor.description]
                cuentasData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            # Cierra la conexión
            udcConn.close()
        except Exception as e:
            # Manejo de excepciones, puedes personalizar esto según tus necesidades
            return JsonResponse({"error": str(e)})

        context = {
            "empresa": empresa,
            "empresasData": empresasData,
            "cuentasData": cuentasData,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "contabilidad/balanza_comprobacion.html", context)


def dataBalanzaComprobacion(request):
    fkEmpresa = request.POST.get("fkEmpresa")
    fecha_inicio = request.POST.get("fecha_inicio")  # 'YYYY-MM-DD'
    fecha_fin = request.POST.get("fecha_fin")  # 'YYYY-MM-DD'
    codigo_filtro = request.POST.get("codigo_filtro", "")  # opcional
    modo_busqueda = int(request.POST.get("modo_busqueda", 2))  # 1 exacto, 2 prefijo

    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_BALANZA_COMPROBACION_v1",
                [fkEmpresa, fecha_inicio, fecha_fin, codigo_filtro, modo_busqueda],
            )
            column_names = [desc[0] for desc in cursor.description]
            balanzaData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        return JsonResponse({"data": balanzaData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def conta_libro_mayor(request, empresa):
    user_id = request.session.get("user_id", "")
    libroMayor_CNT = request.session.get("libroMayor_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if libroMayor_CNT == 1 or adminIT == 1:
            empresasData = obtener_empresas()

            cuentasData = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, 0])
                    column_names = [desc[0] for desc in cursor.description]
                    cuentasData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                # Cierra la conexión
                udcConn.close()
            except Exception as e:
                # Manejo de excepciones, puedes personalizar esto según tus necesidades
                return JsonResponse({"error": str(e)})

            context = {
                "empresa": empresa,
                "empresasData": empresasData,
                "cuentasData": cuentasData,
                "date1": datetime.now().strftime("%Y-%m-01"),
                "date2": datetime.now().strftime("%Y-%m-%d"),
            }

            return render(request, "contabilidad/libro_mayor.html", context)
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def conta_auxiliar_cuentas(request, empresa):
    user_id = request.session.get("user_id", "")
    libroMayor_CNT = request.session.get("libroMayor_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if libroMayor_CNT == 1 or adminIT == 1:
            empresasData = obtener_empresas()

            cuentasData = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_CAT_CUENTAS_X_GRADO", [empresa, 0])
                    column_names = [desc[0] for desc in cursor.description]
                    cuentasData = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                # Cierra la conexión
                udcConn.close()
            except Exception as e:
                # Manejo de excepciones, puedes personalizar esto según tus necesidades
                return JsonResponse({"error": str(e)})

            context = {
                "empresa": empresa,
                "empresasData": empresasData,
                "cuentasData": cuentasData,
                "date1": datetime.now().strftime("%Y-%m-01"),
                "date2": datetime.now().strftime("%Y-%m-%d"),
            }

            return render(request, "contabilidad/auxiliar_cuentas.html", context)
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def dataSolicitudIngreso(request):
    id_gasto = request.POST.get("id_gasto")
    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc("VAIC_PDF_INGRESOS_DETAILS", [id_gasto])
            column_names = [desc[0] for desc in cursor.description]
            solicitudData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        return JsonResponse({"data": solicitudData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataSolicitudGasto(request):
    id_gasto = request.POST.get("id_gasto")
    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc("VAIC_PDF_GASTOS_DETAILS", [id_gasto])
            column_names = [desc[0] for desc in cursor.description]
            solicitudData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        return JsonResponse({"data": solicitudData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataLibroMayor(request):
    codigo = request.POST.get("codigo")
    date1 = request.POST.get("date1")
    date2 = request.POST.get("date2")
    empresa = request.POST.get("empresa")
    modo = request.POST.get("modo_busqueda")

    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_LIBRO_MAYOR_v1", [date1, date2, codigo, empresa, modo]
            )
            column_names = [desc[0] for desc in cursor.description]
            libroMayorData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        return JsonResponse({"data": libroMayorData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def conta_entregar_periodos(request, empresa):
    user_id = request.session.get("user_id", "")
    entregarPeriodo_CNT = request.session.get("entregarPeriodo_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    MesInicial = ""
    AnioInicial = ""
    MesFinal = ""
    AnioFinal = ""
    mesInicial_Text = ""
    mesFinal_Text = ""
    fechaPartida = ""

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if entregarPeriodo_CNT == 1 or adminIT == 1:
            empresasData = obtener_empresas()

            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_PERIODOS_ENTREGADOS", [empresa])
                    column_names = [desc[0] for desc in cursor.description]
                    periodosEntregadosQuery = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                    if periodosEntregadosQuery:
                        for p in periodosEntregadosQuery:

                            MesInicial = p["MesInicial"]
                            AnioInicial = p["AnioInicial"]
                            MesFinal = p["MesFinal"]
                            AnioFinal = p["AnioFinal"]
                            mesInicial_Text = p["mesInicial_Text"]
                            mesFinal_Text = p["mesFinal_Text"]
                            fechaPartida = p["fechaPartida"].strftime("%Y-%m-%d")

                # Cierra la conexión
                udcConn.close()
            except Exception as e:
                # Manejo de excepciones, puedes personalizar esto según tus necesidades
                return JsonResponse({"error": str(e)})

            context = {
                "empresa": empresa,
                "empresasData": empresasData,
                "MesInicial": MesInicial,
                "AnioInicial": AnioInicial,
                "MesFinal": MesFinal,
                "AnioFinal": AnioFinal,
                "mesInicial_Text": mesInicial_Text,
                "mesFinal_Text": mesFinal_Text,
                "fechaPartida": fechaPartida,
                "date1": datetime.now().strftime("%Y-%m-01"),
                "date2": datetime.now().strftime("%Y-%m-%d"),
            }

            return render(request, "contabilidad/entregar_periodos.html", context)
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def autorizar_entregar_periodos(request):
    try:
        empresa = request.session.get("empresa", "")

        mes_inicio = request.POST.get("mes_inicio")
        anio_inicio = request.session.get("anio_inicio", "")

        mes_final = request.session.get("mes_final", "")
        anio_final = request.session.get("anio_final", "")

        mesValidar = 0
        anioValidar = 0

        userName = request.session.get("userName", "CONTABLE")

        mensaje = ""
        msjValue = 0

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_AUTORIZAR_PERIODO_LOOP", [mes_inicio, anio_final, empresa]
            )
            column_names = [desc[0] for desc in cursor.description]
            autorizarPeriodoQuery = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

            if autorizarPeriodoQuery:
                for p in autorizarPeriodoQuery:
                    mensaje = p["mensaje"]
                    msjValue = p["msjValue"]
                    mesValidar = p["mesValidar"]
                    anioValidar = p["anioValidar"]

        appConn.close()

        datos = {"save": 1, "mensaje": mensaje, "msjValue": msjValue}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def generar_partida_entregar_periodo(request):
    try:
        empresa = request.session.get("empresa", "")

        mes_inicio = request.POST.get("mes_inicio")
        anio_inicio = request.session.get("anio_inicio", "")

        mes_final = request.session.get("mes_final", "")
        anio_final = request.session.get("anio_final", "")

        mesValidar = request.session.get("mesValidar", "")
        anioValidar = request.session.get("anioValidar", "")

        userName = request.session.get("userName", "CONTABLE")

        mensaje = ""
        msjValue = 0

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GENERAR_PARTIDA_ENTREGAR_PERIODO",
                [
                    mes_inicio,
                    anio_final,
                    mes_final,
                    anio_final,
                    mesValidar,
                    anioValidar,
                    empresa,
                    userName,
                ],
            )
            column_names = [desc[0] for desc in cursor.description]
            autorizarPeriodoQuery = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        appConn.close()

        datos = {"save": 1, "mensaje": mensaje, "msjValue": msjValue}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def get_meses_cerrados_estado_cerrado(request):

    idEmpresa = request.POST.get("empresa")
    mesesCerradosData = ""
    try:
        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_MESES_CERRADOS_ESTADO_CERRADO", [idEmpresa])
            column_names = [desc[0] for desc in cursor.description]
            mesesCerradosData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        appConn.close()

    except Exception as e:
        return JsonResponse({"error": str(e)})

    return JsonResponse({"data": mesesCerradosData})


def obtener_datos_para_cierre_mes(request):

    fecha = datetime.strptime(request.POST.get("varFecha"), "%Y-%m-%d")
    varfechaFinal = fecha.strftime("%Y-%m-%d %H:%M:%S")

    varEmpresa = request.POST.get("varEmpresa")
    varMaxMes = request.POST.get("varMaxMes")
    varMaxAnio = request.POST.get("varMaxAnio")
    varLastMes = request.POST.get("varLastMes")
    varLastAnio = request.POST.get("varLastAnio")

    datosCierreMes = ""
    try:
        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_OBTENER_DATOS_PARA_CIERRE_MES",
                [
                    varfechaFinal,
                    varEmpresa,
                    varMaxMes,
                    varMaxAnio,
                    varLastMes,
                    varLastAnio,
                ],
            )
            column_names = [desc[0] for desc in cursor.description]
            datosCierreMes = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        appConn.close()

    except Exception as e:
        return JsonResponse({"error": str(e)})

    return JsonResponse({"data": datosCierreMes})


def obtener_nombre_mes(request):

    varNumeroMes = request.POST.get("mes")

    nombreMesData = ""
    try:
        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_NOMBRE_MES", [varNumeroMes])
            column_names = [desc[0] for desc in cursor.description]
            nombreMesData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        appConn.close()

    except Exception as e:
        return JsonResponse({"error": str(e)})

    return JsonResponse({"data": nombreMesData})


def get_ultimo_mes_anio_from_meses_cerrados(request):

    idEmpresa = request.POST.get("empresa")
    estaCerrado = request.POST.get("estaCerrado")
    mesesCerradosData = ""
    try:
        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_ULTIMO_MES_ANIO_FROM_CONTA_MESES_CERRADOS",
                [idEmpresa, estaCerrado],
            )
            column_names = [desc[0] for desc in cursor.description]
            mesesCerradosData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        appConn.close()

    except Exception as e:
        return JsonResponse({"error": str(e)})

    return JsonResponse({"data": mesesCerradosData})


# Datos Iniciales de Cierre/Entrega de Mes
def conta_cierre_entrega_mes(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)
    ciereEntregaMes_CNT = 0

    mes_actual = datetime.now().month

    ciereEntregaMes_CNT = request.session.get("cierreMes_CNT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if ciereEntregaMes_CNT == 1 or adminIT == 1:
            empresasData = obtener_empresas()

            context = {
                "empresasData": empresasData,
                "mes": mes_actual,
            }

            return render(request, "contabilidad/cierre_de_meses.html", context)
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def get_codigo_cuenta_x_fecha(request):
    idEmpresa = request.POST.get("idEmpresa")
    mes = request.POST.get("mes")
    anio = request.POST.get("anio")
    opc = request.POST.get("opc")

    codigoCuentaData = ""
    try:
        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_CODIGO_CUENTA_X_FECHA", [idEmpresa, mes, anio, opc]
            )
            column_names = [desc[0] for desc in cursor.description]
            codigoCuentaData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        appConn.close()

    except Exception as e:
        return JsonResponse({"error": str(e)})

    return JsonResponse({"data": codigoCuentaData})


def insert_cierre_cuenta_mes(request):
    try:

        mes = request.POST.get("mes")
        anio = request.POST.get("anio")
        codigoCuenta = request.POST.get("codigoCuenta")
        saldoInicial = request.POST.get("saldoInicial")
        debe = request.POST.get("debe")
        haber = request.POST.get("haber")
        saldoFinal = request.POST.get("saldoFinal")
        saldoIniciaD = request.POST.get("saldoInicialD")
        debeD = request.POST.get("debeD")
        haberD = request.POST.get("haberD")
        saldoFinalD = request.POST.get("saldoFinalD")
        idEmpresa = request.POST.get("idEmpresa")

        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_INSERT_CIERRE_CUENTA_MES",
                [
                    mes,
                    anio,
                    codigoCuenta,
                    saldoInicial,
                    debe,
                    haber,
                    saldoFinal,
                    saldoIniciaD,
                    debeD,
                    haberD,
                    saldoFinalD,
                    idEmpresa,
                    userName,
                ],
            )
            results = cursor.fetchall()

        appConn.close()

        datos = {"save": 1, "result": results}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def get_saldos_cuenta_x_nic(request):
    idEmpresa = request.POST.get("empresa")
    mes = request.POST.get("mes")
    anio = request.POST.get("anio")
    ultimoMes = request.POST.get("ultimoMes")
    ultimoAnio = request.POST.get("ultimoAnio")

    saldoData = ""
    try:
        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_SALDOS_CUENTA_X_NIC",
                [idEmpresa, mes, anio, ultimoMes, ultimoAnio],
            )
            column_names = [desc[0] for desc in cursor.description]
            saldoData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        appConn.close()

    except Exception as e:
        return JsonResponse({"error": str(e)})

    return JsonResponse({"data": saldoData})


def get_datos_cierre_anterior(request):
    ultimoMes = request.POST.get("ultimoMes")
    ultimoAnio = request.POST.get("ultimoAnio")
    mes = request.POST.get("mes")
    anio = request.POST.get("anio")
    idEmpresa = request.POST.get("idEmpresa")

    cierreAnteriorData = ""
    try:
        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_DATOS_CIERRE_ANTERIOR",
                [ultimoMes, ultimoAnio, mes, anio, idEmpresa],
            )
            column_names = [desc[0] for desc in cursor.description]
            cierreAnteriorData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        appConn.close()

    except Exception as e:
        return JsonResponse({"error": str(e)})

    return JsonResponse({"data": cierreAnteriorData})


def anular_cierre_cuenta_mes(request):
    try:

        idEmpresa = request.POST.get("empresa")
        mes = request.POST.get("mes")
        anio = request.POST.get("anio")

        user_name = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_ANULAR_CIERRE_CUENTA_MES", [user_name, mes, anio, idEmpresa]
            )
            results = cursor.fetchall()

        appConn.close()

        datos = {"save": 1, "result": results}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def dictfetchall(cursor):
    """
    Retorna los resultados del cursor como una lista de diccionarios,
    normalizando tipos no serializables para JSON:
    - BIT (b'\\x00', b'\\x01') -> int (0 / 1)
    - Decimal -> float
    - datetime / date -> string ISO
    """

    columns = [col[0] for col in cursor.description]
    results = []

    for row in cursor.fetchall():
        row_dict = {}

        for col, val in zip(columns, row):

            # BIT / BINARY (b'\x00', b'\x01')
            if isinstance(val, (bytes, bytearray)):
                row_dict[col] = int.from_bytes(val, byteorder="big")

            # Decimal
            elif isinstance(val, Decimal):
                row_dict[col] = float(val)

            # datetime / date
            elif isinstance(val, (datetime, date)):
                row_dict[col] = val.isoformat()

            else:
                row_dict[col] = val

        results.append(row_dict)

    return results


def conta_saldos_iniciales_lotes(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    mes_actual = datetime.now().month

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        context = {
            "empresasData": empresasData,
            "mes": mes_actual,
            "date1": datetime.now().replace(day=1).strftime("%Y-%m-%d"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(
            request, "contabilidad/conta_saldos_iniciales_lotes.html", context
        )


def conta_lotes_listar(request):
    fkEmpresa = request.GET.get("fkEmpresa")

    with connections["contable"].cursor() as cursor:
        cursor.callproc("CONTA_LOTES_LISTAR", [fkEmpresa])
        data = dictfetchall(cursor)

    return JsonResponse({"data": data})


def conta_lote_nuevo(request):
    fkEmpresa = request.POST.get("fkEmpresa")
    fecha_corte = request.POST.get("fecha_corte")

    with connections["contable"].cursor() as cursor:
        cursor.callproc("CONTA_LOTE_NUEVO", [fkEmpresa, fecha_corte])
        row = cursor.fetchone()

    return JsonResponse({"id_lote": row[0], "fecha_corte": row[1]})


def conta_saldo_lote_detalle(request):
    id_lote = request.GET.get("id_lote")
    fkEmpresa = request.GET.get("fkEmpresa")

    with connections["contable"].cursor() as cursor:
        cursor.callproc("CONTA_SALDO_LOTE_DETALLE", [id_lote, fkEmpresa])
        data = dictfetchall(cursor)

    return JsonResponse({"data": data})


def conta_saldo_lote_editar(request):
    id_detalle = request.POST.get("id_detalle_saldo")
    codigo = request.POST.get("codigo")
    saldo = request.POST.get("saldo")
    userName = request.session.get("userName")

    with connections["contable"].cursor() as cursor:
        cursor.callproc(
            "CONTA_SALDO_LOTE_EDITAR", [id_detalle, codigo, saldo, userName]
        )

    return JsonResponse({"success": True})


def conta_saldo_lote_eliminar(request):
    id_detalle = request.POST.get("id_detalle_saldo")

    with connections["contable"].cursor() as cursor:
        cursor.callproc("CONTA_SALDO_LOTE_ELIMINAR", [id_detalle])

    return JsonResponse({"success": True})


def conta_saldo_lote_procesar(request):
    """
    Procesa un lote de saldos iniciales.
    Controla errores del SP y devuelve JSON limpio al frontend.
    """

    # ==========================
    # DATOS DE ENTRADA
    # ==========================
    id_lote = request.POST.get("id_lote")
    fkEmpresa = request.POST.get("fkEmpresa")
    userName = request.session.get("userName")

    # ==========================
    # VALIDACIONES BÁSICAS
    # ==========================
    if not id_lote or not fkEmpresa:
        return JsonResponse(
            {"success": False, "error": "Datos incompletos para procesar el lote."}
        )

    if not userName:
        return JsonResponse(
            {"success": False, "error": "Sesión expirada. Vuelva a iniciar sesión."}
        )

    # ==========================
    # EJECUCIÓN DEL SP
    # ==========================
    try:
        with connections["contable"].cursor() as cursor:
            cursor.callproc("CONTA_SALDO_LOTE_PROCESAR", [id_lote, fkEmpresa, userName])

        return JsonResponse(
            {"success": True, "message": "Lote procesado correctamente."}
        )

    # ==========================
    # ERROR CONTROLADO DEL SP
    # ==========================
    except OperationalError as e:
        """
        Captura errores lanzados por SIGNAL en MySQL
        Ej: SIGNAL SQLSTATE '45000'
        """

        error_msg = str(e)

        # Limpieza del mensaje (opcional pero recomendado)
        if "Existen cuentas que no estan en el Catalogo Contable" in error_msg:
            mensaje = "Existen Cuentas Que No Están En El Catálogo Contable."
        else:
            mensaje = "Error Al Procesar El Lote Contable."

        return JsonResponse({"success": False, "error": mensaje})

    # ==========================
    # CUALQUIER OTRO ERROR
    # ==========================
    except Exception as e:
        return JsonResponse(
            {
                "success": False,
                "error": "Error Interno Del Servidor.",
                "detalle": str(e),
            }
        )


def conta_saldos_iniciales_descargar_plantilla(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "SALDOS_INICIALES"

    headers = ["codigo", "saldo"]
    ws.append(headers)

    # Estilo header
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Ancho columnas
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        "attachment; filename=plantilla_saldos_iniciales.xlsx"
    )

    wb.save(response)
    return response


def conta_saldo_lote_upload_excel(request):
    id_lote = request.POST.get("id_lote")
    fkEmpresa = request.POST.get("fkEmpresa")
    fecha_corte = request.POST.get("fecha_corte")
    userName = request.session.get("userName")

    archivo = request.FILES.get("excel")

    if not archivo:
        return JsonResponse({"error": "Archivo no recibido"}, status=400)

    wb = openpyxl.load_workbook(archivo)
    ws = wb.active

    with connections["contable"].cursor() as cursor:

        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):

            codigo = row[0].value
            saldo = row[1].value or 0

            if not codigo:
                continue

            cursor.callproc(
                "CONTA_SALDO_LOTE_UPSERT",
                [
                    id_lote,
                    fkEmpresa,
                    fecha_corte,
                    str(codigo).strip(),
                    float(saldo),
                    userName,
                ],
            )

    return JsonResponse({"success": True})


def conta_calendario_contable(request):
    user_id = request.session.get("user_id", "")
    adminIT = request.session.get("contabilidadAdminIT", 0)

    mes_actual = datetime.now().month

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        empresasData = obtener_empresas()

        context = {
            "empresasData": empresasData,
            "mes": mes_actual,
        }

        return render(request, "contabilidad/calendario_contable.html", context)


def load_calendario_contable(request):
    empresa = request.POST.get("empresa")
    anio = request.POST.get("anio")
    user = request.session.get("userName", "admin")

    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc("CONTA_INIT_CALENDARIO_CONTABLE", [empresa, anio, user])
            columns = [col[0] for col in cursor.description]
            rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
            return JsonResponse({"data": rows})
    except Exception as e:
        return JsonResponse({"data": [], "error": str(e)})


def guardar_fecha_calendario(request):
    empresa = request.POST.get("empresa")
    anio = request.POST.get("anio")
    mes = request.POST.get("mes")
    fecha = request.POST.get("fecha")
    user = request.session.get("userName", "admin")

    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc(
                "CONTA_SET_CALENDARIO_CONTABLE", [empresa, anio, mes, fecha, user]
            )
        return JsonResponse({"save": 1})
    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)})


def conta_actualizar_datos_mes(request):
    user_id = request.session.get("user_id", "")
    actualizarMes_CNT = request.session.get("actualizarDatosMesesCerrados_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    mes_actual = datetime.now().month

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if actualizarMes_CNT == 1 or adminIT == 1:
            empresasData = obtener_empresas()

            context = {
                "empresasData": empresasData,
                "mes": mes_actual,
            }

            return render(
                request, "contabilidad/actualizar_saldos_de_mes.html", context
            )
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def actualizar_saldos_contables_mes(request):
    empresa = request.POST.get("empresa")
    mes = request.POST.get("mes")
    anio = request.POST.get("anio")
    user = request.session.get("userName", "admin")

    periodo_cerrado = 0

    try:
        with connections["universal"].cursor() as cursor:
            # Validar si ya está cerrado
            cursor.execute(
                "SELECT COUNT(*) FROM conta_cierres_contables WHERE fkempresa = %s AND mes = %s AND anio = %s  AND cerrado = 1 AND estado <> 3",
                [empresa, mes, anio],
            )
            cerrado = cursor.fetchone()[0]

            if cerrado > 0:
                periodo_cerrado = 1
            else:
                # Si no está cerrado, continuar con el cierre por día y mes
                fecha_inicio = f"{anio}-{int(mes):02d}-01"
                cursor.callproc(
                    "CONTA_GENERAR_CIERRE_X_DIA_RANGO", [fecha_inicio, empresa, user]
                )
                cursor.callproc(
                    "CONTA_GENERAR_CIERRE_X_MES",
                    [int(anio), int(mes), int(empresa), user],
                )

        return JsonResponse({"save": 1, "periodo_cerrado": periodo_cerrado})

    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)})


def get_saldos_contables_mes(request):
    empresa = request.POST.get("empresa")
    mes = request.POST.get("mes")
    anio = request.POST.get("anio")

    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc("CONTA_GET_SALDOS_CONTABLES_MENSUAL", [empresa, mes, anio])
            columns = [col[0] for col in cursor.description]
            rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        return JsonResponse({"data": rows})
    except Exception as e:
        return JsonResponse({"data": [], "error": str(e)})


def update_meses_cerrados(request):
    try:
        varMes = request.POST.get("mes")
        varAnio = request.POST.get("anio")
        varEmpresa = request.POST.get("empresa")

        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_UPDATE_MESES_CERRADOS", [varMes, varAnio, varEmpresa, userName]
            )
            results = cursor.fetchall()

        appConn.close()

        datos = {"save": 1, "result": results}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def exportMovimientosDestinosProcedenciasExcel(request):
    empresa = request.POST.get("empresa")
    tipo = request.POST.get("tipo_enlance")
    estado = request.POST.get("estado")
    userName = request.session.get("userName", "")

    tipo_txt = "DESTINOS" if int(tipo) == 1 else "PROCEDENCIAS"

    conn = connections["bankConn"]

    # ================= MOVIMIENTOS =================
    with conn.cursor() as cursor:

        cursor.callproc(
            "CONTA_GET_DESTINOS__PROCEDENCIAS_SIN_ENLAZAR", [tipo, estado, empresa]
        )

        columnas = [col[0] for col in cursor.description]
        filas_raw = cursor.fetchall()

        movimientos = [dict(zip(columnas, f)) for f in filas_raw]

    # ================= ENLACES =================
    with conn.cursor() as cursor:

        cursor.callproc("CONTA_GET_ENLANCES_DESTINOS_PROCEDENCIAS", [tipo, 1])

        columnas_e = [col[0] for col in cursor.description]
        filas_e = cursor.fetchall()

        enlaces = [dict(zip(columnas_e, f)) for f in filas_e]

    # ================= EXCEL =================
    wb = Workbook()

    # =====================================================
    # HOJA 1 - MOVIMIENTOS (ÚNICA EDITABLE)
    # =====================================================
    ws = wb.active
    ws.title = "Movimientos"

    headers_mov = [
        "pk_registro",
        "FechaMovimientoText",
        "NumeroMovimiento",
        "Concepto",
        "CuentaBancaria",
        "TipoCuenta",
        "Banco",
        "valor_movimiento",
        "FechaHora",
        "ElaboradoPor",
        "id_enlace_asignar",
    ]

    ws.append(headers_mov)

    for m in movimientos:
        ws.append(
            [
                m.get("pk_registro"),
                m.get("FechaMovimientoText"),
                m.get("NumeroMovimiento"),
                m.get("ConceptoTrim"),
                m.get("CuentaBancaria"),
                m.get("TipoCuenta"),
                m.get("Banco"),
                m.get("valor_movimiento"),
                m.get("FechaHora"),
                m.get("ElaboradoPor"),
                m.get("id_enlace"),
            ]
        )

    # =====================================================
    # HOJA 2 - ENLACES (SOLO LECTURA)
    # =====================================================
    ws2 = wb.create_sheet("Enlaces")

    headers_enl = ["id_enlace", "nombre_enlace", "cuenta_contable", "tipo_enlance"]

    ws2.append(headers_enl)

    for e in enlaces:
        ws2.append(
            [
                e.get("id_enlace"),
                e.get("nombre_enlace"),
                e.get("cuenta_contable"),
                e.get("tipo_enlance"),
            ]
        )

    # =====================================================
    # HOJA 3 - CONTROL (SOLO LECTURA)
    # =====================================================
    wsC = wb.create_sheet("Control")

    wsC.append(["clave", "valor"])
    wsC.append(["tipo", tipo])
    wsC.append(["descripcion", tipo_txt])
    wsC.append(["empresa", empresa])
    wsC.append(["usuario", userName])
    wsC.append(["generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wsC.append(["identificador", f"CARGA_{tipo_txt}_{empresa}"])

    # =====================================================
    # FORMATO VISUAL
    # =====================================================
    bold = Font(bold=True)

    # ----- Hoja 1 -----
    for col in range(1, len(headers_mov) + 1):
        c = ws.cell(row=1, column=col)
        c.font = bold
        c.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    col_concepto = 4

    for column_cells in ws.columns:

        col_index = column_cells[0].column
        letra = get_column_letter(col_index)

        if col_index == col_concepto:
            ws.column_dimensions[letra].width = 43

            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            ws.column_dimensions[letra].width = 22

    # ----- Hoja 2 -----
    for col in range(1, len(headers_enl) + 1):
        c = ws2.cell(row=1, column=col)
        c.font = bold
        c.alignment = Alignment(horizontal="center")

    for column_cells in ws2.columns:
        letra = get_column_letter(column_cells[0].column)
        ws2.column_dimensions[letra].width = 22

    # =====================================================
    # 🔒 PROTECCIÓN DE HOJAS
    # =====================================================

    # Movimientos → editable
    ws.protection.sheet = False

    # Enlaces → solo lectura
    ws2.protection.set_password("nexapps")
    ws2.protection.enable()

    # Control → solo lectura
    wsC.protection.set_password("nexapps")
    wsC.protection.enable()

    # =====================================================
    # RESPUESTA
    # =====================================================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="Carga_{tipo_txt}_{empresa}.xlsx"'
    )

    wb.save(response)

    return response


def cargarExcelEnlacesDestinosProcedencias(request):
    archivo = request.FILES.get("file")
    tipo_actual = int(request.POST.get("tipo_enlance"))
    userName = request.session.get("userName", "")

    wb = openpyxl.load_workbook(archivo)

    # ============== VALIDAR CONTROL ==============
    if "Control" not in wb.sheetnames:
        return JsonResponse({"error": "Archivo no válido: falta hoja Control"})

    wsC = wb["Control"]

    data_control = {row[0]: row[1] for row in wsC.iter_rows(values_only=True) if row[0]}

    tipo_excel = int(data_control.get("tipo", 0))

    if tipo_excel != tipo_actual:
        return JsonResponse(
            {
                "error": f'Este Excel es de {data_control.get("descripcion")} '
                f"y estás en otro módulo"
            }
        )

    # ============== LEER MOVIMIENTOS ==============
    ws = wb["Movimientos"]

    asignados = 0
    sin_asignar = 0
    no_encontrados = 0

    conn = connections["bankConn"]

    for row in ws.iter_rows(min_row=2, values_only=True):

        pk_registro = row[0]
        id_enlace = row[10]

        if not pk_registro:
            continue

        if not id_enlace or int(id_enlace) == 0:
            sin_asignar += 1
            continue

        try:
            with conn.cursor() as cursor:

                cursor.callproc(
                    "CONTA_SET_ENLACE_MASIVO",
                    [tipo_actual, pk_registro, id_enlace, userName],
                )

                asignados += 1

        except Exception:
            no_encontrados += 1

    return JsonResponse(
        {
            "success": True,
            "mensaje": f"""
        ✔ Asignados: {asignados}
        ⚠ Vacíos: {sin_asignar}
        ❌ No encontrados: {no_encontrados}
        """,
        }
    )


def insert_update_meses_cerrados(request):
    try:
        varMes = request.POST.get("mes")
        varAnio = request.POST.get("anio")
        varEmpresa = request.POST.get("empresa")
        varId = request.POST.get("id")
        varOpcion = request.POST.get("opc")

        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_INSERT_UPDATE_MESES_CERRADOS",
                [varMes, varAnio, userName, varEmpresa, varId, varOpcion],
            )
            results = cursor.fetchall()

        appConn.close()

        datos = {"save": 1, "result": results}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def update_status_meses_cerrados(request):
    try:
        mes_id = request.POST.get("mes_id")
        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc("CONTA_UPDATE_STATUS_MESES_CERRADOS", [mes_id, userName])
            results = cursor.fetchall()

        appConn.close()

        datos = {"save": 1, "result": results}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def dataMesesCerradosEntregados(request):
    empresa = request.POST.get("empresa")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_MESES_CERRADOS_ENTREGADOS", [empresa])
            column_names = [desc[0] for desc in cursor.description]
            mesesCerradosEntregadosData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"data": mesesCerradosEntregadosData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def conta_meses_cerrados_entregados(request):
    user_id = request.session.get("user_id", "")
    mesesCerradosEntregados_CNT = request.session.get(
        "verMesesCerradosEntregados_CNT", 0
    )
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if mesesCerradosEntregados_CNT == 1 or adminIT == 1:
            empresasData = obtener_empresas()

            context = {"empresasData": empresasData}

            return render(
                request, "contabilidad/meses_cerrados_entregados.html", context
            )
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def drain_results(cursor):
    try:
        while cursor.nextset():
            pass
    except Exception:
        pass


def update_status_partida(request):
    try:
        partida_id = request.POST.get("partida_id")
        userName = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc("CONTA_UPDATE_STATUS_PARTIDAS", [partida_id, userName])
            results = cursor.fetchall()

        appConn.close()

        datos = {"save": 1}
    except Exception as e:
        datos = {"save": 0, "error": str(e)}

    return JsonResponse(datos)


def dataDetallesPartidasAutomaticas(request):
    idencpartida = request.POST.get("idencpartida")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_PARTIDA_DETALLE_X_FECHA_NPARTIDA", [idencpartida]
            )
            column_names = [desc[0] for desc in cursor.description]
            partidasData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"data": partidasData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataFlujoEfectivo(request):
    fechaInicial = request.POST.get("fechaInicial")
    fechaFinal = request.POST.get("fechaFinal")

    empresa = request.POST.get("empresa")
    opcion = request.POST.get("opcion")

    user_name = request.session.get("userName", "CONTABLE")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_FLUJO_EFECTIVO", [fechaInicial, fechaFinal, empresa, opcion]
            )
            column_names = [desc[0] for desc in cursor.description]
            partidasData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"data": partidasData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataPartidasDescuadradas(request):

    fecha = datetime.strptime(request.POST.get("fechaInicial"), "%Y-%m-%d")
    fechaInicial = fecha.strftime("%Y-%m-%d %H:%M:%S")

    fecha = datetime.strptime(request.POST.get("fechaFinal"), "%Y-%m-%d")
    fechaFinal = fecha.strftime("%Y-%m-%d %H:%M:%S")

    empresa = request.POST.get("empresa")
    npartida = request.POST.get("npartida")
    opcion = request.POST.get("opcion")

    user_name = request.session.get("userName", "CONTABLE")
    visualizarTodo = 0

    libroDiario_CNT = request.session.get("libroDiarioVerTodo_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if libroDiario_CNT == 1 or adminIT == 1:
        visualizarTodo = 1

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_PARTIDAS_DESCUADRADAS",
                [
                    fechaInicial,
                    fechaFinal,
                    empresa,
                    user_name,
                    npartida,
                    opcion,
                    visualizarTodo,
                ],
            )
            column_names = [desc[0] for desc in cursor.description]
            partidasData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"data": partidasData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataLibroDiario(request):

    fecha = datetime.strptime(request.POST.get("fechaInicial"), "%Y-%m-%d")
    fechaInicial = fecha.strftime("%Y-%m-%d %H:%M:%S")

    fecha = datetime.strptime(request.POST.get("fechaFinal"), "%Y-%m-%d")
    fechaFinal = fecha.strftime("%Y-%m-%d %H:%M:%S")

    empresa = request.POST.get("empresa")
    npartida = request.POST.get("npartida")
    opcion = request.POST.get("opcion")

    user_name = request.session.get("userName", "CONTABLE")
    visualizarTodo = 0

    libroDiario_CNT = request.session.get("libroDiarioVerTodo_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if libroDiario_CNT == 1 or adminIT == 1:
        visualizarTodo = 1

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_PARTIDAS_X_FECHA_NPARTIDA",
                [
                    fechaInicial,
                    fechaFinal,
                    empresa,
                    user_name,
                    npartida,
                    opcion,
                    visualizarTodo,
                ],
            )
            column_names = [desc[0] for desc in cursor.description]
            partidasData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"data": partidasData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def conta_partidas_descuadradas(request):
    user_id = request.session.get("user_id", "")
    libroDiario_CNT = request.session.get("libroDiarioVerTodo_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if libroDiario_CNT == 1 or adminIT == 1:
            empresasData = obtener_empresas()

            partidasAutomaticas = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_PARTIDAS_AUTOMATICAS", [0])
                    column_names = [desc[0] for desc in cursor.description]
                    partidasAutomaticas = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                # Cierra la conexión
                udcConn.close()
            except Exception as e:
                # Manejo de excepciones, puedes personalizar esto según tus necesidades
                return JsonResponse({"error": str(e)})

            context = {
                "empresasData": empresasData,
                "partidasAutomaticas": partidasAutomaticas,
                "date1": datetime.now().replace(day=1).strftime("%Y-%m-%d"),
                "date2": datetime.now().strftime("%Y-%m-%d"),
            }

            return render(request, "auditorias/partidas_descuadradas.html", context)
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def conta_libro_diario(request):
    user_id = request.session.get("user_id", "")
    libroDiario_CNT = request.session.get("libroDiarioVerTodo_CNT", 0)
    adminIT = request.session.get("contabilidadAdminIT", 0)

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        if libroDiario_CNT == 1 or adminIT == 1:
            empresasData = obtener_empresas()

            partidasAutomaticas = ""
            try:
                udcConn = connections["universal"]
                with udcConn.cursor() as cursor:
                    cursor.callproc("CONTA_GET_PARTIDAS_AUTOMATICAS", [0])
                    column_names = [desc[0] for desc in cursor.description]
                    partidasAutomaticas = [
                        dict(zip(map(str, column_names), row))
                        for row in cursor.fetchall()
                    ]

                # Cierra la conexión
                udcConn.close()
            except Exception as e:
                # Manejo de excepciones, puedes personalizar esto según tus necesidades
                return JsonResponse({"error": str(e)})

            context = {
                "empresasData": empresasData,
                "partidasAutomaticas": partidasAutomaticas,
                "date1": datetime.now().replace(day=1).strftime("%Y-%m-%d"),
                "date2": datetime.now().strftime("%Y-%m-%d"),
            }

            return render(request, "contabilidad/libro_diario.html", context)
        else:
            return HttpResponseRedirect(reverse("panel_contabilidad"))


def logoutRequest(request):
    request.session.flush()

    token = ""

    return HttpResponseRedirect(LOGIN_URL)


def dataTiposClasesCuentas(request):
    tipo = request.POST.get("tipo", "")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_TIPOS_CLASES", [tipo])
            column_names = [desc[0] for desc in cursor.description]
            tiposData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"save": 1, "data": tiposData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def obtener_cuentas_ordenadas(request):
    empresa = request.POST.get("empresa", 0)

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_ORDEN_CUENTAS_CONTABLES", [empresa])
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        udcConn.close()
        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def obtener_partida_completa(request):
    id_partida = request.POST.get("id_partida", 0)

    if not id_partida:
        return JsonResponse({"error": "ID no proporcionado"})

    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc("CONTA_GET_ENCABEZADO_PARTIDA", [id_partida])
            encabezado_cols = [col[0] for col in cursor.description]
            encabezado = dict(zip(encabezado_cols, cursor.fetchone()))

        with connections["universal"].cursor() as cursor:
            cursor.callproc("CONTA_GET_DETALLE_PARTIDA", [id_partida])
            detalle_cols = [col[0] for col in cursor.description]
            detalles = [dict(zip(detalle_cols, row)) for row in cursor.fetchall()]

        return JsonResponse({"save": 1, "encabezado": encabezado, "detalles": detalles})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataCentrosCosto(request):
    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("VAIC_GET_CENTROS")
            column_names = [desc[0] for desc in cursor.description]
            centrosData = [
                dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
            ]

        udcConn.close()

        return JsonResponse({"save": 1, "data": centrosData})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def conta_generar_partidas_view(request):
    user_id = request.session.get("user_id", "")

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        partidasAutomaticas = ""
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_TIPOS_PARTIDAS_SETEADAS")
                column_names = [desc[0] for desc in cursor.description]
                partidasAutomaticas = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            # Cierra la conexión
            udcConn.close()
        except Exception as e:
            # Manejo de excepciones, puedes personalizar esto según tus necesidades
            return JsonResponse({"error": str(e)})

        context = {
            "partidasAutomaticas": partidasAutomaticas,
            "date1": datetime.now().strftime("%Y-%m-%d"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "contabilidad/generar_partidas.html", context)


def conta_asignar_cuentas_movimientos(request):
    user_id = request.session.get("user_id", "")

    if user_id == "":
        return HttpResponseRedirect(reverse("login"))
    else:
        cuentasData = ""
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_CAT_CUENTAS", [3])
                column_names = [desc[0] for desc in cursor.description]
                cuentasData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            # Cierra la conexión
            udcConn.close()
        except Exception as e:
            # Manejo de excepciones, puedes personalizar esto según tus necesidades
            return JsonResponse({"error": str(e)})

        tiposMovimientoData = ""
        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_GET_TIPOS_CUENTAS_MOVIMIENTOS")
                column_names = [desc[0] for desc in cursor.description]
                tiposMovimientoData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            # Cierra la conexión
            udcConn.close()
        except Exception as e:
            # Manejo de excepciones, puedes personalizar esto según tus necesidades
            return JsonResponse({"error": str(e)})

        context = {
            "cuentasData": cuentasData,
            "tiposMovimientoData": tiposMovimientoData,
        }

        return render(request, "catalogos/asignar_cuentas_movimientos.html", context)


def insert_cuenta_tipo_movimiento(request):
    try:
        id_cuenta_movimiento = request.POST.get("id_cuenta_movimiento") or None
        id_cuenta_contable = request.POST.get("id_cuenta_contable")
        codigo_cuenta = request.POST.get("codigo_cuenta")
        nombre_cuenta = request.POST.get("nombre_cuenta")
        id_tipo = request.POST.get("id_tipo_cuenta_movimiento")
        usuario = request.session.get("userName", "CONTABLE")
        opcion = int(request.POST.get("opcion"))

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_INSERT_CUENTA_X_TIPO_MOVIMIENTO",
                [
                    id_cuenta_movimiento,
                    id_cuenta_contable,
                    codigo_cuenta,
                    nombre_cuenta,
                    id_tipo,
                    usuario,
                    opcion,
                ],
            )
            result = cursor.fetchone()  # OUT p_existe
            existe = result[0] if result else 0

        appConn.close()
        return JsonResponse({"save": 1, "existe": existe})

    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)})


def update_status_cuenta_tipo_movimiento(request):
    try:
        id_cuenta_movimiento = request.POST.get("id_cuenta_movimiento")
        usuario = request.session.get("userName", "CONTABLE")

        appConn = connections["universal"]
        with appConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_STATUS_UPDATE_CUENTA_X_TIPO_MOVIMIENTO",
                [id_cuenta_movimiento, usuario],
            )
        appConn.close()

        return JsonResponse({"save": 1})
    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)})


def dataCuentasAsignadas(request):
    opc = request.POST.get("opc")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_CUENTAS_ASIGNADAS_X_MOVIMIENTO", [opc, 0])
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataPartidasHermanasAsignadas(request):
    hermana = request.POST.get("hermana")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_PARTIDAS_HERMANAS_ASSIGNED", [hermana])
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataHistorialCierres(request):
    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_CIERRES_CONTABLES")
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def entregar_periodo_contable(request):
    id_cierre = request.POST.get("id")
    user = request.session.get("userName")

    if not id_cierre:
        return JsonResponse({"save": 0, "error": "ID inválido"}, status=400)

    try:
        with connections["default"].cursor() as cursor:
            cursor.callproc("CONTA_ENTREGAR_PERIODO_CONTABLE", [id_cierre, user])

        return JsonResponse({"save": 1})
    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)}, status=500)


@csrf_exempt
def API_cuentas_asignadas_x_movimiento(request):
    opc = request.POST.get("opc")

    try:
        udcConn = connections["universal"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_CUENTAS_ASIGNADAS_X_MOVIMIENTO", [1, 0])
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def get_estado_calendario_dashboard(request):
    if request.method == "POST":
        empresa_id = request.POST.get("empresa_id", "")
        anio = request.POST.get("anio")

        if not empresa_id or not anio:
            return JsonResponse({"error": "Faltan parámetros"})

        try:
            udcConn = connections["universal"]
            with udcConn.cursor() as cursor:
                cursor.callproc(
                    "CONTA_GET_ESTADO_CALENDARIO_EMPRESA_DASHBOARD", [empresa_id, anio]
                )
                column_names = [col[0] for col in cursor.description]
                raw_rows = cursor.fetchall()

                data = []
                for row in raw_rows:
                    row_dict = {}
                    for col, val in zip(column_names, row):
                        # Convertir campos tipo BIT (bytes) a int (0 o 1)
                        if isinstance(val, bytes):
                            row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                        else:
                            row_dict[str(col)] = val
                    data.append(row_dict)

            return JsonResponse({"save": 1, "data": data})

        except Exception as e:
            return JsonResponse({"error": str(e), "save": 0})


def get_partidas_descuadradas_dashboard(request):
    if request.method == "POST":
        try:
            udcConn = connections["contable"]
            with udcConn.cursor() as cursor:
                cursor.callproc("CONTA_COUNT_PARTIDAS_DESCUADRADAS")
                column_names = [col[0] for col in cursor.description]
                raw_rows = cursor.fetchall()

                data = []
                for row in raw_rows:
                    row_dict = {}
                    for col, val in zip(column_names, row):
                        # Convertir campos tipo BIT (bytes) a int (0 o 1)
                        if isinstance(val, bytes):
                            row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                        else:
                            row_dict[str(col)] = val
                    data.append(row_dict)

            return JsonResponse({"save": 1, "data": data})

        except Exception as e:
            return JsonResponse({"error": str(e), "save": 0})


def conta_reporte_estado_resultados_sucursal(request):
    try:
        sucursalesData = ""
        try:
            udcConn = connections["global_nube"]
            with udcConn.cursor() as cursor:
                cursor.callproc("GS_GET_SUCURSALES_X_EMPRESA", [0])
                column_names = [desc[0] for desc in cursor.description]
                sucursalesData = [
                    dict(zip(map(str, column_names), row)) for row in cursor.fetchall()
                ]

            udcConn.close()
        except Exception as e:
            return JsonResponse({"error": str(e)})

        context = {
            "empresasData": obtener_empresas(),
            "sucursalesData": sucursalesData,
            "date1": datetime.now().strftime("%Y-%m-01"),
            "date2": datetime.now().strftime("%Y-%m-%d"),
        }

        return render(request, "contabilidad/reporte_er_sucursal.html", context)

    except Exception as e:
        return HttpResponse(f"Error: {str(e)}")


def dataReporteERPorSucursal(request):
    try:
        empresa = request.POST.get("empresa")
        date1 = request.POST.get("date1")
        date2 = request.POST.get("date2")
        filtro = request.POST.get("filtro", 0)  # 0 = todo, 1 = solo cuentas con saldo
        sucursal = request.POST.get("sucursal", 0)  # 0 = todas

        with connections["contable"].cursor() as cursor:
            cursor.callproc(
                "CONTA_REPORTE_ER_POR_SUCURSAL",
                [empresa, date1, date2, filtro, sucursal],
            )
            column_names = [col[0] for col in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"data": data})

    except Exception as e:
        return JsonResponse({"error": str(e)})


#           """ ======================================= DATA PARA GENERAR PARTIDAS ======================================= """


def dataMovBancariosDetalles(request):
    referencia = request.POST.get("referencia", None)
    codigo = request.POST.get("codigo")
    fecha_inicio = request.POST.get("fechaInicial")
    fecha_fin = request.POST.get("fechaFinal")

    try:
        udcConn = connections["bankConn"]
        with udcConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_MOVIMIENTOS_BANCARIOS_DETALLE",
                [fecha_inicio, fecha_fin, codigo, referencia],
            )
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataDepositosVentas(request):
    referencia = request.POST.get("referencia", None)
    codigo = request.POST.get("codigo")
    fecha_inicio = request.POST.get("fechaInicial")
    fecha_fin = request.POST.get("fechaFinal")

    try:
        udcConn = connections["bankConn"]
        with udcConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_MOVIMIENTOS_BANCARIOS",
                [fecha_inicio, fecha_fin, codigo, referencia],
            )
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataFacturasCosto(request):
    codigo = request.POST.get("codigo")
    fecha_inicio = request.POST.get("fechaInicial")
    fecha_fin = request.POST.get("fechaFinal")

    try:
        udcConn = connections["super"]
        with udcConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_DATA_COSTOS_SUCURSAL", [fecha_inicio, fecha_fin, codigo]
            )
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataFacturas(request):
    codigo = request.POST.get("codigo")
    fecha_inicio = request.POST.get("fechaInicial")
    fecha_fin = request.POST.get("fechaFinal")

    try:
        udcConn = connections["super"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_DATA_VENTAS", [fecha_inicio, fecha_fin, codigo])
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataFacturasDetallesCosto(request):
    NoDocto = request.POST.get("NoDocto")

    try:
        udcConn = connections["super"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_DETALLE_X_FACTURA_COSTO_SUCURSAL", [NoDocto])
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataDevolDetalles(request):
    NoDocto = request.POST.get("NoDocto")

    try:
        udcConn = connections["super"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_DETALLE_X_DEVOLUCION", [NoDocto])
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataFacturasDetalles(request):
    NoDocto = request.POST.get("NoDocto")

    try:
        udcConn = connections["super"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_DETALLE_X_FACTURA", [NoDocto])
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataProvisionesDetalleEmpresa(request):
    empresa = request.POST.get("empresa")

    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc("TH_GET_PROVISIONES_TODOS", [empresa])
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def download_data_partidas_libro_diario(request):
    fechaInicialStr = request.GET.get("fechaInicial")
    fechaFinalStr = request.GET.get("fechaFinal")
    empresa = request.GET.get("empresa")
    npartida = request.GET.get("npartida")
    opcion = request.GET.get("opcion")

    if not fechaInicialStr or not fechaFinalStr:
        return HttpResponse("Fechas no válidas", status=400)

    try:
        fechaInicial = datetime.strptime(fechaInicialStr, "%Y-%m-%d").strftime(
            "%Y-%m-%d %H:%M:%S"
        )
        fechaFinal = datetime.strptime(fechaFinalStr, "%Y-%m-%d").strftime(
            "%Y-%m-%d %H:%M:%S"
        )
    except ValueError:
        return HttpResponse("Formato de fecha incorrecto", status=400)

    user_name = request.session.get("userName", "CONTABLE")
    visualizarTodo = 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Diario"

    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_DETALLES_PARTIDA_X_DATE_RANGE",
                [
                    fechaInicial,
                    fechaFinal,
                    empresa,
                    user_name,
                    npartida,
                    opcion,
                    visualizarTodo,
                ],
            )
            column_names = [desc[0] for desc in cursor.description]
            rows = cursor.fetchall()

        # Agregar encabezados
        ws.append(column_names)

        # Agregar filas
        for row in rows:
            ws.append(row)

        # Preparar respuesta
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"LibroDiario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"
        wb.save(response)
        return response

    except Exception as e:
        return HttpResponse(f"Error al generar el Excel: {e}", status=500)


def download_data_partidas_descuadradas_libro_diario(request):
    fechaInicialStr = request.GET.get("fechaInicial")
    fechaFinalStr = request.GET.get("fechaFinal")
    empresa = request.GET.get("empresa")
    npartida = request.GET.get("npartida")
    opcion = request.GET.get("opcion")

    if not fechaInicialStr or not fechaFinalStr:
        return HttpResponse("Fechas no válidas", status=400)

    try:
        fechaInicial = datetime.strptime(fechaInicialStr, "%Y-%m-%d").strftime(
            "%Y-%m-%d %H:%M:%S"
        )
        fechaFinal = datetime.strptime(fechaFinalStr, "%Y-%m-%d").strftime(
            "%Y-%m-%d %H:%M:%S"
        )
    except ValueError:
        return HttpResponse("Formato de fecha incorrecto", status=400)

    user_name = request.session.get("userName", "CONTABLE")
    visualizarTodo = 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Diario"

    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_DETALLES_PARTIDA_DESCUADRADAS_X_DATE_RANGE",
                [
                    fechaInicial,
                    fechaFinal,
                    empresa,
                    user_name,
                    npartida,
                    opcion,
                    visualizarTodo,
                ],
            )
            column_names = [desc[0] for desc in cursor.description]
            rows = cursor.fetchall()

        # Agregar encabezados
        ws.append(column_names)

        # Agregar filas
        for row in rows:
            ws.append(row)

        # Preparar respuesta
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"LibroDiario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"
        wb.save(response)
        return response

    except Exception as e:
        return HttpResponse(f"Error al generar el Excel: {e}", status=500)


def dataDevoluciones(request):
    codigo = request.POST.get("codigo")
    fecha_inicio = request.POST.get("fechaInicial")
    fecha_fin = request.POST.get("fechaFinal")

    try:
        udcConn = connections["super"]
        with udcConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_DATA_DEVOLUCIONES", [fecha_inicio, fecha_fin, codigo]
            )
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataOrdenesComprasIngresadas(request):
    codigo = request.POST.get("codigo")
    fecha_inicio = request.POST.get("fechaInicial")
    fecha_fin = request.POST.get("fechaFinal")

    try:
        udcConn = connections["super"]
        with udcConn.cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_COMPRAS_INGRESADAS", [fecha_inicio, fecha_fin, codigo]
            )
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


def dataOrdenesComprasDetallesIngresadas(request):
    NoDocto = request.POST.get("NoDocto")

    try:
        udcConn = connections["super"]
        with udcConn.cursor() as cursor:
            cursor.callproc("CONTA_GET_DETALLE_X_COMPRA_INGRESADA", [NoDocto])
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            data = []
            for row in raw_rows:
                row_dict = {}
                for col, val in zip(column_names, row):
                    # Convertir campos tipo BIT (bytes) a int (0 o 1)
                    if isinstance(val, bytes):
                        row_dict[str(col)] = int.from_bytes(val, byteorder="little")
                    else:
                        row_dict[str(col)] = val
                data.append(row_dict)

        return JsonResponse({"save": 1, "data": data})
    except Exception as e:
        return JsonResponse({"error": str(e)})


@csrf_exempt
def generar_partidas_automaticas_devol(request):
    codigo = request.POST.get("codigo")
    fecha_inicio = request.POST.get("fecha_inicio")
    fecha_fin = request.POST.get("fecha_final")

    if not fecha_inicio or not fecha_fin:
        return JsonResponse({"error": "Debes proporcionar ambas fechas"}, status=400)

    try:
        payloads = []

        with connections["super"].cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_DATA_DEVOLUCIONES", [fecha_inicio, fecha_fin, codigo]
            )
            column_names = [desc[0] for desc in cursor.description]
            rows = cursor.fetchall()

            for row in rows:
                data = dict(zip(column_names, row))

                payload = {
                    "codigo_partida": codigo,
                    "fecha": str(data["Fecha"]),
                    "referencia": data["NoNotaCredito"],
                    "usuario": data["Cajera"],
                    "empresa_id": data["fkEmpresa"],
                    "sucursal_id": int(data["fkSucursal"]),
                    "sistema": "POSCA",
                    "sistema_id": 5,
                    "valores": {
                        "venta_devolucion_15": float(
                            data.get("venta_devolucion_15", 0.0)
                        ),
                        "devolucion_efectivo": float(
                            data.get("devolucion_efectivo", 0.0)
                        ),
                    },
                }

                payloads.append(payload)

        # ========== ENVÍO POR BLOQUES ==========
        resultados = []
        for bloque in chunks(
            payloads, CHUNK_SIZE
        ):  # o usa la función chunks si no tenés la librería
            try:
                r = requests.post(
                    "http://3.230.160.184:81/API/conta/recibir/datos/generar/partida",
                    json=bloque,
                    headers={"API-Token": TOKEN},
                    timeout=120,
                )
                respuesta = r.json()
                resultados.extend(respuesta.get("resultados", []))
            except Exception as api_err:
                for p in bloque:
                    resultados.append(
                        {
                            "referencia": p.get("referencia"),
                            "status": "ERROR",
                            "mensaje": str(api_err),
                        }
                    )

        return JsonResponse({"save": 1, "resultados": resultados})

    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)}, status=500)


#           """ ======================================= FX GENERAR PARTIDAS ======================================= """


# Función para comprimir con GZIP
def compress_json(data):
    out = BytesIO()
    with gzip.GzipFile(fileobj=out, mode="wb") as f:
        f.write(json.dumps(data, default=str).encode("utf-8"))
    return out.getvalue()


# Función auxiliar para convertir Decimal a float
def sanitize_payload(data):
    if isinstance(data, dict):
        return {k: sanitize_payload(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [sanitize_payload(v) for v in data]
    elif isinstance(data, Decimal):
        return float(data)
    return data


def ejecutar_cierre_diario(var_fecha_inicio, empresa_id=3, user_name="API"):
    try:
        with connections["universal"].cursor() as cursor:
            cursor.callproc(
                "CONTA_GENERAR_CIERRE_X_DIA_RANGO",
                [var_fecha_inicio, empresa_id, user_name],
            )
    except Exception as e:
        print(f"[ERROR] Cierre diario fallido: {e}")


def build_payload(data, codigo_partida, exclusions, sistema, sistema_id):

    payload = {
        "codigo_partida": codigo_partida,
        "sistema": sistema,
        "sistema_id": sistema_id,
    }

    # ============= ENCABEZADO =============
    for campo_sql, campo_json in exclusions.items():
        valor = data.get(campo_sql)

        # Convertir fechas
        if isinstance(valor, (date, datetime)):
            valor = valor.isoformat()

        payload[campo_json] = valor

    # Convertir sucursal_id a int
    if "sucursal_id" in payload:
        try:
            payload["sucursal_id"] = int(payload["sucursal_id"])
        except:
            payload["sucursal_id"] = 0

    # ============= VALORES DINÁMICOS =============
    valores = {}

    for key, value in data.items():
        if key in exclusions:
            continue
        if value is None:
            continue

        # Decimal → float
        if isinstance(value, decimal.Decimal):
            valores[key] = float(value)
            continue

        # Fechas → ignorar
        if isinstance(value, (date, datetime)):
            continue

        # Números → float
        try:
            valores[key] = float(value)
            continue
        except:
            pass

    payload["valores"] = valores
    return payload


@csrf_exempt
def generar_partidas_automaticas_costos(request):
    codigo = request.POST.get("codigo")
    fecha_inicio = request.POST.get("fecha_inicio")
    fecha_fin = request.POST.get("fecha_final")

    if not fecha_inicio or not fecha_fin:
        return JsonResponse({"error": "Debes proporcionar ambas fechas"}, status=400)

    try:
        with connections["super"].cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_DATA_COSTOS_SUCURSAL", [fecha_inicio, fecha_fin, codigo]
            )
            columns = [d[0] for d in cursor.description]
            rows = cursor.fetchall()

        payloads = []

        # Campos específicos del encabezado para este SP
        exclusions = {
            "FechaFactura": "fecha",
            "Referencia": "referencia",
            "Vendedor": "usuario",
            "fkEmpresa": "empresa_id",
            "fkSucursal": "sucursal_id",
        }

        for row in rows:
            data = dict(zip(columns, row))

            payload = build_payload(
                data=data,
                codigo_partida=codigo,
                exclusions=exclusions,
                sistema="POSCA",
                sistema_id=5,
            )
            payloads.append(payload)

        # === Envío por bloques ===
        resultados = []
        for bloque in chunks(payloads, CHUNK_SIZE):
            try:
                r = requests.post(
                    "http://3.230.160.184:81/API/conta/recibir/datos/generar/partida",
                    json=bloque,
                    headers={"API-Token": TOKEN},
                    timeout=120,
                )
                resultados.extend(r.json().get("resultados", []))
            except Exception as api_err:
                resultados.append({"status": "ERROR", "mensaje": str(api_err)})

        ejecutar_cierre_diario(fecha_inicio)
        return JsonResponse({"save": 1, "resultados": resultados})

    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)}, status=500)


@csrf_exempt
def generar_partidas_automaticas_ventas(request):
    codigo = request.POST.get("codigo")
    fecha_inicio = request.POST.get("fecha_inicio")
    fecha_fin = request.POST.get("fecha_final")

    if not fecha_inicio or not fecha_fin:
        return JsonResponse({"error": "Debes proporcionar ambas fechas"}, status=400)

    try:
        with connections["super"].cursor() as cursor:
            cursor.callproc("CONTA_GET_DATA_VENTAS", [fecha_inicio, fecha_fin, codigo])
            columns = [d[0] for d in cursor.description]
            rows = cursor.fetchall()

        payloads = []

        # Campos específicos del encabezado para este SP
        exclusions = {
            "FechaFactura": "fecha",
            "Referencia": "referencia",
            "Vendedor": "usuario",
            "fkEmpresa": "empresa_id",
            "fkSucursal": "sucursal_id",
        }

        for row in rows:
            data = dict(zip(columns, row))

            payload = build_payload(
                data=data,
                codigo_partida=codigo,
                exclusions=exclusions,
                sistema="POSCA",
                sistema_id=5,
            )
            payloads.append(payload)

        # === Envío por bloques ===
        resultados = []
        for bloque in chunks(payloads, CHUNK_SIZE):
            try:
                r = requests.post(
                    "http://3.230.160.184:81/API/conta/recibir/datos/generar/partida",
                    json=bloque,
                    headers={"API-Token": TOKEN},
                    timeout=120,
                )
                resultados.extend(r.json().get("resultados", []))
            except Exception as api_err:
                resultados.append({"status": "ERROR", "mensaje": str(api_err)})

        ejecutar_cierre_diario(fecha_inicio)
        return JsonResponse({"save": 1, "resultados": resultados})

    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)}, status=500)


@csrf_exempt
def generar_partidas_automaticas_devoluciones(request):
    codigo = request.POST.get("codigo")
    fecha_inicio = request.POST.get("fecha_inicio")
    fecha_fin = request.POST.get("fecha_final")

    if not fecha_inicio or not fecha_fin:
        return JsonResponse({"error": "Debes proporcionar ambas fechas"}, status=400)

    try:
        with connections["super"].cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_DATA_DEVOLUCIONES", [fecha_inicio, fecha_fin, codigo]
            )
            columns = [d[0] for d in cursor.description]
            rows = cursor.fetchall()

        payloads = []

        # Campos específicos del encabezado para este SP
        exclusions = {
            "FechaFactura": "fecha",
            "Referencia": "referencia",
            "Vendedor": "usuario",
            "fkEmpresa": "empresa_id",
            "fkSucursal": "sucursal_id",
        }

        for row in rows:
            data = dict(zip(columns, row))

            payload = build_payload(
                data=data,
                codigo_partida=codigo,
                exclusions=exclusions,
                sistema="POSCA",
                sistema_id=5,
            )
            payloads.append(payload)

        # === Envío por bloques ===
        resultados = []
        for bloque in chunks(payloads, CHUNK_SIZE):
            try:
                r = requests.post(
                    "http://3.230.160.184:81/API/conta/recibir/datos/generar/partida",
                    json=bloque,
                    headers={"API-Token": TOKEN},
                    timeout=120,
                )
                resultados.extend(r.json().get("resultados", []))
            except Exception as api_err:
                resultados.append({"status": "ERROR", "mensaje": str(api_err)})

        ejecutar_cierre_diario(fecha_inicio)
        return JsonResponse({"save": 1, "resultados": resultados})

    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)}, status=500)


@csrf_exempt
def generar_partidas_automaticas_compras(request):
    codigo = request.POST.get("codigo")
    fecha_inicio = request.POST.get("fecha_inicio")
    fecha_fin = request.POST.get("fecha_final")

    if not fecha_inicio or not fecha_fin:
        return JsonResponse({"error": "Debes proporcionar ambas fechas"}, status=400)

    try:
        with connections["super"].cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_COMPRAS_INGRESADAS", [fecha_inicio, fecha_fin, codigo]
            )
            columns = [d[0] for d in cursor.description]
            rows = cursor.fetchall()

        exclusions = {
            "FechaIngreso": "fecha",
            "Referencia": "referencia",
            "Ingreso": "usuario",
            "fkEmpresa": "empresa_id",
            "fkSucursal": "sucursal_id",
        }

        payloads = []
        for row in rows:
            data = dict(zip(columns, row))

            payload = build_payload(
                data=data,
                codigo_partida=codigo,
                exclusions=exclusions,
                sistema="POSCA",
                sistema_id=5,
            )

            payloads.append(payload)

        # Envío
        resultados = []
        for bloque in chunks(payloads, CHUNK_SIZE):
            try:
                r = requests.post(
                    "http://3.230.160.184:81/API/conta/recibir/datos/generar/partida",
                    json=bloque,
                    headers={"API-Token": TOKEN},
                    timeout=120,
                )
                resultados.extend(r.json().get("resultados", []))
            except Exception as api_err:
                resultados.append({"status": "ERROR", "mensaje": str(api_err)})

        ejecutar_cierre_diario(fecha_inicio)
        return JsonResponse({"save": 1, "resultados": resultados})

    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)}, status=500)


@csrf_exempt
def generar_partidas_automaticas_provisionales(request):
    codigo = request.POST.get("codigo")
    fecha_inicio = request.POST.get("fecha_inicio")
    fecha_fin = request.POST.get("fecha_final")

    empresas = obtener_empresas()

    userName = request.session.get("userName", "")

    resultados = []

    try:
        for e in empresas:
            empresa = e.get("id", 0)

            with connections["universal"].cursor() as cursor:
                cursor.callproc(
                    "TH_GET_PARTIDA_PREVIEW_PROVISIONES",
                    [empresa, codigo, userName, fecha_fin],
                )

            resultados.append(
                {
                    "referencia": "",
                    "status": "OK",
                    "mensaje": "Registrado correctamente",
                }
            )

        return JsonResponse({"save": 1, "resultados": resultados})
    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)}, status=500)


@csrf_exempt
def generar_partidas_automaticas_movimientos_bancarios(request):
    referencia = request.POST.get("referencia")
    codigo = request.POST.get("codigo")
    fecha_inicio = request.POST.get("fecha_inicio")
    fecha_fin = request.POST.get("fecha_final")

    if not fecha_inicio or not fecha_fin:
        return JsonResponse({"error": "Debes proporcionar ambas fechas"}, status=400)

    try:
        with connections["bankConn"].cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_MOVIMIENTOS_BANCARIOS",
                [fecha_inicio, fecha_fin, codigo, referencia],
            )
            columns = [d[0] for d in cursor.description]
            rows = cursor.fetchall()

        exclusions = {
            "FechaMovimiento": "fecha",
            "Referencia": "referencia",
            "ElaboradoPor": "usuario",
            "FkEmpresa": "empresa_id",
            "FkSucursal": "sucursal_id",
        }

        payloads = []
        for row in rows:
            data = dict(zip(columns, row))

            payload = build_payload(
                data=data,
                codigo_partida=codigo,
                exclusions=exclusions,
                sistema="BANCARIO",
                sistema_id=2,
            )

            payloads.append(payload)

        # envío
        resultados = []
        for bloque in chunks(payloads, CHUNK_SIZE):
            try:
                r = requests.post(
                    "http://3.230.160.184:81/API/conta/recibir/datos/generar/partida",
                    json=bloque,
                    headers={"API-Token": TOKEN},
                    timeout=120,
                )
                resultados.extend(r.json().get("resultados", []))
            except Exception as api_err:
                resultados.append({"status": "ERROR", "mensaje": str(api_err)})

        ejecutar_cierre_diario(fecha_inicio)
        return JsonResponse({"save": 1, "resultados": resultados})

    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)}, status=500)


@csrf_exempt
def generar_partidas_automaticas_movimientos_bancarios_detalles(request):
    referencia = request.POST.get("referencia")
    codigo = request.POST.get("codigo")
    fecha_inicio = request.POST.get("fecha_inicio")
    fecha_fin = request.POST.get("fecha_final")

    if not fecha_inicio or not fecha_fin:
        return JsonResponse({"error": "Debes proporcionar ambas fechas"}, status=400)

    try:
        with connections["bankConn"].cursor() as cursor:
            cursor.callproc(
                "CONTA_GET_MOVIMIENTOS_BANCARIOS_DETALLE",
                [fecha_inicio, fecha_fin, codigo, referencia],
            )
            columns = [d[0] for d in cursor.description]
            rows = cursor.fetchall()

        exclusions = {
            "FechaMovimiento": "fecha",
            "Referencia": "referencia",
            "ElaboradoPor": "usuario",
            "FkEmpresa": "empresa_id",
            "FkSucursal": "sucursal_id",
        }

        payloads = []
        for row in rows:
            data = dict(zip(columns, row))

            payload = build_payload(
                data=data,
                codigo_partida=codigo,
                exclusions=exclusions,
                sistema="BANCARIO",
                sistema_id=2,
            )

            payloads.append(payload)

        # envío
        resultados = []
        for bloque in chunks(payloads, CHUNK_SIZE):
            try:
                r = requests.post(
                    "http://3.230.160.184:81/API/conta/recibir/datos/generar/partida",
                    json=bloque,
                    headers={"API-Token": TOKEN},
                    timeout=120,
                )
                resultados.extend(r.json().get("resultados", []))
            except Exception as api_err:
                resultados.append({"status": "ERROR", "mensaje": str(api_err)})

        ejecutar_cierre_diario(fecha_inicio)
        return JsonResponse({"save": 1, "resultados": resultados})

    except Exception as e:
        return JsonResponse({"save": 0, "error": str(e)}, status=500)


""" 

Desglose para ver detalles

Voucher
Cod Provee / Num Factura

"""
