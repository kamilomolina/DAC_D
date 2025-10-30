from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.db import connections
from django.http import JsonResponse
from django.conf import settings
from datetime import datetime, timedelta
import requests
import pandas as pd
import time
import hashlib
import os
import subprocess
import traceback
from rest_framework_simplejwt.tokens import RefreshToken
import jwt
from django.views.decorators.csrf import csrf_exempt
import json
import base64
from django.shortcuts import redirect, render
from django.http import HttpResponseRedirect
import logging
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import render, HttpResponseRedirect
import xlrd
from datetime import date, datetime
from django.http import JsonResponse, Http404
from django.urls import reverse
import csv
import tempfile
import pyodbc


#*************************************** Planilla ********************************
logger = logging.getLogger(__name__)  # Inicializa el logger


def gestion_planilla(request):
    id = request.session.get('user_id', '')

    if id == '':
        return HttpResponseRedirect(reverse('login'))

    id_planilla = request.GET.get('id_planilla')
    if not id_planilla:
        return render(request, 'planilla/gestion_planilla.html', {'error': 'ID de planilla no proporcionado'})

    try:
        id_planilla = int(id_planilla)

        with connections['universal'].cursor() as cursor:
            bancos = obtener_bancos()
   
            cursor.callproc('TH_OBTENER_PLANILLA_POR_ID', [id_planilla])
            resultado = cursor.fetchone()

            if not resultado:
                raise Http404("Planilla no encontrada")

            id_planilla, nombre_planilla, tipo_planilla, periodo, estado, nombre_empresa, id_empresa, fecha_por_aplicar = resultado
            logger.debug(f"Resultado obtenido: {resultado}")
            

            context = {
                'codigo': id_planilla,
                'nombre_planilla': nombre_planilla,
                'descripcion_cargo_laboral': tipo_planilla,
                'periodo': periodo,
                'estado': estado,
                'NombreEmpresa': nombre_empresa,
                'PkEmpresa': id_empresa,
                'fecha_por_aplicar':fecha_por_aplicar,
                'bancos': bancos
            }
            logger.debug(f"Contexto generado: {context}")

        return render(request, 'planilla/gestion_planilla.html', context)

    except ValueError:
        return render(request, 'planilla/gestion_planilla.html', {'error': 'ID de planilla inválido'})
    except Http404 as e:
        logger.error(f"Planilla no encontrada: {e}")
        return render(request, 'planilla/gestion_planilla.html', {'error': 'Planilla no encontrada'})
    except Exception as e:
        logger.error(f"Error al obtener la planilla: {e}")
        return render(request, 'planilla/gestion_planilla.html', {'error': f'Error inesperado: {e}'})

def verificar_registros_deducciones(request):
    try:
        id_planilla = request.POST.get('id_planilla')

        with connections['universal'].cursor() as cursor:
            cursor.callproc('TH_VERIFICAR_REGISTROS_AUTOMATICOS', [id_planilla])
            resultado = cursor.fetchall()
            if not resultado:
                return JsonResponse({
                    "hay_registros": False,
                    "cantidad_registros": 0
                }, status=200)

            cantidad_registros = resultado[0][0]
            hay_registros = cantidad_registros > 0

        return JsonResponse({
            "hay_registros": hay_registros,
            "cantidad_registros": cantidad_registros
        }, status=200)
    except Exception as e:
        return JsonResponse({"error": f"Error al verificar los registros: {str(e)}"}, status=400)


from openpyxl import load_workbook

def procesar_archivo_excel(archivo, fecha_planilla, creado_por, codigo_planilla):
    with connections['universal'].cursor() as cursor:
        # Llamar al procedimiento para actualizar registros automáticos de la planilla específica
        cursor.execute("CALL TH_ACTUALIZAR_REGISTROS_AUTOMATICOS(%s);", [codigo_planilla])
        
        # Limpiar la tabla temporal antes de insertar nuevos datos
        cursor.execute("CALL TH_TRUNCATE_TEMP_DEDUCCIONES_BONIFICACIONES();")

    # Leer el archivo Excel usando openpyxl
    workbook = load_workbook(archivo)
    sheet = workbook.active  # Selecciona la primera hoja

    # Validar encabezados esperados
    columnas_esperadas = ['Identidad', 'Nombre', 'Valor', 'Ajuste Nomina', 'Tipo Ajuste', 'Empresa']
    header = [cell.value.strip() for cell in sheet[1]]  # Lee la primera fila como encabezado

    for col in columnas_esperadas:
        if col not in header:
            raise ValueError(f"Falta la columna '{col}' en el archivo.")

    registros = []  # Lista para almacenar los datos

    # Procesar filas del archivo Excel
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Comenzar en la fila 2
        identidad = str(row[header.index('Identidad')]).strip()
        nombre = str(row[header.index('Nombre')]).strip()
        id_tipo = int(row[header.index('Tipo Ajuste')]) if row[header.index('Tipo Ajuste')] else 0
        valor = float(row[header.index('Valor')]) if row[header.index('Valor')] else 0.0
        tipo_nomina_archivo = int(row[header.index('Ajuste Nomina')]) if row[header.index('Ajuste Nomina')] else 0
        id_empresa = int(row[header.index('Empresa')]) if row[header.index('Empresa')] else 0

        registros.append((
            codigo_planilla,
            fecha_planilla,
            identidad,
            nombre,
            id_tipo,
            valor,
            tipo_nomina_archivo,
            2,
            id_empresa,  
            creado_por
        ))

    # Inserción masiva de los registros procesados
    insertar_registros_masivos(registros, connections['universal'])

    # Llamar al procedimiento para insertar deducciones
    with connections['universal'].cursor() as cursor:
        cursor.execute("CALL TH_INSERTAR_DEDUCCIONES();")


def insertar_registros_masivos(registros, conexion):
    """Inserta registros de manera masiva en la tabla de MySQL."""
    sql = """
        INSERT INTO temp_deducciones_bonificaciones 
        (codigo_planilla, fecha_planilla, identidad, nombre, id_tipo, valor, tipo_nomina_archivo, estado, id_empresa, creado_por)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    with conexion.cursor() as cursor:
        cursor.executemany(sql, registros)
    conexion.commit()


def subir_deducciones_bonificaciones(request):
    if request.method != 'POST':
        return JsonResponse({"error": "Método no permitido."}, status=405)

    codigo_planilla = request.POST.get('codigo')
    fecha_planilla = request.POST.get('fecha_planilla')
    creado_por = request.POST.get('creado_por')
    archivo_subida = request.FILES.get('archivoSubida')

    if not archivo_subida:
        return JsonResponse({"error": "No se ha subido ningún archivo."}, status=400)

    try:
        procesar_archivo_excel(archivo_subida, fecha_planilla, creado_por, codigo_planilla)
        return JsonResponse({"mensaje": "Archivo procesado y guardado con éxito."}, status=200)
    except Exception as e:
        return JsonResponse({"error": f"Error al procesar el archivo: {str(e)}"}, status=400)



def nueva_planilla(request):
    id = request.session.get('user_id', '')

    if id == '':
        return HttpResponseRedirect(reverse('login'))
    else:
        empresas = obtener_empresas()
        tipo_planillas = obtener_tipos_planillas()
        print(tipo_planillas)  
        context = {
            'empresas': empresas,
            'tipo_planillas': tipo_planillas,
        }
        return render(request, 'planilla/planilla.html', context)

def obtener_empresas():
    try:
        with connections['global_local'].cursor() as cursor:
            cursor.callproc('TH_GET_EMPRESAS')
            results = cursor.fetchall()
        return results
    except Exception as e:
        print(f"Error al obtener empresas: {e}")
        return []


def obtener_tipos_planillas():
    try:
        with connections['universal'].cursor() as cursor:
            cursor.callproc('TH_GET_CT_CLASES_DE_PLANILLAS')
            results = cursor.fetchall()
        return results
    except Exception as e:
        print(f"Error al obtener tipos de planillas: {e}")
        return []



def obtener_empleados(request):
    try:
        # Obtén el parámetro de empresa desde el request (puede ser GET o POST)
        p_id_empresa = request.GET.get('id_empresa')
        if not p_id_empresa:
            return JsonResponse({'status': 'fail', 'error': 'El parámetro id_empresa es requerido.'})

        with connections['universal'].cursor() as cursor:
            cursor.callproc('TH_OBTENER_NOMBRE_COMPLETO_EMPLEADOS', [p_id_empresa])
            results = cursor.fetchall()

        empleados = [
            {
                'id_empleado': row[0],
                'nombre_completo': row[1].split(' | ')[0],
                'identidad': row[1].split(' | ')[1]
            }
            for row in results
        ]

        return JsonResponse({'status': 'success', 'empleados': empleados}, safe=False)

    except Exception as e:
        return JsonResponse({'status': 'fail', 'error': str(e)})




def validar_fecha_aplicar_planilla(request):
    try:
        aplica = 0
        pendientes = 0

        id_planilla = request.POST.get('id_planilla')
        userName = request.session.get('userName', '')

        fecha_actual = datetime.datetime.now()

        udcConn = connections['universal']
        with udcConn.cursor() as cursor:
            cursor.callproc('TH_VALIDAR_FECHA_APLICAR_PLANILLA', [
                id_planilla
            ]) 
            results = cursor.fetchall()

            if results:
                for result in results:
                    aplica = result[0]
                    desde = result[1]
                    hasta = result[2]
            else:
                aplica = 0
                desde = 0
                hasta = 0
        
        udcConn.commit()

        udcConn = connections['universal']
        with udcConn.cursor() as cursor:
            cursor.callproc('TH_ASISTENCIA_VALIDAR_ACCIONES_MARCACIONES_v2', [desde, hasta, 0, 0])
            results = cursor.fetchall()

            if results:
                for result in results:
                    pendientes = result[0]


        udcConn.close()

        datos = {'save': 1, 'aplica': aplica, 'pendientes': pendientes}
    except Exception as e:
        datos = {'save': 0, 'error': str(e)}
    
    return JsonResponse(datos)






def obtener_bonificaciones(request):
    try:
        with connections['universal'].cursor() as cursor:
            # Llamar al procedimiento almacenado
            cursor.callproc('TH_GET_BONIFICACIONES')  
            results = cursor.fetchall()

        # Formatear los resultados en una lista de diccionarios
        bonificaciones = [{'id': row[0], 'descripcion': row[1]} for row in results]

        # Devolver los resultados en formato JSON
        return JsonResponse({'bonificaciones': bonificaciones})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
def obtener_tipos_manuales(request):
    tipo = request.GET.get('tipo')  # Obtener el parámetro 'tipo' del request
    
    if not tipo:
        return JsonResponse({'error': 'El parámetro "tipo" es requerido.'}, status=400)
    
    try:
        tipo = int(tipo)
    except ValueError:
        return JsonResponse({'error': 'El parámetro "tipo" debe ser un número entero.'}, status=400)
    
    resultados = []
    try:
        with connections['universal'].cursor() as cursor:
            cursor.callproc('TH_GET_TIPOS_MANUALES', [tipo])  # Llamada al procedimiento almacenado
            for row in cursor.fetchall():
                resultados.append({'id': row[0], 'descripcion': row[1]})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'resultados': resultados}, safe=False)
    


def obtener_cat_categorias_tipos_deduccion(request):
    try:
        with connections['universal'].cursor() as cursor:
            cursor.callproc('TH_MOSTRAR_CATEGORIAS_TIPOS_DEDUCCION')  
            results = cursor.fetchall()

        # Formatear los resultados en una lista de diccionarios
        categorias = [{'id': row[0], 'nombre': row[1]} for row in results]
        
        # Devolver los resultados en formato JSON
        return JsonResponse({'categorias': categorias})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def mostrar_tipos_deducciones(request):
    try:
        print("Iniciando el procedimiento almacenado...")  # Mensaje de depuración
        with connections['universal'].cursor() as cursor:
            cursor.execute("CALL TH_MOSTRAR_TIPOS_DEDUCCIONES()")
            rows = cursor.fetchall()
            print(f"Filas obtenidas: {len(rows)}")  # Número de filas obtenidas

            # Formatear los datos en una lista de diccionarios
            deducciones = [
                {
                    'id_deduccion': row[0],
                    'id_categoria': row[1],
                    'nombre_categoria': row[2],  # Nombre de la categoría
                    'descripcion': row[3],
                    'automatizacion': row[4],
                    'forma': row[5],
                    'nombre_tipo': row[6],  # Nombre del tipo de deducción
                    'tipo': row[7],  # Tipo como entero
                    'creado_por': row[8],
                    'fecha_creacion': row[9],
                    'fecha_modificacion': row[10],
                    'modificado_por': row[11],
                    'estado': row[12]
                }
                for row in rows
            ]

            print("Formato de deducciones completado.")  # Mensaje de depuración

        # Retornar los datos en formato JSON
        print("Devolviendo respuesta JSON...")  # Mensaje de depuración
        return JsonResponse({'deducciones': deducciones}, safe=False)

    except Exception as e:
        error_traceback = traceback.format_exc()  # Obtener el rastreo completo del error
        print(f"Error al ejecutar la vista: {str(e)}")  # Imprimir el error en la consola
        print(error_traceback)  # Imprimir el rastreo del error completo

        # Retornar un JSON con el mensaje de error y el rastreo de la pila
        return JsonResponse({'error': str(e), 'traceback': error_traceback}, status=500)


def guardar_tipo_deduccion(request):
    if request.method == 'POST':
        try:
            # Leer los datos de la solicitud en formato JSON
            body_unicode = request.body.decode('utf-8')
            print(f"Datos recibidos en la solicitud: {body_unicode}")

            data = json.loads(body_unicode)
            id_deduccion = data.get('id_deduccion')
            p_descripcion = data.get('descripcion', '')
            p_automatizacion = data.get('automatizacion')
            p_forma = data.get('forma')
            p_tipo = data.get('tipo')
            p_id_categoria = data.get('id_categoria')  
            p_creado_por = data.get('creado_por')

           
            # Verificar que todos los campos requeridos estén presentes
            if not all([p_forma, p_tipo, p_id_categoria, p_creado_por]):
                return JsonResponse({'status': 'fail', 'error': 'Faltan campos requeridos'}, status=400)


            with connections['universal'].cursor() as cursor:
                if id_deduccion:
                    print(f"Actualizando tipo de deducción con id_deduccion={id_deduccion}")
                    cursor.callproc('TH_EDITAR_TIPO_DEDUCCION', [
                    id_deduccion,
                    p_descripcion,
                    p_automatizacion,
                    p_forma,
                    p_tipo,
                    p_id_categoria,
                    p_creado_por  # Cambia a p_modificado_por si es diferente
                ])
                else:
                    print("Insertando nuevo tipo de deducción")
                    cursor.callproc('TH_GUARDAR_TIPO_DEDUCCION', [
                        p_descripcion,
                        p_automatizacion,
                        p_forma,
                        p_tipo,
                        p_id_categoria,
                        p_creado_por
                    ])
            print("Operación exitosa")
            return JsonResponse({'status': 'success'})

        except Exception as e:
            print(f"Error en la operación: {str(e)}")
            return JsonResponse({'status': 'fail', 'error': str(e)})

    return JsonResponse({'status': 'fail', 'error': 'Método no permitido'}, status=405)



def guardar_tipo_bonificacion(request):
    if request.method == 'POST':
        try:
            # Leer los datos de la solicitud en formato JSON
            body_unicode = request.body.decode('utf-8')
            print(f"Datos recibidos en la solicitud: {body_unicode}")

            data = json.loads(body_unicode)
            id_bonificacion = data.get('id_bonificacion')  # Ajustado para coincidir con el input oculto
            p_id_categoria = data.get('bonificacion_nombre')  # Coincide con el select de categoría
            p_descripcion = data.get('descripcion_bonificacion')  # Descripción de la bonificación
            p_forma = data.get('forma_bonificacion')  # Forma de la bonificación
            p_automatizacion = data.get('automatizacion_bonificacion')  # Automatización (0 o 1)
            p_creado_por = data.get('creado_por')  # Usuario que realiza la acción

          
            # Verificar que los campos requeridos estén presentes (excluyendo p_descripcion)
            if not all([p_id_categoria, p_forma, p_creado_por]):
                return JsonResponse({'status': 'fail', 'error': 'Faltan campos requeridos'}, status=400)

            with connections['universal'].cursor() as cursor:
                if id_bonificacion:  # Actualizar si existe id_bonificacion
                    print(f"Actualizando tipo de bonificación con id_bonificacion={id_bonificacion}")
                    cursor.callproc('TH_ACTUALIZAR_TIPO_BONIFICACION', [
                        id_bonificacion,
                        p_id_categoria,
                        p_descripcion,
                        p_forma,
                        p_automatizacion,
                        p_creado_por
                    ])
                else:  # Insertar nueva bonificación si no existe id_bonificacion
                    print("Insertando nuevo tipo de bonificación")
                    cursor.callproc('TH_GUARDAR_TIPO_BONIFICACION', [
                        p_id_categoria,
                        p_descripcion,
                        p_forma,
                        p_automatizacion,
                        p_creado_por
                    ])

            print("Operación exitosa")
            return JsonResponse({'status': 'success'})

        except Exception as e:
            print(f"Error en la operación: {str(e)}")
            return JsonResponse({'status': 'fail', 'error': str(e)})

    return JsonResponse({'status': 'fail', 'error': 'Método no permitido'}, status=405)

def guardar_tipo_Comisiones(request):
    if request.method == 'POST':
        try:
            # Leer los datos de la solicitud en formato JSON
            body_unicode = request.body.decode('utf-8')
            print(f"Datos recibidos en la solicitud: {body_unicode}")

            data = json.loads(body_unicode)
            id_comision = data.get('id_comision')  # ID de la comisión
            p_nombre = data.get('Comision_nombre')  # Nombre de la comisión
            p_descripcion = data.get('descripcion_Comision')  # Descripción
            p_forma = data.get('forma_Comision')  # Forma de la comisión
            p_automatizacion = data.get('automatizacion_Comision')  # Automatización (0 o 1)
            p_creado_por = data.get('creado_por')  # Usuario que realiza la acción

            # Validar los campos requeridos
            if not all([p_nombre, p_forma, p_creado_por]):
                return JsonResponse({'status': 'fail', 'error': 'Faltan campos requeridos'}, status=400)

            with connections['universal'].cursor() as cursor:
                if id_comision and id_comision.strip():  # Actualizar si `id_comision` tiene un valor válido
                    print(f"Actualizando tipo de COMISION con id_comision={id_comision}")
                    cursor.callproc('TH_ACTUALIZAR_TIPO_COMISION', [
                        id_comision,
                        p_nombre,
                        p_descripcion,
                        p_forma,
                        p_automatizacion,
                        p_creado_por
                    ])
                else:  # Insertar nueva comisión si `id_comision` está vacío
                    print("Insertando nuevo tipo de COMISION")
                    cursor.callproc('TH_GUARDAR_TIPO_COMISION', [
                        p_nombre,
                        p_descripcion,
                        p_forma,
                        p_automatizacion,
                        p_creado_por
                    ])

            print("Operación exitosa")
            return JsonResponse({'status': 'success'})

        except Exception as e:
            print(f"Error en la operación: {str(e)}")
            return JsonResponse({'status': 'fail', 'error': str(e)})

    return JsonResponse({'status': 'fail', 'error': 'Método no permitido'}, status=405)

def mostrar_tipos_Comisiones(request):
    if request.method == 'GET':
        try:
            # Conexión a la base de datos y ejecución del procedimiento almacenado
            with connections['universal'].cursor() as cursor:
                cursor.callproc('TH_MOSTRAR_TIPOS_COMISIONES')
                resultados = cursor.fetchall()
                # Obtener los nombres de las columnas
                columnas = [col[0] for col in cursor.description]

            # Formatear los resultados como una lista de diccionarios
            comisiones = [dict(zip(columnas, fila)) for fila in resultados]

            # Retornar los datos en formato JSON
            return JsonResponse({'Comisiones': comisiones}, status=200)
        except Exception as e:
            # En caso de error, retornar el mensaje de error
            return JsonResponse({'error': str(e)}, status=500)
    else:
        # Método no permitido
        return JsonResponse({'error': 'Método no permitido'}, status=405)


def eliminar_tipo_Comision(request):
    if request.method == 'POST':
        try:
            # Obtener los datos de la solicitud
            id_comision = request.POST.get('id_Comision')
            modificado_por = request.POST.get('modificado_por')

            if not id_comision or not modificado_por:
                return JsonResponse({'status': 'fail', 'message': 'Datos insuficientes para eliminar la comisión.'}, status=400)

            # Ejecutar el procedimiento almacenado
            with connections['universal'].cursor() as cursor:
                cursor.callproc('TH_ELIMINAR_TIPO_COMISION', [id_comision, modificado_por])

            return JsonResponse({'status': 'success', 'message': 'Comisión eliminada correctamente.'})
        except Exception as e:
            return JsonResponse({'status': 'fail', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'fail', 'message': 'Método no permitido.'}, status=405)




def insertar_planilla(request):
    if request.method == 'POST':
        # Obtener los datos del formulario
        id_planilla = request.POST.get('id_planilla')  # Para actualización
        nombre_planilla = request.POST.get('nombre_planilla')
        id_tipo_planilla = request.POST.get('tipo_planillas')
        id_empresa = request.POST.get('empresa_planillas')
        periodo_desde = request.POST.get('periodo_desde')
        periodo_hasta = request.POST.get('periodo_hasta')
        fecha_aplicar = request.POST.get('fecha_aplicar')
        modificado_por = request.POST.get('creado_por')  # Usuario que modifica

        try:
            with connections['universal'].cursor() as cursor:
                if id_planilla:  # Si existe `id_planilla`, actualizar
                    print(f"Actualizando planilla con id_planilla={id_planilla}")
                    cursor.callproc('TH_EDITAR_PLANILLA', [
                        id_planilla,
                        nombre_planilla,
                        id_tipo_planilla,
                        id_empresa,
                        periodo_desde,
                        periodo_hasta,
                        modificado_por
                    ])
                else:  # Si no existe `id_planilla`, insertar nueva
                    print("Insertando nueva planilla")
                    cursor.callproc('TH_INSERTAR_PLANILLA', [
                        nombre_planilla,
                        id_tipo_planilla,
                        id_empresa,
                        periodo_desde,
                        periodo_hasta,
                        fecha_aplicar,
                        modificado_por
                    ])
            print("Operación exitosa")
            return JsonResponse({'status': 'success', 'message': 'Operación realizada correctamente'})

        except Exception as e:
            print(f"Error en la operación: {str(e)}")  # Imprimir error

            # Verificar si el error es específico del procedimiento almacenado
            if 'Ya existe una planilla con el mismo tipo, empresa y fechas de periodo' in str(e):
                error_message = 'Ya existe una planilla con el mismo tipo, empresa y periodo especificado. Por favor, verifica las fechas.'
            elif 'No se puede crear una nueva planilla' in str(e):
                error_message = 'No se puede crear una nueva planilla. Existen planillas en estado "Creada" o "En proceso".'
            elif 'La planilla especificada no existe' in str(e):
                error_message = 'La planilla especificada no existe.'
            elif 'La fecha de inicio del período no puede ser mayor que la fecha de fin.' in str(e):
                error_message = 'La fecha de inicio del período no puede ser mayor que la fecha de fin..'
            elif 'La planilla solo puede ser creada para el mes actual y el mes anterior.' in str(e):
                error_message = 'La planilla solo puede ser creada para el mes actual y el mes anterior.'
            else:
                error_message = 'Hubo un problema al realizar la solicitud.'

            return JsonResponse({'status': 'fail', 'message': error_message}, status=400)

    else:
        print("Método no permitido")
        return JsonResponse({'status': 'fail', 'error': 'Método no permitido'}, status=405)


def obtener_planillas(request):
    fecha_inicio = request.POST.get('fecha_inicio')
    fecha_final = request.POST.get('fecha_final')

    try:
        udcConn = connections['universal']
        with udcConn.cursor() as cursor:
            cursor.callproc("TH_OBTENER_PLANILLAS", [fecha_inicio, fecha_final])
            column_names = [desc[0] for desc in cursor.description]
            planillasData = [dict(zip(map(str, column_names), row)) for row in cursor.fetchall()]
        
        # Cierra la conexión
        udcConn.close()
        
        # Devuelve los resultados como JSON
        return JsonResponse({'data': planillasData})
    except Exception as e:
        # Manejo de excepciones, puedes personalizar esto según tus necesidades
        return JsonResponse({'error': str(e)})





def eliminar_planilla(request):
    if request.method == 'POST':
        id_planilla = request.POST.get('id_planilla')

        if not id_planilla:
            return JsonResponse({'status': 'fail', 'message': 'ID de planilla no proporcionado.'}, status=400)

        try:
            with connections['universal'].cursor() as cursor:
                # Llamar al procedimiento almacenado para cambiar el estado a 4
                cursor.callproc('TH_ELIMINAR_PLANILLA', [id_planilla])

            return JsonResponse({'status': 'success', 'message': 'Planilla eliminada correctamente.'})

        except Exception as e:
            # Capturar cualquier error y devolverlo como respuesta
            error_message = str(e)
            if 'La planilla especificada no existe' in error_message:
                error_message = 'La planilla especificada no existe o ya fue eliminada.'

            return JsonResponse({'status': 'fail', 'message': error_message}, status=500)

    # Responder con error si no es una solicitud POST
    return JsonResponse({'status': 'fail', 'message': 'Método de solicitud no permitido.'}, status=405)


def obtener_codigos_planillas(request):
    logger.debug("Entrando en la vista obtener_codigos_planillas")

    if request.method == 'GET':
        try:
            logger.debug("Obteniendo la lista de planillas")

            with connections['universal'].cursor() as cursor:
                # Llamar al procedimiento almacenado
                cursor.callproc('TH_OBTENER_CODIGO_PLANILLAS')
                resultados = cursor.fetchall()

                planillas = []
                for row in resultados:
                    planilla = {
                        'id_planilla': row[0],
                        'nombre_planilla': row[1],
                        'descripcion_cargo_laboral': row[2],  # Nuevo campo para la descripción del cargo laboral
                        'periodo': row[3]
                        }
                    planillas.append(planilla)

                logger.debug("Planillas obtenidas correctamente")
                return JsonResponse({'status': 'success', 'planillas': planillas})

        except Exception as e:
            logger.error(f"Error en la vista obtener_codigos_planillas: {e}")
            return JsonResponse({'status': 'fail', 'error': str(e)}, status=500)

    logger.warning("Método de solicitud no válido")
    return JsonResponse({'status': 'fail', 'error': 'Invalid request method'}, status=400)



def obtener_detalles_planilla_v2(request):
    id_planilla = request.POST.get('id_planilla')
    modo_agrupacion = request.POST.get('modo_agrupacion')

    try:
        udcConn = connections['universal']
        with udcConn.cursor() as cursor:
            cursor.callproc("TH_GET_DETALLES_PLANILLAS_v2", [id_planilla, modo_agrupacion])
            column_names = [desc[0] for desc in cursor.description]
            detallesData = [dict(zip(map(str, column_names), row)) for row in cursor.fetchall()]
        
        # Cierra la conexión
        udcConn.close()
        
        # Devuelve los resultados como JSON
        return JsonResponse({'data': detallesData})
    except Exception as e:
        # Manejo de excepciones, puedes personalizar esto según tus necesidades
        return JsonResponse({'error': str(e)})




def obtener_detalles_planilla(request):
    if request.method == 'POST':
        id_planilla = request.POST.get('id_planilla')
        modo_agrupacion = request.POST.get('modo_agrupacion', 'AGRUPADO').upper()

        if not id_planilla:
            return JsonResponse({'status': 'fail', 'error': 'ID de planilla no proporcionado'}, status=400)

        if modo_agrupacion not in ['AGRUPADO', 'DETALLE']:
            return JsonResponse({'status': 'fail', 'error': 'Modo de agrupación inválido'}, status=400)

        try:
            with connections['universal'].cursor() as cursor:
                cursor.callproc('TH_GET_DETALLES_PLANILLAS_v2', [id_planilla, modo_agrupacion])
                resultados = cursor.fetchall()

            detalles_planilla = [
                {
                    
                    'id_deduccion': row[0],
                    'id_empleado': row[1],
                    'fecha_planilla': row[2],
                    'fecha_hora_creado': row[3],
                    'creado_por': row[4],
                    'fecha_hora_modificado': row[5],
                    'modificado_por': row[6],
                    'nombre_completo': row[7],
                    'identidad': row[8],
                    'salario_base': row[9],
                    'total_deducciones': row[10],
                    'total_bonificaciones': row[11],
                    'total_comisiones': row[12],
                    'estado': row[13],
                    'salario_bruto': row[30],
                    'salario_neto': row[31],
                    'salario_x_dia': row[16],
                    'dias_trabajados': row[17],
                    'Banco': row[18],
                    'cuenta_bancaria': row[19],
                    'tipo_planilla': row[20],
                    'aplica_salario_completo': row[21],
                    'nombre_planilla': row[22],
                    'salario_esperado_bruto': row[30],
                    'total_pagar': row[31],
                   
                }
                for row in resultados
            ]

            return JsonResponse({'status': 'success', 'detalles_planilla': detalles_planilla}, status=200)

        except connections['universal'].DatabaseError as db_error:
            logger.error(f"Error de base de datos: {str(db_error)}")
            return JsonResponse({'status': 'fail', 'error': 'Error en la base de datos.'}, status=500)

        except Exception as e:
            logger.error(f"Error en la vista obtener_detalles_planilla: {str(e)}")
            return JsonResponse({'status': 'fail', 'error': str(e)}, status=500)

    return JsonResponse({'status': 'fail', 'error': 'Método de solicitud no permitido'}, status=405)



def obtener_detalles_planilla_no_asociados(request):
    if request.method == 'POST':
        id_planilla = request.POST.get('id_planilla')
        if not id_planilla:
            return JsonResponse({'status': 'fail', 'error': 'ID de planilla no proporcionado'}, status=400)

        try:
            with connections['universal'].cursor() as cursor:
                cursor.callproc('TH_GET_DETALLES_PLANILLAS_SIN_ASOCIAR', [id_planilla])
                resultados = cursor.fetchall()

                detalles_planilla_no_asociados = []
                for row in resultados:
                    detalle = {
                        'id_deduccion': row[0],
                        'id_planilla': row[1],
                        'id_empleado': row[2],
                        'fecha_planilla': row[3],
                        'fecha_hora_creado': row[4],
                        'creado_por': row[5],
                        'nombre_completo': row[6],
                        'identidad': row[7],
                        'monto': row[8],
                        'tipo': row[9],
                        'estado': row[10]
                    }
                    detalles_planilla_no_asociados.append(detalle)

            return JsonResponse({'status': 'success', 'detalles_planilla_no_asociados': detalles_planilla_no_asociados})

        except Exception as e:
            logger.error(f"Error en la vista obtener_detalles_planilla_no_asociados: {e}")
            return JsonResponse({'status': 'fail', 'error': str(e)}, status=500)

    return JsonResponse({'status': 'fail', 'error': 'Invalid request method'}, status=400)

def obtener_deducciones_por_empleado(request):
    if request.method == 'POST':
        id_empleado = request.POST.get('id_empleado')  # Obtener el ID del empleado del cuerpo de la solicitud
        id_planilla = request.POST.get('id_planilla') 
        try:
            with connections['universal'].cursor() as cursor:
                # Llamar al procedimiento almacenado
                cursor.callproc('TH_OBTENER_DETALLES_DEDUCCION_PLANILLA_POR_EMPLEADO', [id_empleado, id_planilla])
                
                # Obtener los resultados
                resultados = cursor.fetchall()
                
                # Nombres de las columnas para mapear los resultados
                columnas = [col[0] for col in cursor.description]
                
                # Convertir resultados a una lista de diccionarios
                deducciones = [
                    dict(zip(columnas, fila))
                    for fila in resultados
                ]
            
            return JsonResponse({'status': 'success', 'deducciones': deducciones}, safe=False)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)


def obtener_bonificaciones_por_empleado(request):
    if request.method == 'POST':
        id_empleado = request.POST.get('id_empleado')  # Obtener el ID del empleado del cuerpo de la solicitud
        id_planilla = request.POST.get('id_planilla')  # Obtener el ID de la planilla del cuerpo de la solicitud

        if not id_empleado or not id_planilla:
            return JsonResponse({'status': 'error', 'message': 'ID de empleado o planilla no proporcionado'}, status=400)

        try:
            with connections['universal'].cursor() as cursor:
                # Llamar al procedimiento almacenado con los dos parámetros
                cursor.callproc('TH_OBTENER_DETALLES_BONIFICACION_PLANILLA_POR_EMPLEADO', [id_empleado, id_planilla])
                
                # Obtener los resultados
                resultados = cursor.fetchall()
                
                # Nombres de las columnas para mapear los resultados
                columnas = [col[0] for col in cursor.description]
                
                # Convertir resultados a una lista de diccionarios
                bonificaciones = [
                    dict(zip(columnas, fila))
                    for fila in resultados
                ]
            
            return JsonResponse({'status': 'success', 'bonificaciones': bonificaciones}, safe=False)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)


def insertar_deduccion_manual(request):
    if request.method == 'POST':
        try:
            # Obtener los datos del formulario desde request.POST
            p_id_planilla = request.POST.get('cod_planilla_manual')
            p_fecha_planilla = request.POST.get('fecha_planilla_manual')
            p_id_empleado = request.POST.get('id_empleado')  
            p_identidad = request.POST.get('identidad_empleado')
            p_nombre_completo = request.POST.get('nombre_completo')
            p_id_tipo = request.POST.get('deduccion_planilla_manual')
            p_valor = request.POST.get('valor_deduccion')
            p_creado_por = request.POST.get('creado_por')
            tipo_archivo = request.POST.get('tipoArchivo')

            # Determinar el tipo de registro basado en tipoArchivo
            if tipo_archivo == '1':
                p_tipo_registro = 'DEDUCCION'
            elif tipo_archivo == '2':
                p_tipo_registro = 'BONIFICACION'
            elif tipo_archivo == '3':
                p_tipo_registro = 'COMISION'
            else:
                return JsonResponse({'status': 'fail', 'error': 'Tipo de archivo no válido'}, status=400)

            # Verificar campos faltantes
            missing_fields = {
                'cod_planilla_manual': p_id_planilla,
                'fecha_planilla_manual': p_fecha_planilla,
                'id_empleado': p_id_empleado,  # Usar id_empleado aquí
                'identidad_empleado': p_identidad,
                'nombre_completo': p_nombre_completo,
                'deduccion_planilla_manual': p_id_tipo,
                'valor_deduccion': p_valor,
                'creado_por': p_creado_por
            }

            campos_vacios = [campo for campo, valor in missing_fields.items() if not valor]

            if campos_vacios:
                return JsonResponse(
                    {
                        'status': 'fail',
                        'error': 'Todos los campos son requeridos',
                        'missing_fields': campos_vacios
                    },
                    status=400
                )

            with connections['universal'].cursor() as cursor:
                # Llamar al procedimiento almacenado
                cursor.callproc('TH_INSERTAR_DEDUCCION_MANUAL', [
                    p_id_planilla,
                    p_fecha_planilla,
                    p_id_empleado,
                    p_identidad,
                    p_nombre_completo,
                    p_id_tipo,
                    p_valor,
                    p_creado_por,
                    p_tipo_registro
                ])

            return JsonResponse({'status': 'success', 'message': 'Registro insertado correctamente'})

        except Exception as e:
            return JsonResponse({'status': 'fail', 'error': str(e)}, status=500)

    return JsonResponse({'status': 'fail', 'error': 'Método de solicitud no permitido'}, status=405)

def insertar_bonificacion_manual(request):
    if request.method == 'POST':
        try:
            # Obtener los datos del formulario desde request.POST
            p_id_planilla = request.POST.get('cod_planilla_bonificacion_manual')
            p_fecha_planilla = request.POST.get('fecha_planilla_bonificacion_manual')
            p_id_empleado = request.POST.get('empleado_planilla_bonificacion_manual')
            p_id_tipo = request.POST.get('deduccion_planilla_bonificacion_manual')
            p_valor = request.POST.get('sbruto_empleado_bonificacion_planilla')
            p_creado_por = request.POST.get('creado_por')
            p_tipo_registro = 'BONIFICACION'

            # Log de los valores obtenidos
            logger.debug(f"p_id_planilla: {p_id_planilla}")
            logger.debug(f"p_fecha_planilla: {p_fecha_planilla}")
            logger.debug(f"p_id_empleado: {p_id_empleado}")
            logger.debug(f"p_id_tipo: {p_id_tipo}")
            logger.debug(f"p_valor: {p_valor}")
            logger.debug(f"p_creado_por: {p_creado_por}")

            # Validar que todos los parámetros sean proporcionados
            if not all([p_id_planilla, p_fecha_planilla, p_id_empleado, p_id_tipo, p_valor, p_creado_por]):
                missing_fields = {
                    'p_id_planilla': p_id_planilla,
                    'p_fecha_planilla': p_fecha_planilla,
                    'p_id_empleado': p_id_empleado,
                    'p_id_tipo': p_id_tipo,
                    'p_valor': p_valor,
                    'p_creado_por': p_creado_por
                }
                logger.error(f"Campos faltantes o vacíos: {missing_fields}")
                return JsonResponse({'status': 'fail', 'error': 'Todos los campos son requeridos', 'missing_fields': missing_fields}, status=400)

            with connections['universal'].cursor() as cursor:
                # Llamar al procedimiento almacenado
                cursor.callproc('TH_INSERTAR_DEDUCCION_MANUAL', [
                    p_id_planilla,
                    p_fecha_planilla,
                    p_id_empleado,
                    p_id_tipo,
                    p_valor,
                    p_creado_por,
                    p_tipo_registro
                ])

            return JsonResponse({'status': 'success', 'message': 'Registro insertado correctamente'})
        
        except Exception as e:
            logger.exception("Error al insertar la bonificación manual")
            return JsonResponse({'status': 'fail', 'error': str(e)}, status=500)

    return JsonResponse({'status': 'fail', 'error': 'Método de solicitud no permitido'}, status=405)

def mostrar_tipos_bonificaciones(request):
    try:
        with connections['universal'].cursor() as cursor:
            # Llamar al procedimiento almacenado
            cursor.callproc('TH_MOSTRAR_TIPOS_BONIFICACIONES')
            results = cursor.fetchall()

        # Formatear los resultados en una lista de diccionarios
        bonificaciones = [
            {
                'id_bonificacion': row[0],
                'id_categoria': row[1],
                'nombre_bonificacion': row[2],  
                'descripcion_tipo': row[3],    
                'automatizacion': row[4],
                'forma': row[5],
                'fecha_creacion': row[6],
                'creado_por': row[7],
                'fecha_modificacion': row[8],
                'modificado_por': row[9] 
            }
            for row in results
        ]

        # Devolver los resultados en formato JSON
        return JsonResponse({'bonificaciones': bonificaciones})

    except Exception as e:
        error_traceback = traceback.format_exc()
        print(f"Error al ejecutar la vista: {str(e)}")
        print(error_traceback)

        # Devolver el error en formato JSON
        return JsonResponse({'error': str(e), 'traceback': error_traceback}, status=500)


def aplicar_planilla(request):
    if request.method == 'POST':
        try:
            id_planilla = request.POST.get('id_planilla')
            
            if not id_planilla:
                return JsonResponse({'success': False, 'message': 'ID de planilla no proporcionado.'})

            # Llamar al procedimiento almacenado para aplicar la planilla
            with connections['universal'].cursor() as cursor:
                cursor.callproc('TH_APLICAR_PLANILLA', [id_planilla])

            return JsonResponse({'success': True, 'message': 'Planilla aplicada con éxito.'})

        except Exception as e:
            error_message = str(e)
            # Verificar si el error es el mensaje del procedimiento almacenado
            if 'Para aplicar la planilla, es necesario que su estado sea "En proceso"' in error_message:
                error_message = 'Para aplicar la planilla, asegúrate de que esté en estado "En proceso".'
            elif 'La planilla especificada no existe o ya está aplicada' in error_message:
                error_message = 'La planilla seleccionada no existe o ya ha sido aplicada.'

            return JsonResponse({'success': False, 'message': error_message}, status=400)

    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)




def aplicar_nueva_fecha_por_aplicar_planilla(request):
    try:
        id_planilla = request.POST.get('id_planilla')

        userName = request.session.get('userName', '')

        udcConn = connections['universal']
        with udcConn.cursor() as cursor:
            cursor.callproc('TH_UPDATE_NEW_FECHA_APLICAR_PLANILLA', [id_planilla, userName])

        udcConn.close()

        datos = {'save': 1}
    except Exception as e:
        datos = {'save': 0, 'error': str(e)}
    
    return JsonResponse(datos)


def set_new_fecha_aplicar_planilla(request):
    try:
        id_planilla = request.POST.get('id_planilla')
        fechaNuevaAplicarPlanilla = request.POST.get('fechaNuevaAplicarPlanilla')

        userName = request.session.get('userName', '')

        udcConn = connections['universal']
        with udcConn.cursor() as cursor:
            cursor.callproc('TH_SET_NEW_FECHA_APLICAR_PLANILLA', [id_planilla, fechaNuevaAplicarPlanilla, userName])

        udcConn.close()

        datos = {'save': 1}
    except Exception as e:
        datos = {'save': 0, 'error': str(e)}
    
    return JsonResponse(datos)



def aplica_salario_completo(request):
    try:
        id_empleado = request.POST.get('id_empleado')
        id_planilla = request.POST.get('id_planilla')
        aplica_salario = request.POST.get('aplica_salario')

        userName = request.session.get('userName', '')

        udcConn = connections['universal']
        with udcConn.cursor() as cursor:
            cursor.callproc('TH_INSERT_APLICA_SALARIO_COMPLETO_X_PLANILLA', [id_empleado, id_planilla, aplica_salario, userName])

        udcConn.close()

        datos = {'save': 1}
    except Exception as e:
        datos = {'save': 0, 'error': str(e)}
    
    return JsonResponse(datos)

@csrf_exempt
def desaplicar_planilla(request):
    try:
        id_planilla = request.POST.get('id_planilla')

        userName = request.session.get('userName', '')

        udcConn = connections['universal']
        with udcConn.cursor() as cursor:
            cursor.callproc('TH_DESAPLICAR_PLANILLA', [id_planilla, userName])

        udcConn.close()

        datos = {'save': 1}
    except Exception as e:
        datos = {'save': 0, 'error': str(e)}
    
    return JsonResponse(datos)



@csrf_exempt
def registrar_saldos_empleados_planillas(request):
    try:
        # Leer los datos del cuerpo de la solicitud
        body = json.loads(request.body)
        id_planilla = body.get('id_planilla')  # Planilla enviada desde AJAX
        payload = body.get('data', [])  # Arreglo de datos

        userName = request.session.get('userName', '')

        # Recorremos la lista para ejecutar el procedimiento almacenado
        for item in payload:
            id_empleado = item.get("id_empleado")
            salario_base = item.get("salario_base")
            salario_pagado = item.get("salario_pagado")
            saldo = item.get("saldo")

            # Llamar al procedimiento almacenado
            udcConn = connections['universal']
            with udcConn.cursor() as cursor:
                cursor.callproc(
                    'TH_INSERT_SALDO_EMPLEADOS_SALARIOS_PLANILLAS',
                    [id_empleado, id_planilla, salario_base, salario_pagado, saldo, userName]
                )
            udcConn.commit()

        # Aplicar la planilla
        with connections['universal'].cursor() as cursor:
            cursor.callproc('TH_APLICAR_PLANILLA', [id_planilla])

        datos = {'save': 1}
    except Exception as e:
        datos = {'save': 0, 'error': str(e)}
    
    return JsonResponse(datos)



def eliminar_deduccion(request):
    try:
        # Obtener los datos enviados por AJAX
        arr_empleado_detalle = []
        for key in request.POST.keys():
            if key.startswith('arrEmpleadoDetalle'):
                arr_empleado_detalle.append(request.POST.get(key))
        
        id_planilla = request.POST.get('id_planilla')
        agrupacion = request.POST.get('agrupacion')
        
        userName = request.session.get('userName', '')

        # Recorremos la lista para ejecutar el procedimiento almacenado
        for id_detalle_x_empleado in arr_empleado_detalle:
            print(id_detalle_x_empleado)
            
            with connections['universal'].cursor() as cursor:
                cursor.callproc('TH_ELIMINAR_DETALLE_PLANILLA', [id_detalle_x_empleado, id_planilla, agrupacion, userName])

        datos = {'save':1}
    except Exception as e:
        datos = {'save': 0, 'error': str(e)}
    
    return JsonResponse(datos)


def eliminar_deduccion_x_empleado(request):
    if request.method == 'POST':
        try:
            id_deduccion = request.POST.get('id_deduccion')
            id_planilla = request.POST.get('id_planilla')
            modificado_por = request.POST.get('modificado_por')


            with connections['universal'].cursor() as cursor:
                cursor.callproc('TH_ELIMINAR_DETALLE_X_EMPLEADO', [id_deduccion, id_planilla, modificado_por])

            return JsonResponse({'status': 'success', 'message': 'Detalle eliminado correctamente'})

        except Exception as e:
            return JsonResponse({'status': 'fail', 'error': str(e)}, status=500)

    return JsonResponse({'status': 'fail', 'error': 'Método no permitido'}, status=405)

def eliminar_empleados_no_asociados(request):
    if request.method == 'POST':
        try:
            id_deduccion = request.POST.get('id_deduccion')
            modificado_por = request.POST.get('modificado_por')

            with connections['universal'].cursor() as cursor:
                cursor.callproc('TH_ELIMINAR_DETALLE_NO_ASOCIADOS', [id_deduccion, modificado_por])

            return JsonResponse({'status': 'success', 'message': 'Detalle eliminado correctamente'})

        except Exception as e:
            return JsonResponse({'status': 'fail', 'error': str(e)}, status=500)

    return JsonResponse({'status': 'fail', 'error': 'Método no permitido'}, status=405)



def editar_deduccion(request):
    if request.method == 'POST':
        try:
            # Obtén los datos del formulario
            id_deduccion = request.POST.get('id_deduccion')
            id_tipo = request.POST.get('id_tipo')
            valor = request.POST.get('valor')
            modificado_por = request.POST.get('modificado_por')
            tipo_registro = request.POST.get('tipoArchivo')

            # Valida los datos recibidos
            if not (id_deduccion and id_tipo and valor and tipo_registro):
                return JsonResponse({'status': 'error', 'message': 'Faltan datos requeridos.'})

            # Conecta y ejecuta el procedimiento almacenado
            with connections['universal'].cursor() as cursor:
                cursor.callproc('TH_EDITAR_DEDUCCION_PLANILLA', [id_deduccion, id_tipo, valor, modificado_por, tipo_registro])

            # Responde exitosamente
            return JsonResponse({'status': 'success', 'message': 'Deducción actualizada correctamente.'})

        except Exception as e:
            # Maneja errores
            return JsonResponse({'status': 'error', 'message': str(e)})

    # Si el método no es POST, retorna un error
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'})



def mostrar_tipos_deduccion_banco(request):
    # Obtener el parámetro `id_categoria` desde la URL
    id_categoria = request.GET.get('id_categoria')
    logger.debug("ID Categoría recibido:", id_categoria)  # Verificar el valor recibido

    if id_categoria is None:
        return JsonResponse({"error": "El parámetro id_categoria es necesario"}, status=400)
    try:
        with connections['universal'].cursor() as cursor:
            logger.debug(f"Llamando al procedimiento almacenado con id_categoria = {id_categoria}")
            cursor.execute("CALL TH_MOSTRAR_TIPOS_DEDUCCION_BANCO(%s)", [id_categoria])
            resultados = cursor.fetchall()
            logger.debug(f"Resultados obtenidos del procedimiento: {resultados}")

        deducciones = [{"id_tipo_deduccion": fila[0], "nombre_tipo": fila[1]} for fila in resultados]
        return JsonResponse({"deducciones": deducciones}, status=200)

    except Exception as e:
        logger.error(f"Error al ejecutar el procedimiento almacenado: {e}")
        return JsonResponse({"error": str(e)}, status=500)

   
def eliminar_tipo_deduccion(request):
    if request.method == 'POST':
        try:
            # Obtener los datos del POST
            id_deduccion = request.POST.get('id_deduccion')
            modificado_por = request.POST.get('modificado_por')

            # Verificar que ambos campos estén presentes
            if not all([id_deduccion, modificado_por]):
                return JsonResponse({'status': 'fail', 'error': 'Faltan campos requeridos'}, status=400)

            # Llamar al procedimiento almacenado para cambiar el estado a 3
            with connections['universal'].cursor() as cursor:
                cursor.callproc('TH_ELIMINAR_TIPO_DEDUCCION', [
                    id_deduccion,
                    modificado_por
                ])

            return JsonResponse({'status': 'success', 'message': 'Deducción eliminada correctamente'})

        except Exception as e:
            return JsonResponse({'status': 'fail', 'error': str(e)})

    return JsonResponse({'status': 'fail', 'error': 'Método no permitido'}, status=405)


def eliminar_tipo_bonificacion(request):
    if request.method == 'POST':
        try:
            # Obtener los datos del POST
            id_bonificacion = request.POST.get('id_bonificacion')
            modificado_por = request.POST.get('modificado_por')

            # Verificar que ambos campos estén presentes
            if not all([id_bonificacion, modificado_por]):
                return JsonResponse({'status': 'fail', 'error': 'Faltan campos requeridos'}, status=400)

            # Llamar al procedimiento almacenado para cambiar el estado a 3
            with connections['universal'].cursor() as cursor:
                cursor.callproc('TH_ELIMINAR_TIPO_BONIFICACION', [
                    id_bonificacion,
                    modificado_por
                ])

            return JsonResponse({'status': 'success', 'message': 'Bonificación eliminada correctamente'})

        except Exception as e:
            error_traceback = traceback.format_exc()
            print(f"Error al ejecutar la vista: {str(e)}")
            print(error_traceback)
            return JsonResponse({'status': 'fail', 'error': str(e), 'traceback': error_traceback}, status=500)

    return JsonResponse({'status': 'fail', 'error': 'Método no permitido'}, status=405)



def insertar_categoria_deduccion(request):
    if request.method == 'POST':
        nombre_categoria = request.POST.get('nombre_categoria')
        creado_por = request.POST.get('creado_por')

        try:
            with connections['universal'].cursor() as cursor:
                # Llama al procedimiento almacenado e intenta insertar la categoría
                cursor.callproc('TH_INSERTAR_CATEGORIAS_DEDUCCIONES', [nombre_categoria, creado_por])
                last_id = cursor.fetchone()[0]  # Captura el resultado de SELECT LAST_INSERT_ID()
            
            # Devuelve éxito y el último ID insertado
            return JsonResponse({'status': 'success', 'last_id': last_id})
        
        except Exception as e:
            # Verifica si el error es por duplicación
            if 'La categoría ya existe' in str(e):
                return JsonResponse({'status': 'fail', 'error': 'El nombre de la categoría ya existe. Por favor, elige otro nombre.'}, status=400)
            else:
                # Otro tipo de error
                return JsonResponse({'status': 'fail', 'error': 'Hubo un problema al realizar la solicitud.'}, status=500)

    return JsonResponse({'status': 'fail', 'error': 'Método no permitido'}, status=405)


def insertar_bonificacion_tipo(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            descripcion_bonificacion = data.get('descripcion_bonificacion')
            creado_por = data.get('creado_por')

            if not descripcion_bonificacion or not creado_por:
                return JsonResponse({'status': 'fail', 'error': 'Faltan campos requeridos'}, status=400)

            with connections['universal'].cursor() as cursor:
                cursor.callproc('TH_INSERTAR_BONIFICACION_TIPO', [
                    descripcion_bonificacion,
                    creado_por,
                    0  # Valor temporal para el parámetro de salida
                ])
                cursor.execute("SELECT LAST_INSERT_ID()")
                last_id = cursor.fetchone()[0]

            return JsonResponse({'status': 'success', 'message': 'Bonificación creada exitosamente', 'last_id': last_id})

        except Exception as e:
            return JsonResponse({'status': 'fail', 'error': str(e)})

    return JsonResponse({'status': 'fail', 'error': 'Método no permitido'}, status=405)


def insertar_tipo_deduccion_banco(request):
    if request.method == 'POST':
        try:
            # Obtener los datos del cuerpo de la solicitud
            body_unicode = request.body.decode('utf-8')
            data = json.loads(body_unicode)
            
            # Extraer los parámetros necesarios
            p_id_categoria = data.get('id_categoria')
            p_nombre_tipo = data.get('nombre_tipo')
            p_creado_por = data.get('creado_por')

            # Validación de campos requeridos
            if not p_id_categoria or not p_nombre_tipo or not p_creado_por:
                return JsonResponse({'status': 'fail', 'error': 'Faltan campos requeridos'}, status=400)
            
            with connections['universal'].cursor() as cursor:
                # Llamar al procedimiento almacenado con los parámetros necesarios
                cursor.callproc('TH_INSERTAR_TIPO_DEDUCCION_BANCO', [
                    p_id_categoria,
                    p_nombre_tipo,
                    p_creado_por
                ])
                
                # Obtener el último ID insertado
                last_id = cursor.fetchone()[0]
            
            # Respuesta de éxito con el último ID
            return JsonResponse({'status': 'success', 'last_id': last_id})

        except Exception as e:
            # En caso de error, devolver el mensaje de error
            return JsonResponse({'status': 'fail', 'error': str(e)}, status=500)

    return JsonResponse({'status': 'fail', 'error': 'Método no permitido'}, status=405)



def obtener_tipos_archivo(request):
    if request.method == 'GET':
        try:
            with connections['universal'].cursor() as cursor:
                cursor.callproc('TH_GET_TIPOS_ARCHIVO')  # Llamar al procedimiento almacenado
                results = cursor.fetchall()  # Obtener los resultados

            # Construir la respuesta en formato JSON
            data = [
                {
                    'id': row[0],
                    'nombre': row[1]
                }
                for row in results
            ]
            
            return JsonResponse({'status': 'success', 'tipos_archivo': data}, safe=False)
        
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'fail', 'message': 'Método no permitido'}, status=405)

def actualizar_detalle_planilla_no_asociada(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            p_id_deduccion = data.get('deducciones')[0].get('id_deduccion')  # Tomar id_deduccion de las deducciones seleccionadas
            p_id_empleado = data.get('empleado_id')
            p_modificado_por = data.get('modificado_por')
            p_tipo_registro = data.get('tipo_registro')

            missing_fields = []
            if not p_id_deduccion:
                missing_fields.append('id_deduccion')
            if not p_id_empleado:
                missing_fields.append('id_empleado')
            if not p_modificado_por:
                missing_fields.append('modificado_por')

            if missing_fields:
                return JsonResponse({'error': f'Faltan los siguientes campos: {", ".join(missing_fields)}'}, status=400)

          
            with connections['universal'].cursor() as cursor:
                for deduccion in data.get('deducciones'):
                    cursor.callproc('TH_ACTUALIZAR_DETALLE_PLANILLA_NO_ASOCIADA', [
                        int(deduccion['id_deduccion']),
                        int(p_id_empleado),
                        p_modificado_por,
                        p_tipo_registro
                    ])
            return JsonResponse({'success': 'Detalle de planilla actualizado correctamente.'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Método no permitido.'}, status=405)


def get_detalles_x_empleado(request):
    try:
        id_planilla = request.POST.get('id_planilla')
        id_empleado = request.POST.get('id_empleado')

        if not id_planilla or not id_empleado:
            return JsonResponse({'status': 'error', 'message': 'Parámetros faltantes.'}, status=400)

        with connections['universal'].cursor() as cursor:
            cursor.callproc('TH_GET_DETALLES_X_EMPLEADO', [id_planilla, id_empleado])
            resultados = cursor.fetchall()
            columnas = [col[0] for col in cursor.description]

        data = [dict(zip(columnas, fila)) for fila in resultados]
        return JsonResponse({'status': 'success', 'data': data})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    

def actualizar_automatico_manual(request):
    try:
        # Llamada al procedimiento almacenado
        with connections['universal'].cursor() as cursor:
            cursor.callproc('TH_VERIFICAR_ACTUALIZAR_AUTOMATICO_MANUAL')
        
        return JsonResponse({'success': True, 'message': 'Procedimiento ejecutado correctamente.'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})



#Bonificaciones 
def nueva_bonificacion(request):
    id = request.session.get('user_id', '')

    if id == '':
        return HttpResponseRedirect(reverse('login'))
    else:
        return render(request, 'planilla/bonificaciones.html')


#Deducciones
def nueva_deduccion(request):
    id = request.session.get('user_id', '')

    if id == '':
        return HttpResponseRedirect(reverse('login'))
    else:
        return render(request, 'planilla/deducciones.html')
    
def nueva_comision(request):
    id = request.session.get('user_id', '')

    if id == '':
        return HttpResponseRedirect(reverse('login'))
    else:
        return render(request, 'planilla/comisiones.html')

def reportes(request):
    id = request.session.get('user_id', '')

    if id == '':
        return HttpResponseRedirect(reverse('login'))
    else:
        # Obtener los reportes de planillas
        planilla_reporte = obtener_planillas_reportes()
      

        # Pasar `planilla_reporte` al contexto
        context = {
            'planilla_reporte': planilla_reporte,
            'empleados_empresa':obtener_empleado_empresas()
        }

        return render(request, 'planilla/reportes.html', context)



def obtener_bancos():
    with connections['universal'].cursor() as cursor:
        cursor.callproc('TH_GET_BANCOS')  
        results = cursor.fetchall()

    return results

def obtener_planillas_reportes():
    with connections['universal'].cursor() as cursor:
        cursor.callproc('TH_GET_ALL_PLANILLAS')
        results = cursor.fetchall()

    return results

def obtener_empleado_empresas():
    with connections['universal'].cursor() as cursor:
        cursor.callproc('TH_OBTENER_NOMBRE_EMPLEADO_EMPRESAS')
        results = cursor.fetchall()

    return results



def comprobante_bancos(request):
    try:
        id_planilla = request.POST.get('id_planilla')
        tipo_registro = 'AGRUPADO'
        # Conexión y llamada al procedimiento
        with connections['universal'].cursor() as cursor:
            cursor.callproc("TH_GET_DETALLES_PLANILLAS_v2", [id_planilla, 1])
            column_names = [desc[0] for desc in cursor.description]
            crudComprobanteBancos = [dict(zip(column_names, row)) for row in cursor.fetchall()]
        
        return JsonResponse({'comprobante': crudComprobanteBancos}, safe=False)

    except Exception as e:
        print(f"Error al mostrar comprobantes de bancos: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)



def reporte_deducciones(request):
    try:
        # Obtener los valores enviados desde el cliente
        id_planilla = request.POST.get('id_planilla')
        tipo_registro = 'AGRUPADO'
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_final = request.POST.get('fecha_final')
        filtro = request.POST.get('filtro')
        id_empleado = request.POST.get('id_empleado', None)  # Por defecto, None si no está presente

        # Validar parámetros obligatorios
        if not id_planilla or not fecha_inicio or not fecha_final or not filtro:
            return JsonResponse({'error': 'Faltan datos obligatorios'}, status=400)

        # Si el filtro es EMPLEADO, id_empleado es obligatorio
        if filtro == 'EMPLEADO' and not id_empleado:
            return JsonResponse({'error': 'Debe especificar un empleado si el filtro es EMPLEADO'}, status=400)

        # Si el filtro no es EMPLEADO, establece id_empleado como NULL
        if filtro != 'EMPLEADO':
            id_empleado = None

        # Llamar al procedimiento almacenado
        with connections['universal'].cursor() as cursor:
            cursor.callproc("TH_GET_REPORTE_PLANILLA_v3", [
                id_planilla,
                tipo_registro,
                fecha_inicio,
                fecha_final,
                filtro,
                id_empleado
            ])
            column_names = [desc[0] for desc in cursor.description]
            crudReporteDeducciones = [dict(zip(column_names, row)) for row in cursor.fetchall()]

        # Devolver los datos en formato JSON
        return JsonResponse({'reporte': crudReporteDeducciones}, safe=False)

    except Exception as e:
        # Manejar errores y devolver una respuesta con el error
        print(f"Error al mostrar reporte de deducciones: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

